import base64
from operator import eq
from pyexpat import features
import re
from io import BytesIO
from typing import List
from wsgiref.util import FileWrapper
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework.views import APIView
from django.http import Http404, request
from rest_framework.response import Response
from .decorators import jwt_required
from django.http.response import HttpResponse
from django.http import FileResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from rest_framework import status
import mimetypes
import pytz
from datetime import datetime, timedelta, date, time as dt_time
# import cups
import os, tempfile
import pickle
import glob
import pathlib
import random

from pyodbc import Cursor, Error, connect, lowercase
from django.http.response import JsonResponse
from rest_framework.decorators import api_view, authentication_classes, permission_classes, renderer_classes
from django.db import connections, transaction
from support.database import encloseColumn, Filters, DBSql, TypeDml, fetchall, Check
from support.myUtils import  ifNull

from rest_framework.renderers import JSONRenderer, MultiPartRenderer, BaseRenderer
from rest_framework.utils import encoders, json
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
import collections
import hmac
import hashlib
import math
from django.core.files.storage import FileSystemStorage
from sistema.settings.appSettings import AppSettings
import time
import requests
import psycopg2
from api.exports import export
import face_recognition
from PIL import Image, ImageEnhance, ImageOps, ImageFilter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import traceback
import threading
import shutil
import uuid
import json

connGatewayName = "postgres"
connMssqlName = "sqlserver"
connSage100cName = "sage100c"
dbsage100c = DBSql(connections[connSage100cName].alias)
dbgw = DBSql(connections[connGatewayName].alias)
db = DBSql(connections["default"].alias)
dbmssql = DBSql(connections[connMssqlName].alias)

fotos_base_path = '../fotos'
records_base_path = '../records'
records_invalid_base_path = '../records_invalid'
faces_base_path = 'faces'
cropped_faces_base_path = 'cropped_faces'
tolerance = 0.45
jitters = 1
model = 'large'
JUSTIFICACOES_BASE_PATH = "../justificacoes"
PAUSA_ALMOCO_MIN = 45   
PAUSA_ALMOCO_MAX = 75 
FERIAS_ESTADOS = {
'pendente':             {'label': 'Pendente', 'cor': 'orange'},
    'aprovado_chefe':   {'label': 'Aprovado pelo Chefe','cor': 'blue'},
    'rejeitado_chefe':  {'label': 'Rejeitado pelo Chefe','cor': 'red'},
    'aprovado_rh':      {'label': 'Aprovado por RH',    'cor': 'green'},
    'rejeitado_rh':     {'label': 'Rejeitado por RH',   'cor': 'red'},
}


def filterMulti(data, parameters, forceWhere=True, overrideWhere=False, encloseColumns=True, logicOperator="and"):
    p = {}
    txt = ''
    _forceWhere = forceWhere
    _overrideWhere = overrideWhere
    hasFilters = False
    for mainKey, mainValue in parameters.items():
        if (hasFilters):
            _forceWhere = False
            _overrideWhere = logicOperator
        if data.get(mainKey) is not None:
            sp = {}
            for key in mainValue.get('keys'):
                table = f'{mainValue.get("table")}.' if (mainValue.get("table") and encloseColumns) else mainValue.get("table", '')
                field = f'{table}"{key}"' if encloseColumns else f'{table}{key}'
                sp[key] = {"value": data.get(mainKey).lower(), "field": f'lower({field})'}
            f = Filters(data)
            f.setParameters(sp, True)
            f.where(_forceWhere, _overrideWhere)
            f.auto()
            f.value('or')
            p = {**p, **f.parameters}
            txt = f'{txt}{f.text}'
            if (not hasFilters):
                hasFilters = f.hasFilters
    return {"hasFilters": hasFilters, "text": txt, "parameters": p}

def rangeP(data, key, field, fieldDiff=None,pName=None):
    ret = {}
    if data is None:
        return ret
    if isinstance(key, list):
        hasNone = False
        for i, v in enumerate(data):
            if v is not None:
                ret[f'{pName}{key[i]}_{i}'] = {"key": key[i], "value": v, "field": field}
            else:
                hasNone = True
        if hasNone == False and len(data)==2 and fieldDiff is not None:
            ret[f'{pName}{key[0]}_{key[1]}'] = {"key": key, "value": ">=0", "field": fieldDiff}
    else:    
        for i, v in enumerate(data):
            if v is not None:
                ret[f'{pName}{key}_{i}'] = {"key": key, "value": v, "field": field}
    return ret

def rangeP2(data, key, field1, field2, fieldDiff=None):
    ret = {}
    field=False
    if data is None:
        return ret
    if isinstance(key, list):
        hasNone = False
        for i, v in enumerate(data):
            if v is not None:
                ret[f'{key[i]}_{i}'] = {"key": key[i], "value": v, "field": field1 if field is False else field2}
            else:
                hasNone = True
        if hasNone == False and len(data)==2 and fieldDiff is not None:
            ret[f'{key[0]}_{key[1]}'] = {"key": key, "value": ">=0", "field": fieldDiff}
    else:    
        for i, v in enumerate(data):
            if v is not None:
                ret[f'{key}_{i}'] = {"key": key, "value": v, "field": field1 if field is False else field2}
    return ret

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR',None)
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR',None)
        if not ip:
            ip = request.META.get('HTTP_X_REAL_IP', None)
    return ip


@api_view(['POST'])
@renderer_classes([JSONRenderer])
def Sql(request, format=None):
    ips_allowed = ["*","192.168.0.254"]
    #ip_address = request.META.get("HTTP_X_REAL_IP")
    if "*" not in ips_allowed and ip_address not in ips_allowed:
        return Response({"status": "error", "title": "Erro de acesso!"})
    if "parameters" in request.data and "method" in request.data["parameters"]:
        method=request.data["parameters"]["method"]
        func = globals()[method]
        response = func(request, format)
        return response
    return Response({})


@api_view(['POST'])
@renderer_classes([JSONRenderer])
@permission_classes([IsAuthenticated])
@jwt_required
def SqlProtected(request):
    if "parameters" in request.data and "method" in request.data["parameters"]:
        method=request.data["parameters"]["method"]
        func = globals()[method]
        response = func(request, format)
        return response
    return Response({})

@api_view(['GET'])
@renderer_classes([JSONRenderer])
def Sync(request, format=None):
    faces = loadFaces(faces_base_path,True)
    return Response({"status":"success","nums":faces.get("nums"),"matrix":faces.get("matrix")})

def EmployeesLookup(request, format=None):
    connection = connections[connMssqlName].cursor()
    f = Filters(request.data['filter'])
    f.setParameters({}, True)
    f.where()
    f.auto()
    f.value()

    fmulti = filterMulti(request.data['filter'], {
        'fmulti': {"keys": ['REFNUM_0', "FULLNAME"], "table": 'T.'}
    }, False, "and" if f.hasFilters else "where" ,False)
    parameters = {**f.parameters, **fmulti['parameters']}

    dql = dbmssql.dql(request.data, False)
    cols = f"""*"""
    dql.columns=encloseColumn(cols,False)
    dql.sort = " ORDER BY(SELECT NULL) " if not dql.sort else dql.sort #Obrigatório se PAGING em sqlserver
    sql = lambda p, c, s: (
        f"""  
            select * from (
            select DISTINCT e.REFNUM_0, NAM_0,SRN_0, CONCAT(SRN_0,' ',NAM_0) FULLNAME FROM x3peoplesql.PEOPLELTEK.EMPLOID e 
            JOIN x3peoplesql.PEOPLELTEK.EMPLOCTR c on c.REFNUM_0 = e.REFNUM_0 
            WHERE c.PROPRF_0 = 'STD' 
            ) T
            {f.text} {fmulti["text"]}
            {s(dql.sort)}
             {p(dql.paging)} {p(dql.limit)}
        """
    )
    if ("export" in request.data["parameters"]):
        dql.limit=f"""OFFSET 0 ROWS FETCH NEXT {request.data["parameters"]["limit"]} ROWS ONLY"""
        dql.paging=""
        return export(sql(lambda v:v,lambda v:v,lambda v:v), db_parameters=parameters, parameters=request.data["parameters"],conn_name=AppSettings.reportConn["sage"],dbi=dbmssql,conn=connection)
    try:
        response = dbmssql.executeList(sql, connection, parameters,[],None,f"""
            select * from (
            select DISTINCT e.REFNUM_0, NAM_0,SRN_0, CONCAT(SRN_0,' ',NAM_0) FULLNAME FROM x3peoplesql.PEOPLELTEK.EMPLOID e 
            JOIN x3peoplesql.PEOPLELTEK.EMPLOCTR c on c.REFNUM_0 = e.REFNUM_0 
            WHERE c.PROPRF_0 = 'STD' 
            ) T
            {f.text} {fmulti["text"]}
        """)
    except Exception as error:
        print(str(error))
        return Response({"status": "error", "title": str(error)})
    return Response(response)

#CHANGED
def loadFaces(path,sync=False):
    if os.path.isfile(os.path.join("faces.dictionary")) and sync==False:
        with open('faces.dictionary', 'rb') as faces_file:
            return pickle.load(faces_file)
    else:
        faces ={"nums": []}
        for filename in os.listdir(path):
            processedimage = preProcessImage(os.path.join(path, filename)).save(os.path.join(cropped_faces_base_path,filename),"JPEG")
            ki = face_recognition.load_image_file(os.path.join(cropped_faces_base_path,filename))
            #f = os.path.join(path, filename)
            #if os.path.isfile(f):
            #    ki = face_recognition.load_image_file(f)
            matrix=face_recognition.face_encodings(ki,None,jitters,model)
            if matrix and len(matrix)>0:
                faces.get("nums").append({"num":filename.split('_')[0],"t_stamp":datetime.today(),"file":filename,"matrix":matrix[0]})
        with open('faces.dictionary', 'wb') as faces_file:
            pickle.dump(faces, faces_file)
        return faces

def getBiggestFace(face_locations):
    max_size = 0
    max_location = None
    for location in face_locations:
        top, right, bottom, left = location
        size = (bottom - top) * (right - left)
        if size > max_size:
            max_size = size
            max_location = location
    return max_location

def preProcessImage(filepath,radius=None,brightness_factor=None):
    #f = os.path.join(faces_base_path, filename)
    if os.path.isfile(filepath):
        image = face_recognition.load_image_file(filepath)
        face_locations = face_recognition.face_locations(image)
        if face_locations and len(face_locations) > 0:
            top, right, bottom, left = getBiggestFace(face_locations)
            #top, right, bottom, left = face_locations[0]
            # Crop the face from the image
            face_image = Image.fromarray(image[top:bottom, left:right])
            gray_image = face_image.convert('L')
            equalized_image = ImageOps.equalize(gray_image)
            if radius is not None:
                equalized_image = equalized_image.filter(ImageFilter.GaussianBlur(radius))
            
            if (brightness_factor is None):
                return equalized_image.convert("RGB")
            else:
                #average_pixel = int(sum(list(blurred_image.getdata())) / len(list(blurred_image.getdata())))
                gamma_corrected_image = ImageEnhance.Brightness(equalized_image).enhance(1.5)
                return gamma_corrected_image.convert("RGB")
                
    

@api_view(['GET'])
@renderer_classes([JSONRenderer])
def SimulateRecordAdd(request, format=None):
    #for testes only to remove from here
    connection = connections[connMssqlName].cursor()
    num ="F00030"
    record = processRecord(num,datetime(2022, 1, 19, 23, 00,00)) #saveRecord("F00160",datetime(2023, 3, 3, 13, 44,00),None,{"type":"in","timestamp":datetime.today().strftime("%Y-%m-%d %H:%M:%S")})
    #reg = [{"id":242,"num":"F00030","dt":"2022-01-20","dts":"2022-01-20","nt":2,"ts_01":"2022-01-019 23:50:03","ss_01":"2022-01-019 23:50:03","ty_01":"in","ts_02":None,"ss_02":None,"ty_02":None,"ts_03":None,"ss_03":None,"ty_03":None,"ts_04":None,"ss_04":None,"ty_04":None,"ts_05":None,"ss_05":None,"ty_05":None,"ts_06":None,"ss_06":None,"ty_06":None,"ts_07":None,"ss_07":None,"ty_07":None,"ts_08":None,"ss_08":None,"ty_08":None,"status":1}]
    #n_records = 8
    #for n in reversed(range(8)):
    #    if reg[0].get(f"ss_{str(n+1).zfill(2)}") is None:
    #        n_records = n_records -1
    print(record)
    # if record.get("date_ref") is not None:
    #     f = Filters({"num": num,"dts": record.get("date_ref").strftime("%Y-%m-%d") })
    #     f.where()
    #     f.add(f'num = :num', True)
    #     f.add(f'dts = :dts', True)
    #     f.value("and")
    #     reg = dbmssql.executeSimpleList(lambda: (f'SELECT * from rponto.dbo.time_registration {f.text}'), connection, f.parameters)['rows']
        
    return Response({"record":record})

def processRecord(num, ts):
    connection = connections[connSage100cName].cursor()
    sql = f"""
        SELECT F1.NFUNC, F1.NOME, F1.DEPARTAMENTO, F1.CALENDARIO
        FROM TRIMTEK_1GEP.dbo.FUNC1 F1
        WHERE F1.NFUNC = '{num}' AND F1.DEMITIDO = 0
    """
    print(f"[DEBUG processRecord] Procurando num='{num}' (string interpolation) em FUNC1")
    try:
        connection.execute(sql)
        row = connection.fetchone()
        print("[DEBUG processRecord] Row:", row)
        if row:
            nfunc, nome, departamento, calendario = row
            return {
                "date_ref": ts.strftime("%Y-%m-%d"),
                "nfunc": nfunc,
                "nome": nome,
                "dep": departamento,
                "tp_hor": calendario
            }
        else:
            print(f"[DEBUG processRecord] NÃO ENCONTROU {num} em FUNC1!")
            return {
                "date_ref": ts.strftime("%Y-%m-%d"),
                "nfunc": None,
                "dep": None,
                "tp_hor": None
            }
    except Exception as error:
        print("[ERROR processRecord]:", error)
        return {
            "date_ref": ts.strftime("%Y-%m-%d"),
            "nfunc": None,
            "dep": None,
            "tp_hor": None
        }
    finally:
        connection.close()


def saveRecord(num, ts, hsh, data, ip):
    pln = processRecord(num, ts)
    connection = connections[connMssqlName].cursor()
    dep = pln.get("dep")
    tp_hor = pln.get("tp_hor")

    if hsh is None:
        f = Filters({"num": num, "dts": pln.get("date_ref")})
        f.where()
        f.add(f'num = :num', True)
        f.add(f'dts = :dts', True)
        f.value("and")
        reg = dbmssql.executeSimpleList(lambda: (f'SELECT * from rponto.dbo.time_registration {f.text}'), connection, f.parameters)['rows']
        if len(reg) == 0:
            dti = {
                "num": f.parameters["num"],
                "nt": 1,
                "hsh": hashlib.md5(f"""{f.parameters["num"]}-{ts.strftime("%Y-%m-%d")}""".encode('utf-8')).hexdigest(),
                "dts": pln.get("date_ref"),
                "dt": pln.get("date_ref"),
                f"ss_01": ts.strftime("%Y-%m-%d %H:%M:%S"),
                f"ts_01": data["timestamp"],
                f"ty_01": "in",
                f"auto_01": 1 if data.get("auto") else 0,
                f"source_01": ip,
                "dep": dep,
                "tp_hor": tp_hor
            }
            dml = dbmssql.dml(TypeDml.INSERT, dti, "rponto.dbo.time_registration", None, None, False)
            dbmssql.execute(dml.statement, connection, dml.parameters)
            return {"status": "success", "hsh": dti.get("hsh")}
        else:
            nt = reg[0].get("nt")
            if nt == 8:
                raise Exception("Atingiu o número máximo de registos! Por favor entre em contacto com os Recursos Humanos.")
            dti = {
                "nt": nt + 1,
                f"ss_{str(nt+1).zfill(2)}": ts.strftime("%Y-%m-%d %H:%M:%S"),
                f"ts_{str(nt+1).zfill(2)}": data["timestamp"],
                f"ty_{str(nt+1).zfill(2)}": "in" if reg[0].get(f"ty_{str(nt).zfill(2)}") == "out" else "out",
                f"auto_{str(nt+1).zfill(2)}": 1 if data.get("auto") else 0,
                f"source_{str(nt+1).zfill(2)}": ip,
                "dep": dep,
                "tp_hor": tp_hor
            }
            f = Filters({"num": num, "hsh": reg[0].get("hsh")})
            f.where()
            f.add(f'num = :num', True)
            f.add(f'hsh = :hsh', True)
            f.value("and")
            dml = dbmssql.dml(TypeDml.UPDATE, dti, "rponto.dbo.time_registration", f.parameters, None, False)
            dbmssql.execute(dml.statement, connection, dml.parameters)
            return {"status": "success", "hsh": reg[0].get("hsh")}
    else:
        f = Filters({"num": num, "hsh": hsh})
        f.where()
        f.add(f'num = :num', True)
        f.add(f'hsh = :hsh', True)
        f.value("and")
        reg = dbmssql.executeSimpleList(lambda: (f'SELECT * from rponto.dbo.time_registration {f.text}'), connection, f.parameters)['rows']
        if len(reg) > 0:
            nt = reg[0].get("nt")
            dti = {f"ty_{str(nt).zfill(2)}": data.get("type"), "dep": dep, "tp_hor": tp_hor}
            dml = dbmssql.dml(TypeDml.UPDATE, dti, "rponto.dbo.time_registration", f.parameters, None, False)
            dbmssql.execute(dml.statement, connection, dml.parameters)
            return {"status": "success"}



@api_view(['GET'])
@renderer_classes([JSONRenderer])
def PreProcessImages(request, format=None):
    faces ={}
    for filename in os.listdir(faces_base_path):
        image = preProcessImage(os.path.join(faces_base_path, filename))
        image.save(f"{cropped_faces_base_path}/{filename}")
    return Response({"status":"success"})

def getConfig():
    if os.path.isfile(os.path.join("config.json")):
        with open('config.json', 'rb') as config_file:
            return json.load(config_file)

#CHANGED
def addFace(path,img):
    faces = {"nums": []}
    if os.path.isfile(os.path.join("faces.dictionary")):
        with open('faces.dictionary', 'rb') as faces_file:
            faces = pickle.load(faces_file)
    f = os.path.join(path,img)
    if os.path.isfile(f):
        ki = face_recognition.load_image_file(f)
        faces.get("nums").append({"num":img.split('_')[0],"t_stamp":datetime.today(),"file":img,"matrix":face_recognition.face_encodings(ki,None,jitters,model)[0]})
        with open('faces.dictionary', 'wb') as faces_file:
            pickle.dump(faces, faces_file)
            return True
    return False

#CHANGED
def DelFace(request, format=None):
    filter = request.data['filter']
    if filter.get("num") and filter.get("file"):
        if os.path.isfile(os.path.join("faces.dictionary")):
            with open('faces.dictionary', 'rb') as faces_file:
                faces = pickle.load(faces_file)
                idx = next((index for (index, d) in enumerate(faces.get("nums")) if d["num"] == filter.get("num") and d["file"] == filter.get("file")), None)
                if idx is not None:
                    faces.get("nums").pop(idx)
                    with open('faces.dictionary', 'wb') as faces_file:
                        pickle.dump(faces, faces_file)
                    return Response({"status":"success"})
    return Response({"status":"error"})

def filePathByNum(path,num):
    for i in os.listdir(path):
        if os.path.isfile(os.path.join(path,i)) and i.startswith(num):
            return os.path.join("media",i)
    return None



def SetUser(request, format=None):
    print("ededededededede")
    # connection = connections[connMssqlName].cursor()  # NÃO usar connMssqlName para a FUNC1/SAGE!!
    data = request.data['parameters']
    filter = request.data['filter']
    portugal_timezone = pytz.timezone('Europe/Lisbon')
    current_time_with_dst = datetime.now(portugal_timezone)
    current_time_naive = current_time_with_dst.replace(tzinfo=None)
    ts = current_time_naive
    try:
        if "save" in data and data["save"]==True:
            num = filter["num"]
            func_info = processRecord(num, ts)
            if not func_info.get("nfunc"):
                return Response({"status": "error", "title": f"Funcionário {num} não encontrado na base FUNC1"})
            hsh = data.get("hsh") if data.get("hsh") is not None else None
            if hsh is None:
                if data.get("learn"):
                    fname = f"{num}_{int(datetime.timestamp(datetime.now()))}.jpg"
                    with open(f"{faces_base_path}/{fname}", "wb") as fh:
                        fh.write(base64.b64decode(data["snapshot"].replace('data:image/jpeg;base64,','')))
                    preProcessImage(f"{faces_base_path}/{fname}").save(os.path.join(cropped_faces_base_path, fname), "JPEG")
                    addFace(cropped_faces_base_path, fname)
                try:
                    os.makedirs(f"{records_base_path}/{ts.strftime('%Y%m%d')}")
                except FileExistsError:
                    pass
                try:
                    os.makedirs(f"{records_base_path}/{ts.strftime('%Y%m%d')}/{num}")
                except FileExistsError:
                    pass
                with open(f"{records_base_path}/{ts.strftime('%Y%m%d')}/{num}/{ts.strftime('%Y%m%d.%H%M%S')}.jpg", "wb") as fh:
                    fh.write(base64.b64decode(data["snapshot"].replace('data:image/jpeg;base64,','')))
            # Chama saveRecord tal como em AutoCapture!
            res = saveRecord(num, ts, hsh, data, get_client_ip(request))
            # ADICIONA SEMPRE "rows" com colaborador proveniente do SAGE/FUNC1
            sage100c_connection = connections[connSage100cName].cursor()
            func_sql = f"""
                SELECT F1.NFUNC, F1.NOME
                FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                WHERE F1.NFUNC = '{num}' AND F1.DEMITIDO = 0
            """
            sage100c_connection.execute(func_sql)
            row = sage100c_connection.fetchone()
            if row:
                nfunc, nome = row
                rows = [{"NFUNC": nfunc, "NOME": nome}]
            else:
                rows = []
            return Response({**res, "rows": rows})

        else:
            existsInBd = True
            result = False
            unknown_encoding = []
            unknown_image = None
            filepath = filePathByNum(fotos_base_path, filter["num"])
            faces = loadFaces(faces_base_path)
            tmp = tempfile.NamedTemporaryFile(delete=False)
            try:
                tmp.write(base64.b64decode(data["snapshot"].replace('data:image/jpeg;base64,','')))
                ppi = preProcessImage(tmp.name)
                if ppi is not None:
                    ppi.save(tmp.name, "JPEG")
                    unknown_image = face_recognition.load_image_file(tmp)
                unknown_image = face_recognition.load_image_file(tmp)
            finally:
                tmp.close()
                os.unlink(tmp.name)
            if unknown_image is not None:
                unknown_encoding = face_recognition.face_encodings(unknown_image, None, jitters, model)
            if len(unknown_encoding)==0:
                saveSnapshot(records_invalid_base_path,data["snapshot"],ts,"no_face",filter["num"])
                return Response({"status": "error", "title": "Não foi reconhecida nenhuma face!"})
            unknown_encoding = unknown_encoding[0]

            valid_nums = []
            valid_filepaths = []
            valid_names = []

            try:
                result=False
                existsInBd=False
                for f in faces.get("nums"):
                    if f['num'] == filter["num"]:
                        existsInBd=True
                        results = face_recognition.compare_faces([f["matrix"]], unknown_encoding, tolerance)
                        if len(results) > 0 and True in results:
                            result = True
                            break
            except ValueError:
                existsInBd = False

            if result == False:
                saveSnapshot(records_invalid_base_path, data["snapshot"], ts, "not_identified", filter["num"])

                distances = face_recognition.face_distance([_f['matrix'] for _f in faces.get("nums")], unknown_encoding)
                items = []
                for idx, x in enumerate(distances):
                    if x <= tolerance:
                        items.append({"num": faces.get("nums")[idx].get("num"), "distance": x})
                items = sorted(items, key=lambda x: x["distance"])
                for idx, x in enumerate(items):
                    valid_nums.append(x.get("num"))
                    valid_filepaths.append(filePathByNum(fotos_base_path, x.get("num")))
                if len(valid_nums):
                    sql = lambda: (
                        f"""
                            SELECT F1.NFUNC, F1.NOME 
                            FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                            WHERE F1.DEMITIDO = 0 AND F1.NFUNC IN ({','.join(f"'{w}'" for w in valid_nums)})
                            ORDER BY F1.NFUNC ASC
                        """
                    )
                    response = dbmssql.executeSimpleList(sql, connections[connSage100cName].cursor(), {})
                    if len(response["rows"]) > 0:
                        valid_names = response["rows"]
                if existsInBd == False:
                    added = False
                    fname = f"""{filter["num"]}_{int(datetime.timestamp(datetime.now()))}.jpg"""
                    with open(f"""{faces_base_path}/{fname}""", "wb") as fh:
                        fh.write(base64.b64decode(data["snapshot"].replace('data:image/jpeg;base64,','')))
                    preProcessImage(f"""{faces_base_path}/{fname}""").save(os.path.join(cropped_faces_base_path, fname), "JPEG")
                    added = addFace(cropped_faces_base_path, fname)
                    return Response({"status": "error", "title": f"""O colaborador indicado não existe no sistema! {"A recolha dos dados biométricos foi efetuada." if added else ""}"""})

            f = Filters(request.data['filter'])
            f.setParameters({
                "F1.NFUNC": {"value": lambda v: f"=={v.get('num')}", "field": lambda k, v: f'e.{k}'}
            }, True)
            f.where(False, "and")
            f.auto()
            f.value("and")
            parameters = {**f.parameters}
            dql = dbmssql.dql(request.data, False, False, [])
            if valid_nums:
                sql = lambda: (
                    f"""
                        SELECT F1.NFUNC, F1.NOME 
                        FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                        WHERE F1.DEMITIDO = 0 AND F1.NFUNC IN ({','.join(f"'{w}'" for w in valid_nums)})
                        ORDER BY F1.NFUNC ASC
                    """
                )
                response = dbsage100c.executeSimpleList(sql, connections[connSage100cName].cursor(), {})
            else:
                response = {"rows": []}

            # Sempre buscar colaborador por número na base correta (SAGE):
            func_num = filter["num"]
            sage100c_connection = connections[connSage100cName].cursor()
            func_sql = f"""
                SELECT F1.NFUNC, F1.NOME
                FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                WHERE F1.NFUNC = '{func_num}' AND F1.DEMITIDO = 0
            """
            sage100c_connection.execute(func_sql)
            row = sage100c_connection.fetchone()
            if row:
                nfunc, nome = row
                rows = [{"NFUNC": nfunc, "NOME": nome}]
            else:
                rows = []

            return Response({
                **response,
                "rows": rows,
                "result": result,
                "foto": filepath,
                "valid_nums": valid_nums,
                "valid_filepaths": valid_filepaths,
                "valid_names": valid_names,
                "config": getConfig(),
                "existsInBd": existsInBd
            })
    except Exception as error:
        print(error)
        return Response({"status": "error", "title": str(error)})



def saveSnapshot(basepath,snapshot,tstamp,suffix="",num=None):
    try:
        os.makedirs(f"""{basepath}/{tstamp.strftime("%Y%m%d")}""")
    except FileExistsError:
        pass
    if num is not None:
        try:
            os.makedirs(f"""{basepath}/{tstamp.strftime("%Y%m%d")}/{num}""")
        except FileExistsError:
            pass

    if num is None:
        pth=f"""{basepath}/{tstamp.strftime("%Y%m%d")}/{tstamp.strftime("%Y%m%d.%H%M%S")}.{suffix}.jpg"""
    else:
        pth=f"""{basepath}/{tstamp.strftime("%Y%m%d")}/{num}/{tstamp.strftime("%Y%m%d.%H%M%S")}.{suffix}.jpg"""

    with open(pth, "wb") as fh:
        fh.write(base64.b64decode(snapshot.replace('data:image/jpeg;base64,','')))


def AutoCapture(request, format=None):
    connection = connections[connMssqlName].cursor()
    data = request.data['parameters']
    filter = request.data['filter']
    ts = datetime.now()
    try:
        if "save" in data and data["save"] == True:
            num = filter["num"]
            # Vai buscar os dados atuais do colaborador (dep e tp_hor)
            func_info = processRecord(num, ts)
            if not func_info.get("nfunc"):
                return Response({"status": "error", "title": f"Funcionário {num} não encontrado na base FUNC1"})
            hsh = data.get("hsh") if data.get("hsh") is not None else None
            if hsh is None:
                try:
                    os.makedirs(f"{records_base_path}/{ts.strftime('%Y%m%d')}")
                except FileExistsError:
                    pass
                try:
                    os.makedirs(f"{records_base_path}/{ts.strftime('%Y%m%d')}/{num}")
                except FileExistsError:
                    pass
                with open(f"{records_base_path}/{ts.strftime('%Y%m%d')}/{num}/{ts.strftime('%Y%m%d.%H%M%S')}.jpg", "wb") as fh:
                    fh.write(base64.b64decode(data["snapshot"].replace('data:image/jpeg;base64,','')))
                    
            return Response(saveRecord(num, ts, hsh, data, get_client_ip(request)))

        # RECONHECIMENTO FACIAL (NÃO-SAVE)
        existsInBd = True
        result = False
        unknown_encoding = []
        unknown_image = None
        filepath = None
        faces = loadFaces(faces_base_path)
        tmp = tempfile.NamedTemporaryFile(delete=False)
        try:
            tmp.write(base64.b64decode(data["snapshot"].replace('data:image/jpeg;base64,','')))
            ppi = preProcessImage(tmp.name)
            if ppi is not None:
                ppi.save(tmp.name, "JPEG")
                unknown_image = face_recognition.load_image_file(tmp)
        finally:
            tmp.close()
            os.unlink(tmp.name)
        if unknown_image is not None:
            unknown_encoding = face_recognition.face_encodings(unknown_image, None, jitters, model)
        if len(unknown_encoding) == 0:
            saveSnapshot(records_invalid_base_path, data["snapshot"], ts, "no_face")
            return Response({"status": "error", "title": "Não foi reconhecida nenhuma face!"})
        unknown_encoding = unknown_encoding[0]

        valid_nums = []
        valid_filepaths = []
        valid_names = []
        valid_num = None

        distances = face_recognition.face_distance([_f['matrix'] for _f in faces.get("nums")], unknown_encoding)
        items = []
        for idx, x in enumerate(distances):
            if x <= tolerance:
                items.append({"num": faces.get("nums")[idx].get("num"), "distance": x})
        items = sorted(items, key=lambda x: x["distance"])
        for idx, x in enumerate(items):
            if idx == 0:
                result = True
                valid_num = x.get("num")
                request.data['filter']["num"] = x.get("num")
                filepath = filePathByNum(fotos_base_path, x.get("num"))
            else:
                if x.get("num") != valid_num:
                    valid_nums.append(x.get("num"))
                    valid_filepaths.append(filePathByNum(fotos_base_path, x.get("num")))

        response = {"rows": []}
        if len(valid_nums):
            sage100c_connection = connections[connSage100cName].cursor()
            sql = lambda: (
                f"SELECT F1.NFUNC, F1.NOME FROM TRIMTEK_1GEP.dbo.FUNC1 F1 WHERE F1.DEMITIDO = 0 AND F1.NFUNC = '{filter['num']}'"
            )
            response = dbsage100c.executeSimpleList(sql, sage100c_connection, {})
            if len(response["rows"]) > 0:
                valid_names = response["rows"]

        sage100c_connection = connections[connSage100cName].cursor()
        f = Filters(request.data['filter'])
        f.setParameters({
            "NFUNC": {"value": lambda v: f"=={v.get('num')}", "field": lambda k, v: f'F1.{k}'}
        }, True)
        f.where(False, "and")
        f.auto()
        f.value("and")
        parameters = {**f.parameters}
        dql = dbmssql.dql(request.data, False, False, [])
        sql = lambda: (
            f"""
                SELECT F1.NFUNC, F1.NOME 
                FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                WHERE F1.DEMITIDO = 0 {f.text}
                ORDER BY F1.NFUNC ASC
                {dql.limit}
            """
        )
        print("DBG dbsage100c is:", dbsage100c)
        response = dbsage100c.executeSimpleList(sql, sage100c_connection, parameters)
        if result == False and request.data['filter'].get("num") is None:
            saveSnapshot(records_invalid_base_path, data["snapshot"], ts, "not_identified")
            return Response({"status": "error", "title": "O sistema não o(a) identificou!"})
        return Response({**response, "result": result, "num": request.data['filter'].get("num"), "foto": filepath,
                         "valid_nums": valid_nums, "valid_filepaths": valid_filepaths, "valid_names": valid_names,
                         "config": getConfig()})
    except Exception as error:
        print(error)
        return Response({"status": "error", "title": str(error)})



def BiometriasList(request, format=None):
    bios = []
    if os.path.isfile(os.path.join("faces.dictionary")):
        with open('faces.dictionary', 'rb') as faces_file:
            bios = pickle.load(faces_file).get("nums")
    return Response({"rows":bios})


def InvalidRecordsList(request, format=None):
    records = []
    dates = request.data.get("filter").get("fdata")
    num = request.data.get("filter").get("fnum")
    start_date = datetime.today()
    end_date = datetime.today()
    if (dates and len(dates)>0):
        if dates[0] is None and dates[1] is not None:
            start_date = datetime.strptime(dates[1].replace("<=",""), '%Y-%m-%d')
            end_date = start_date
        if dates[1] is None and dates[0] is not None:
            start_date = datetime.strptime(dates[0].replace(">=",""), '%Y-%m-%d')
            end_date = start_date
        if dates[0] is not None and dates[1] is not None:
            start_date = datetime.strptime(dates[0].replace(">=",""), '%Y-%m-%d')
            end_date = datetime.strptime(dates[1].replace("<=",""), '%Y-%m-%d')
    start_date=start_date.date()
    end_date=end_date.date()
    if (num is not None):
        num = f"""F{num.replace("F","").replace("f","").zfill(5)}"""
    for root, dirs, files in os.walk(records_invalid_base_path):
        for idx,file in enumerate(files):
            # Get the full path of the file
            creation_date = datetime.fromtimestamp(pathlib.Path(os.path.join(root, file)).stat().st_ctime)
            if creation_date.date()>=start_date and creation_date.date()<=end_date:
                fullpath = os.path.join(root, file).replace("\\","/")
                if num is not None:
                    print(fullpath)
                    if num in fullpath:
                        records.append({"k":f"f-{idx}-{random.randint(111111, 999999)}", "filename":fullpath,"tstamp":creation_date.strftime("%Y-%m-%d %H:%M:%S")})
                else:
                    records.append({"k":f"f-{idx}-{random.randint(111111, 999999)}", "filename":fullpath,"tstamp":creation_date.strftime("%Y-%m-%d %H:%M:%S")})
    return Response({"rows":records})


def UpdatePicagem(request, format=None):
    try:
        payload = request.data.get('filter', {}).get('payload', {})

        num = payload.get('num')
        dts = payload.get('dts')

        if not num or not dts:
            return Response({"status": "error", "title": "num e dts são obrigatórios"})

        # Build update dict
        update_data = {}
        nt = 0
        for i in range(1, 9):
            padded = str(i).zfill(2)
            ss_key = f'ss_{padded}'
            ty_key = f'ty_{padded}'
            ss_val = payload.get(ss_key)
            ty_val = payload.get(ty_key)

            if ss_val:
                ss_val = ss_val.replace('T', ' ')
                if len(ss_val) == 16:          
                    ss_val = ss_val + ':00'
                nt += 1

            update_data[ss_key] = ss_val if ss_val else None
            update_data[ty_key] = ty_val.strip() if ty_val and ty_val.strip() else None

        update_data['nt'] = nt
        update_data['edited'] = 1

        with connections[connMssqlName].cursor() as cursor:
            set_parts = []
            params = []
            for key, val in update_data.items():
                set_parts.append(f"{key} = %s")
                params.append(val)

            params.extend([num, dts[:10]])  

            sql = f"""
                UPDATE rponto.dbo.time_registration
                SET {', '.join(set_parts)}
                WHERE num = %s
                  AND CONVERT(DATE, dts) = %s
            """
            cursor.execute(sql, params)

        return Response({"status": "success", "title": "Registo actualizado com sucesso!"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})



# def RegistosRH(request, format=None):
#     print("RegistosRH")
    
#     connection_rponto = connections[connMssqlName].cursor()
#     connection_sage = connections[connSage100cName].cursor()
    
#     try:
#         # Handle filters manually
#         filter_data = request.data.get('filter', {})
#         fnum_value = filter_data.get('fnum')
#         fdata_value = filter_data.get('fdata')
#         fnome_value = filter_data.get('fnome', '').lower()
        
#         print(f" DEBUG filter_data: {filter_data}")
#         print(f" DEBUG fdata_value: {fdata_value}")
#         print(f"DEBUG fdata_value type: {type(fdata_value)}")
        
#         where_clause = ""
#         parameters = {}
        
#         if fnum_value:
#             where_clause = "WHERE TR.num LIKE %(fnum)s"
#             parameters['fnum'] = fnum_value
        
#                 # CORREÇÃO: Processar corretamente o array de datas
#         if fdata_value:
#             print(f" Processando fdata_value: {fdata_value}")
            
#             # Se for um dicionário com 'formatted' (formato atual)
#             if isinstance(fdata_value, dict) and 'formatted' in fdata_value:
#                 formatted = fdata_value['formatted']
#                 if isinstance(formatted, dict):
#                     start_date = formatted.get('startValue')
#                     end_date = formatted.get('endValue')
                    
#                     print(f" start_date extraída: {start_date}")
#                     print(f" end_date extraída: {end_date}")
                    
#                     if start_date and end_date:
#                         and_or_where = " AND " if where_clause else "WHERE "
#                         where_clause += f"{and_or_where}TR.dts >= %(fdata_start)s AND TR.dts <= %(fdata_end)s"
#                         parameters['fdata_start'] = start_date + ' 00:00:00'  
#                         parameters['fdata_end'] = end_date + ' 23:59:59' 
            
#             elif isinstance(fdata_value, list) and len(fdata_value) >= 2:
#                 start_date = str(fdata_value[0]).replace(">=", "").strip()
#                 end_date = str(fdata_value[1]).replace("<=", "").strip()
                
#                 print(f" start_date extraída: {start_date}")
#                 print(f" end_date extraída: {end_date}")
                
#                 and_or_where = " AND " if where_clause else "WHERE "
#                 where_clause += f"{and_or_where}TR.dts >= %(fdata_start)s AND TR.dts <= %(fdata_end)s"
#                 parameters['fdata_start'] = start_date
#                 parameters['fdata_end'] = end_date
            
#             elif isinstance(fdata_value, dict):
#                 start_date = fdata_value.get(">=") or fdata_value.get("formatted", [None])[0]
#                 end_date = fdata_value.get("<=") or fdata_value.get("formatted", [None, None])[1]
                
#                 if start_date and end_date:
#                     and_or_where = " AND " if where_clause else "WHERE "
#                     where_clause += f"{and_or_where}TR.dts >= %(fdata_start)s AND TR.dts <= %(fdata_end)s"
#                     parameters['fdata_start'] = start_date
#                     parameters['fdata_end'] = end_date

#         print(f" DEBUG WHERE CLAUSE: {where_clause}")
#         print(f" DEBUG PARAMS: {parameters}")

#         dql = dbmssql.dql(request.data, False)
        
#         cols = """
#             TR.id, TR.num, TR.dts, TR.dep, TR.tp_hor,
#             TR.ss_01, TR.ty_01, TR.ss_02, TR.ty_02,
#             TR.ss_03, TR.ty_03, TR.ss_04, TR.ty_04,
#             TR.ss_05, TR.ty_05, TR.ss_06, TR.ty_06,
#             TR.ss_07, TR.ty_07, TR.ss_08, TR.ty_08,
#             TR.nt
#         """
        
#         dql.columns = encloseColumn(cols, False)
        
#         # Função de paginação segura para SQL Server
#         def sql_rponto(paging_func, columns_func, sort_func):
#             offset = (dql.currentPage - 1) * dql.pageSize  
#             limit = dql.pageSize
#             order_clause = sort_func(dql.sort) if dql.sort else "ORDER BY TR.dts DESC, TR.num ASC"
#             paging_clause = f"OFFSET {offset} ROWS FETCH NEXT {limit} ROWS ONLY"
            
#             sql_query = f"""
#                 SELECT {columns_func(dql.columns)}
#                 FROM rponto.dbo.time_registration TR
#                 {where_clause}
#                 {order_clause}
#                 {paging_clause}
#             """
#             print("🔧 DEBUG SQL:", sql_query)
#             return sql_query
        
#         response_rponto = dbmssql.executeList(
#             sql_rponto, 
#             connection_rponto, 
#             parameters,  
#             [],  #
#             None,  
#             f"select {dql.currentPage * dql.pageSize + 1}"  
#         )
        
#         if not response_rponto.get('rows'):
#             return Response({
#                 "rows": [], "total": 0, "page": dql.currentPage,
#                 "pageSize": dql.pageSize, "status": "success"
#             })
        
#         registos = response_rponto['rows']
#         total_records = response_rponto.get('total', 0)
        
#         nums_list = list(set([r['num'] for r in registos if r.get('num')]))
#         funcionarios_dict = {}
#         if nums_list:
#             placeholders = ','.join(['%s' for _ in nums_list])
#             sql_func1 = f"SELECT NFUNC, NOME FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE NFUNC IN ({placeholders})"
#             connection_sage.execute(sql_func1, tuple(nums_list))
#             columns_func1 = [col[0] for col in connection_sage.description]
#             for row in connection_sage.fetchall():
#                 func_data = dict(zip(columns_func1, row))
#                 funcionarios_dict[func_data['NFUNC']] = func_data

#         # 3. Normalização e Processamento de Turnos
#         registos_normalizados = []
#         for registro in registos:
#             primeira_picagem_dt = None
            
#             # Limpar e normalizar os tipos de picagem
#             for i in range(1, 9):
#                 ty_key = f'ty_{i:02d}'
#                 if registro.get(ty_key):
#                     # Garante que é string, remove espaços e converte para minúsculas
#                     registro[ty_key] = str(registro[ty_key]).strip().lower()
            
#             for i in range(1, 9):
#                 ss_key = f'ss_{i:02d}'
#                 val = registro.get(ss_key)
#                 if val:
#                     dt_obj = val
#                     if isinstance(val, str):
#                         try: dt_obj = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
#                         except: pass
                    
#                     if not primeira_picagem_dt:
#                         primeira_picagem_dt = dt_obj
                    
#                     if isinstance(dt_obj, (datetime, date)):
#                         registro[ss_key] = dt_obj.strftime('%Y-%m-%d %H:%M:%S')

#             if primeira_picagem_dt:
#                 hora = primeira_picagem_dt.hour
#                 if hora < 6:
#                     data_turno = (primeira_picagem_dt - timedelta(days=1)).date()
#                 else:
#                     data_turno = primeira_picagem_dt.date()
                
#                 registro['data_turno'] = data_turno.strftime('%Y-%m-%d')
#                 registro['tipo_turno'] = identificar_tipo_turno(hora)
#             else:
#                 dts = registro.get('dts')
#                 registro['data_turno'] = dts.strftime('%Y-%m-%d') if isinstance(dts, (datetime, date)) else str(dts)
#                 registro['tipo_turno'] = 'N/A'

#             num = registro.get('num')
#             registro['nome_colaborador'] = funcionarios_dict.get(num, {}).get('NOME', 'Nome não disponível')
            
#             if isinstance(registro.get('dts'), (datetime, date)):
#                 registro['dts'] = registro['dts'].strftime('%Y-%m-%d %H:%M:%S')

#             registos_normalizados.append(registro)

#         # 4. Filtro manual por nome
#         if fnome_value:
#             registos_normalizados = [r for r in registos_normalizados if fnome_value in r.get('nome_colaborador', '').lower()]
#             total_records = len(registos_normalizados)

#         return Response({
#             "rows": registos_normalizados,
#             "total": total_records,
#             "page": dql.currentPage,
#             "pageSize": dql.pageSize,
#             "status": "success"
#         })
        
#     except Exception as error:
#         print(f" Erro em RegistosRH: {str(error)}")
#         import traceback
#         traceback.print_exc()
#         return Response({"status": "error", "title": str(error)})
    
#     finally:
#         connection_rponto.close()
#         connection_sage.close()


#em cima esta a função original
def RegistosRH(request, format=None):
    connection_rponto = connections[connMssqlName].cursor()
    connection_sage   = connections[connSage100cName].cursor()

    try:
        filter_data = request.data.get('filter', {})

        # Filtros funcionais
        fnum_value  = filter_data.get('fnum')
        fdata_value = filter_data.get('fdata')
        fnome_value = (filter_data.get('fnome') or '').lower().strip()

        # Contexto/permissões
        is_rh      = bool(filter_data.get('isRH', False))
        is_admin   = bool(filter_data.get('isAdmin', False))
        is_chefe   = bool(filter_data.get('isChefe', False))
        deps_chefe = filter_data.get('deps_chefe', []) or []

        # Colaborador autenticado (para restringir acesso pessoal)
        num_auth = str(filter_data.get('num') or '').strip()

        # Fallback: RegistosRHPessoal envia apenas fnum quando num está vazio.
        # Aceitar fnum como identidade SOMENTE quando não é RH/Admin/Chefe,
        # para não interferir com a pesquisa por número dos outros papéis.
        if not num_auth and not is_rh and not is_admin and not is_chefe:
            fnum_fallback = str(filter_data.get('fnum') or '').replace('%', '').strip()
            if fnum_fallback:
                num_auth = fnum_fallback

        # Toggle para listagem geral no picagensv3
        is_picagens_v3_list = bool(filter_data.get('isPicagensV3List', False))

        # ══════════════════════════════════════════════════════════
        # SEGURANÇA
        # ══════════════════════════════════════════════════════════

        # Chefe sem departamentos -> vazio
        if is_chefe and not is_rh and not is_admin and not deps_chefe:
            return Response({
                "rows": [],
                "total": 0,
                "page": 1,
                "pageSize": 0,
                "status": "success",
                "warn": "Chefe sem departamentos configurados"
            })

        # Colaborador sem num autenticado -> vazio
        if not is_rh and not is_admin and not is_chefe and not num_auth:
            return Response({
                "rows": [],
                "total": 0,
                "page": 1,
                "pageSize": 0,
                "status": "success",
                "warn": "Colaborador sem número autenticado"
            })

        # ══════════════════════════════════════════════════════════
        # WHERE dinâmico
        # ══════════════════════════════════════════════════════════
        where_parts = []
        parameters  = {}

        # Restrição base por papel
        if is_rh or is_admin:
            # RH/Admin: sem restrição base — filtros opcionais abaixo
            pass

        elif is_chefe:
            # Chefe só vê os seus departamentos
            dep_placeholders = []
            for i, dep in enumerate(deps_chefe):
                key = f"dep_{i}"
                dep_placeholders.append(f"%({key})s")
                parameters[key] = str(dep).strip()
            where_parts.append(f"RTRIM(LTRIM(TR.dep)) IN ({', '.join(dep_placeholders)})")

        else:
            # Colaborador normal -> só o próprio
            where_parts.append("TR.num = %(num_auth)s")
            parameters["num_auth"] = num_auth

        # Filtro por número
        # Para colaborador normal: o fnum é a identidade (já tratado acima via num_auth),
        # por isso só aplica fnum como filtro adicional para RH/Admin/Chefe.
        fnum_clean = str(fnum_value or '').replace('%', '').strip()
        if fnum_clean:
            if is_rh or is_admin or is_chefe:
                where_parts.append("TR.num LIKE %(fnum)s")
                parameters["fnum"] = f"%{fnum_clean}%"
            # Para colaborador normal não adicionamos filtro fnum separado
            # porque num_auth já garante "TR.num = %(num_auth)s" acima.
            # Evita duplicação do mesmo num no WHERE.

        _ = is_picagens_v3_list  # reconhecido mas sem restrição extra

        # Filtro por data
        if fdata_value:
            start_date = None
            end_date   = None

            # formato: { formatted: { startValue, endValue } }
            if isinstance(fdata_value, dict) and 'formatted' in fdata_value:
                formatted = fdata_value.get('formatted') or {}
                if isinstance(formatted, dict):
                    start_date = formatted.get('startValue')
                    end_date   = formatted.get('endValue')

            # formato: [">=YYYY-MM-DD", "<=YYYY-MM-DD"]
            elif isinstance(fdata_value, list) and len(fdata_value) >= 2:
                raw_start  = str(fdata_value[0]).replace(">=", "").strip()
                raw_end    = str(fdata_value[1]).replace("<=", "").strip()
                start_date = raw_start.split(' ')[0].strip()
                end_date   = raw_end.split(' ')[0].strip()

            # formato: {">=": "...", "<=": "..."}
            elif isinstance(fdata_value, dict):
                start_date = fdata_value.get(">=")
                end_date   = fdata_value.get("<=")

            if start_date and end_date:
                where_parts.append("TR.dts >= %(fdata_start)s AND TR.dts <= %(fdata_end)s")
                parameters["fdata_start"] = f"{start_date} 00:00:00"
                parameters["fdata_end"]   = f"{end_date} 23:59:59"

        where_clause = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""

        # ══════════════════════════════════════════════════════════
        # Query principal
        # ══════════════════════════════════════════════════════════
        dql = dbmssql.dql(request.data, False)

        cols = """
            TR.id, TR.num, TR.dts, TR.dep, TR.tp_hor,
            TR.ss_01, TR.ty_01, TR.ss_02, TR.ty_02,
            TR.ss_03, TR.ty_03, TR.ss_04, TR.ty_04,
            TR.ss_05, TR.ty_05, TR.ss_06, TR.ty_06,
            TR.ss_07, TR.ty_07, TR.ss_08, TR.ty_08,
            TR.nt
        """
        dql.columns = encloseColumn(cols, False)

        def sql_rponto(paging_func, columns_func, sort_func):
            offset        = (dql.currentPage - 1) * dql.pageSize
            limit         = dql.pageSize
            order_clause  = sort_func(dql.sort) if dql.sort else "ORDER BY TR.dts DESC, TR.num ASC"
            paging_clause = f"OFFSET {offset} ROWS FETCH NEXT {limit} ROWS ONLY"

            return f"""
                SELECT {columns_func(dql.columns)}
                FROM rponto.dbo.time_registration TR
                {where_clause}
                {order_clause}
                {paging_clause}
            """

        response_rponto = dbmssql.executeList(
            sql_rponto,
            connection_rponto,
            parameters,
            [],
            None,
            f"select {dql.currentPage * dql.pageSize + 1}"
        )

        if not response_rponto.get('rows'):
            return Response({
                "rows":     [],
                "total":    0,
                "page":     dql.currentPage,
                "pageSize": dql.pageSize,
                "status":   "success"
            })

        registos      = response_rponto.get('rows', [])
        total_records = response_rponto.get('total', 0)

        # ══════════════════════════════════════════════════════════
        # Enriquecer com nome colaborador (SAGE)
        # ══════════════════════════════════════════════════════════
        nums_list         = list({r.get('num') for r in registos if r.get('num')})
        funcionarios_dict = {}

        if nums_list:
            placeholders_sage = ','.join(['%s' for _ in nums_list])
            sql_func = (
                f"SELECT NFUNC, NOME "
                f"FROM TRIMTEK_1GEP.dbo.FUNC1 "
                f"WHERE NFUNC IN ({placeholders_sage})"
            )
            connection_sage.execute(sql_func, tuple(nums_list))
            cols_func = [col[0] for col in connection_sage.description]

            for row in connection_sage.fetchall():
                data = dict(zip(cols_func, row))
                funcionarios_dict[data['NFUNC']] = data

        # ══════════════════════════════════════════════════════════
        # Normalização + cálculo turno
        # ══════════════════════════════════════════════════════════
        registos_normalizados = []

        for registro in registos:
            primeira_picagem_dt = None

            # ty_01..ty_08 — normalizar para lowercase sem espaços
            for i in range(1, 9):
                ty_key = f"ty_{i:02d}"
                if registro.get(ty_key):
                    registro[ty_key] = str(registro[ty_key]).strip().lower()

            # ss_01..ss_08 — converter para string normalizada
            for i in range(1, 9):
                ss_key = f"ss_{i:02d}"
                val = registro.get(ss_key)
                if val:
                    dt_obj = val
                    if isinstance(val, str):
                        try:
                            dt_obj = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
                        except Exception:
                            pass

                    if not primeira_picagem_dt:
                        primeira_picagem_dt = dt_obj

                    if isinstance(dt_obj, (datetime, date)):
                        registro[ss_key] = dt_obj.strftime('%Y-%m-%d %H:%M:%S')

            # Calcular data e tipo de turno
            if primeira_picagem_dt:
                hora       = primeira_picagem_dt.hour
                data_turno = (
                    (primeira_picagem_dt - timedelta(days=1)).date()
                    if hora < 6
                    else primeira_picagem_dt.date()
                )
                registro['data_turno'] = data_turno.strftime('%Y-%m-%d')
                registro['tipo_turno'] = identificar_tipo_turno(hora)
            else:
                dts = registro.get('dts')
                registro['data_turno'] = (
                    dts.strftime('%Y-%m-%d')
                    if isinstance(dts, (datetime, date))
                    else str(dts)
                )
                registro['tipo_turno'] = 'N/A'

            # Nome colaborador
            num = registro.get('num')
            registro['nome_colaborador'] = (
                funcionarios_dict.get(num, {}).get('NOME', 'Nome não disponível')
            )

            # Normalizar dts para string
            if isinstance(registro.get('dts'), (datetime, date)):
                registro['dts'] = registro['dts'].strftime('%Y-%m-%d %H:%M:%S')

            registos_normalizados.append(registro)

        # Filtro por nome (pós-processamento — só usado quando fnome enviado)
        if fnome_value:
            registos_normalizados = [
                r for r in registos_normalizados
                if fnome_value in (r.get('nome_colaborador') or '').lower()
            ]
            total_records = len(registos_normalizados)

        return Response({
            "rows":     registos_normalizados,
            "total":    total_records,
            "page":     dql.currentPage,
            "pageSize": dql.pageSize,
            "status":   "success"
        })

    except Exception as error:
        traceback.print_exc()
        return Response({"status": "error", "title": str(error)})

    finally:
        connection_rponto.close()
        connection_sage.close()



def identificar_tipo_turno(hora_entrada):
    """
    Identifica o tipo de turno baseado na hora de entrada
    """
    if 6 <= hora_entrada < 14:
        return "MANHÃ (08:00-14:00)"
    elif 14 <= hora_entrada < 22:
        return "TARDE (14:00-22:00)"
    else:  # 22:00-06:00
        return "NOITE (22:00-06:00)"



    

def GetTurnosEquipas(request, format=None):
    parameters = request.data. get('parameters', {})
    data_inicio_str = parameters.get('data_inicio')
    data_fim_str = parameters. get('data_fim')
    
    if not data_inicio_str: 
        dt_inicio = datetime.now().replace(day=1)
        data_inicio_str = dt_inicio.strftime('%Y-%m-%d')
    else: 
        dt_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
    
    if not data_fim_str: 
        proximo_mes = (dt_inicio. replace(day=28) + timedelta(days=4)).replace(day=1)
        data_fim = proximo_mes - timedelta(days=1)
        data_fim_str = data_fim. strftime('%Y-%m-%d')
    
    query = f"""
    SET DATEFIRST 1;
    
    SELECT 
        FORMAT(DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01'), 'yyyy-MM-dd') AS data,
        DATENAME(WEEKDAY, DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01')) AS dia_semana,
        c.equipa_letra AS equipa,
        c.esquema_tipo AS esquema,
        CASE 
            WHEN c.ordem_rotacao IN (1, 358, 359, 365) THEN 'DSC'
            ELSE c.turno_sigla
        END AS turno_sigla,
        CASE
            WHEN c. ordem_rotacao IN (1, 358, 359, 365) THEN 'Feriado'
            ELSE COALESCE(t.nome, 'Sem turno')
        END AS turno_nome,
        CASE 
            WHEN c.ordem_rotacao IN (1, 358, 359, 365) THEN NULL 
            ELSE t.hora_inicio 
        END AS hora_inicio,
        CASE 
            WHEN c.ordem_rotacao IN (1, 358, 359, 365) THEN NULL 
            ELSE t.hora_fim 
        END AS hora_fim,
        CASE 
            WHEN c.ordem_rotacao IN (1, 358, 359, 365) THEN '#E0E0E0'
            ELSE t.cor_hex 
        END AS cor_hex,
        CASE 
            WHEN c.ordem_rotacao IN (1, 358, 359, 365) THEN 1 
            ELSE 0 
        END AS is_feriado,
        h.name AS nome_feriado
    FROM rponto. dbo.ciclo_laboracao c
    LEFT JOIN rponto. dbo.turnos t ON t.sigla = c.turno_sigla
    LEFT JOIN rponto.dbo.holidays h ON h.holiday_date = DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01')
    WHERE DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01') >= '{data_inicio_str}'
      AND DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01') <= '{data_fim_str}'
    ORDER BY c.ordem_rotacao, c.esquema_tipo, c.equipa_letra;
    """
    
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor. execute(query)
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        escalas_agrupadas = {}
        
        dt_current = dt_inicio
        dt_end = datetime.strptime(data_fim_str, '%Y-%m-%d')
        
        while dt_current <= dt_end:
            data_str = dt_current. strftime('%Y-%m-%d')
            escalas_agrupadas[data_str] = {
                'data':  data_str,
                'dia_semana': dt_current. strftime('%A'),
                'equipas': []
            }
            dt_current += timedelta(days=1)
        
        for row in rows:
            data = row['data']
            if data in escalas_agrupadas:
                escalas_agrupadas[data]['equipas'].append({
                    'equipa': row['equipa'],
                    'esquema': row['esquema'],
                    'turno_sigla': row['turno_sigla'],
                    'turno_nome':  row['turno_nome'],
                    'hora_inicio':  str(row['hora_inicio']) if row['hora_inicio'] else None,
                    'hora_fim': str(row['hora_fim']) if row['hora_fim'] else None,
                    'cor_hex': row['cor_hex'],
                    'is_feriado':  bool(row['is_feriado']),
                    'nome_feriado': row['nome_feriado']
                })
        
        return Response({
            'success': True,
            'data_inicio': data_inicio_str,
            'data_fim':  data_fim_str,
            'total_dias': len(escalas_agrupadas),
            'escalas':  list(escalas_agrupadas.values())
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)





def CalendarList(request, format=None):
    connection = connections[connMssqlName].cursor()
    f = Filters(request.data['filter'])
    f.setParameters({
        #**rangeP(f.filterData.get('fdata'), 'dts', lambda k, v: f'CONVERT(DATE, dts)'),
        "REFNUM_0": {"value": lambda v: f"==F{str(v.get('fnum')).zfill(5)}" if v.get('fnum') is not None else None, "field": lambda k, v: f'T.{k}'},
        "num": {"value": lambda v: f"=={v.get('num')}" if v.get('num') is not None else None, "field": lambda k, v: f'T.REFNUM_0'},
    }, True)
    f.where("")
    f.auto()
    f.value()

    _year = request.data['filter'].get("y") if request.data['filter'].get("y") is not None else datetime.now().year
    _month = request.data['filter'].get("m") if request.data['filter'].get("m") is not None else None
    f2 = Filters(request.data['filter'])
    f2.setParameters({
        #**rangeP(f.filterData.get('fdata'), 'dts', lambda k, v: f'CONVERT(DATE, dts)'),
        "FULLNAME": {"value": lambda v: v.get('fnome').lower() if v.get('fnome') is not None else None, "field": lambda k, v: f'lower({k})'},
        "y": {"value": lambda v: f"=={_year}", "field": lambda k, v: f'C.{k}'},
        "m": {"value": lambda v: f"=={_month}" if _month is not None else None, "field": lambda k, v: f'C.{k}'}
    }, True)
    f2.where()
    f2.auto()
    f2.value()


    def filterMonthMultiSelect(data,name,operator):
        f = Filters(data)
        fP = {}
        if name in data:
            dt = [o['value'] for o in data[name]]
            for idx,v in enumerate(dt):
                fP[f"m{idx}"] = {"key":"m", "value": f"=={v}", "field": lambda k, v: f'C.{k}'}
        f.setParameters({**fP}, True)
        f.auto()
        f.where(False, operator)
        f.value("or")
        return f
    fmonths = filterMonthMultiSelect(request.data['filter'],'months',"and" if f2.hasFilters else "where")
    
    fmulti = filterMulti(request.data['filter'], {
        # 'flotenw': {"keys": ['lotenwinf', 'lotenwsup'], "table": 'mb.'},
        # 'ftiponw': {"keys": ['tiponwinf', 'tiponwsup'], "table": 'mb.'},
        # 'fbobine': {"keys": ['nome'], "table": 'mb.'},
    }, False, "and" if f.hasFilters else "where" ,False)
    fmulti["text"] = f""" """

    parameters = {**f.parameters, **fmulti['parameters'],**f2.parameters,**fmonths.parameters}
    dql = dbmssql.dql(request.data, False)
    cols = f"""*"""
    dql.columns=encloseColumn(cols,False)
    dql.sort = " ORDER BY(SELECT NULL) " if not dql.sort else dql.sort #Obrigatório se PAGING em sqlserver
    sql = lambda p, c, s: (
        f"""            
            WITH [CTE_CALENDAR] AS
            (SELECT CAST('{_year}-01-01' AS DATE) AS [date]
            union all
            select DATEADD(dd,1,[date]) FROM [CTE_CALENDAR]
            WHERE DATEADD(dd,1,[date]) <= CAST('{_year}-12-31' AS DATE)
            ), [CALENDAR] AS 
            (SELECT 
            [date],
            DATEPART(ISO_WEEK,[date]) isowyear,
            DATEPART(WEEK,[date]) wyear,
            DATEPART(WEEKDAY,[date]) wday,
            FORMAT([date], 'dddd', 'pt-pt') wdayname,
            --DATENAME(WEEKDAY,[date]) wdayname,
            DATEPART(MONTH,[date]) m,
            DATEPART(YEAR,[date]) y,
            CASE WHEN DATEPART(ISO_WEEK,[date])>DATEPART(WEEK,[date]) THEN DATEPART(YEAR,[date])-1 ELSE DATEPART(YEAR,[date]) END isoy
            FROM [CTE_CALENDAR]
            )
            SELECT {c(f'{dql.columns}')} FROM (
            SELECT [YEA_0],[WEEK],[DAYWEEK],[REFNUM_0],[PLNTYP_0], EN_MANHA,SA_MANHA,EN_TARDE,SA_TARDE, SRN_0,NAM_0, CONCAT(SRN_0,' ',NAM_0) FULLNAME
            FROM (
            SELECT T.*,EID.SRN_0, EID.NAM_0 FROM (
            select DISTINCT 1 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_0
            UNION ALL
            select DISTINCT 2 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_1
            UNION ALL
            select DISTINCT 3 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_2
            UNION ALL
            select DISTINCT 4 WEEK,YEA_0, REFNUM_0, STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_3
            UNION ALL
            select DISTINCT 5 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_4
            UNION ALL
            select DISTINCT 6 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_5
            UNION ALL
            select DISTINCT 7 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_6
            UNION ALL
            select DISTINCT 8 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_7
            UNION ALL
            select DISTINCT 9 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_8
            UNION ALL
            select DISTINCT 10 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_9
            UNION ALL
            select DISTINCT 11 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_10
            UNION ALL
            select DISTINCT 12 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_11
            UNION ALL
            select DISTINCT 13 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_12
            UNION ALL
            select DISTINCT 14 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_13
            UNION ALL
            select DISTINCT 15 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_14
            UNION ALL
            select DISTINCT 16 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_15
            UNION ALL
            select DISTINCT 17 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_16
            UNION ALL
            select DISTINCT 18 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_17
            UNION ALL
            select DISTINCT 19 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_18
            UNION ALL
            select DISTINCT 20 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_19
            UNION ALL
            select DISTINCT 21 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_20
            UNION ALL
            select DISTINCT 22 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_21
            UNION ALL
            select DISTINCT 23 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_22
            UNION ALL
            select DISTINCT 24 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_23
            UNION ALL
            select DISTINCT 25 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_24
            UNION ALL
            select DISTINCT 26 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_25
            UNION ALL
            select DISTINCT 27 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_26
            UNION ALL
            select DISTINCT 28 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_27
            UNION ALL
            select DISTINCT 29 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_28
            UNION ALL
            select DISTINCT 30 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_29
            UNION ALL
            select DISTINCT 31 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_30
            UNION ALL
            select DISTINCT 32 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_31
            UNION ALL
            select DISTINCT 33 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_32
            UNION ALL
            select DISTINCT 34 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_33
            UNION ALL
            select DISTINCT 35 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_34
            UNION ALL
            select DISTINCT 36 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_35
            UNION ALL
            select DISTINCT 37 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_36
            UNION ALL
            select DISTINCT 38 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_37
            UNION ALL
            select DISTINCT 39 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_38
            UNION ALL
            select DISTINCT 40 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_39
            UNION ALL
            select DISTINCT 41 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_40
            UNION ALL
            select DISTINCT 42 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_41
            UNION ALL
            select DISTINCT 43 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_42
            UNION ALL
            select DISTINCT 44 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_43
            UNION ALL
            select DISTINCT 45 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_44
            UNION ALL
            select DISTINCT 46 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_45
            UNION ALL
            select DISTINCT 47 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_46
            UNION ALL
            select DISTINCT 48 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_47
            UNION ALL
            select DISTINCT 49 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_48
            UNION ALL
            select DISTINCT 50 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_49
            UNION ALL
            select DISTINCT 51 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_50
            UNION ALL
            select DISTINCT 52 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_51
            UNION ALL
            select DISTINCT 53 WEEK,YEA_0, REFNUM_0,STRTIM0_0, ENDTIM0_0,STRTIM1_0, ENDTIM1_0,STRTIM0_1, ENDTIM0_1,STRTIM1_1, ENDTIM1_1,STRTIM0_2, ENDTIM0_2,STRTIM1_2, ENDTIM1_2,STRTIM0_3, ENDTIM0_3,STRTIM1_3, ENDTIM1_3,
            STRTIM0_4, ENDTIM0_4,STRTIM1_4, ENDTIM1_4,STRTIM0_5, ENDTIM0_5,STRTIM1_5, ENDTIM1_5,STRTIM0_6, ENDTIM0_6,STRTIM1_6, ENDTIM1_6,PLNTYP_0
            from x3peoplesql.[PEOPLELTEK].[EMPLOCTR] CT
            JOIN x3peoplesql.[PEOPLELTEK].PLANTYP PT ON PT.COD_0 = CT.PLNTYP_0
            JOIN x3peoplesql.[PEOPLELTEK].TYPWEEK PW0 ON PW0.COD_0 = PT.WEKTYP_52
            ) T 
            JOIN x3peoplesql.[PEOPLELTEK].EMPLOID EID on EID.REFNUM_0 = T.REFNUM_0
            {f.text} {fmulti["text"]}
            --WHERE T.REFNUM_0='F00085' -- AND YEA_0=2020
            ) MyTable
            CROSS APPLY (
            SELECT DAY_ORDER,DAYWEEK,EN_MANHA,SA_MANHA,EN_TARDE,SA_TARDE
            FROM (VALUES
                (0,2,[STRTIM0_0],[ENDTIM0_0],[STRTIM1_0],[ENDTIM1_0]),
                (1,3,[STRTIM0_1],[ENDTIM0_1],[STRTIM1_1],[ENDTIM1_1]),
                (2,4,[STRTIM0_2],[ENDTIM0_2],[STRTIM1_2],[ENDTIM1_2]),
                (3,5,[STRTIM0_3],[ENDTIM0_3],[STRTIM1_3],[ENDTIM1_3]),
                (4,6,[STRTIM0_4],[ENDTIM0_4],[STRTIM1_4],[ENDTIM1_4]),
                (5,7,[STRTIM0_5],[ENDTIM0_5],[STRTIM1_5],[ENDTIM1_5]),
                (6,1,[STRTIM0_6],[ENDTIM0_6],[STRTIM1_6],[ENDTIM1_6])
            ) AS [SourceTable](DAY_ORDER,DAYWEEK,EN_MANHA,SA_MANHA,EN_TARDE,SA_TARDE)
            ) AS [UnpivotTable]
            ) H
            JOIN CALENDAR AS C ON C.wday=H.DAYWEEK and C.isoy=H.YEA_0 AND C.isowyear=H.WEEK
            {f2.text} {fmonths.text}
            {s(dql.sort)} {p(dql.paging)} {p(dql.limit)}
            OPTION(MAXRECURSION 400)
            
        """
    )
    if ("export" in request.data["parameters"]):
        dql.limit=f"""OFFSET 0 ROWS FETCH NEXT {request.data["parameters"]["limit"]} ROWS ONLY"""
        dql.paging=""
        dql.sort = " ORDER BY(date) " if dql.sort == " ORDER BY(SELECT NULL) " else dql.sort #Obrigatório se PAGING em sqlserver
        return export(sql(lambda v:v,lambda v:v,lambda v:v), db_parameters=parameters, parameters={**request.data["parameters"],"filter":request.data.get('filter')},conn_name=AppSettings.reportConn["sage"],dbi=dbmssql,conn=connection)
    try:
        response = dbmssql.executeList(sql, connection, parameters,[],None,None)
    except Exception as error:
        print(str(error))
        return Response({"status": "error", "title": str(error)})
    return Response(response)    




def GetCameraRecords(request, format=None):
    records = []
    parameters = request.data['parameters']
    if parameters.get('date') and parameters.get('num'):
        path = os.path.join(parameters.get('date'),parameters.get('num'))        
        files = os.listdir(os.path.join(records_base_path,path))
        # Create a list of tuples where each tuple contains the filename and its modification time
        file_times = [(f, datetime.fromtimestamp(os.path.getmtime(os.path.join(os.path.join(records_base_path,path), f)))) for f in files]
        # Sort the list of tuples by the modification time
        file_times_sorted = sorted(file_times, key=lambda x: x[1])        
        for f in file_times_sorted:
            filename = f[0]
            v = datetime.strptime(filename.replace(".jpg",""), '%Y%m%d.%H%M%S').strftime("%Y-%m-%d %H:%M:%S")
            records.append({"filename":os.path.join(path,filename).replace("\\","/"),"tstamp":v})
    return Response(records)




# mapeamento de pastas/dep
DEPT_MAPPING = {
    'DAF': 'Contabilidade e Financeiro',
    'DSE': 'Manutenção e SI',
    'DPLAN': 'Planeamento',
    'DPROD': 'Produção',
    'DQUAL': 'Qualidade',
    'DRH': 'RH',
    'DTEC': 'Técnico'
}

#caminho para a pasta alterar mais tarde para Picagem 2026
BASE_PATH = '/mnt/Picagem/Picagem 2026_2'

#helpers

def parse_date_field(date_val):
    if isinstance(date_val, str):
        return date_val[:10]
    if hasattr(date_val, 'strftime'):
        return date_val.strftime('%Y-%m-%d')
    return str(date_val)[:10]


def parse_datetime_field(dt_val):
    if isinstance(dt_val, datetime):
        return dt_val
    if isinstance(dt_val, str):
        try:
            return datetime.strptime(dt_val[:19], '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            return None
    return None


def is_midnight(dt_obj):
    if not dt_obj:
        return False
    return dt_obj.hour == 0 and dt_obj.minute == 0 and dt_obj.second == 0


def format_fnum_filter(fnum_filter):
    if not fnum_filter:
        return None
    
    fnum_filter = fnum_filter.strip()
    
    if '%' in fnum_filter:
        return fnum_filter
    
    if fnum_filter.isdigit():
        return f"F{fnum_filter.zfill(5)}"
    
    if fnum_filter.startswith('F') and len(fnum_filter) < 6:
        num_part = fnum_filter[1:]
        if num_part.isdigit():
            return f"F{num_part.zfill(5)}"
    
    return fnum_filter


#region FUNÇÕES PARA AJUSTE DE TURNO NOTURNO

def _extract_date_str(dts):
    if isinstance(dts, str):
        return dts.split(' ')[0] if ' ' in dts else dts[:10]
    elif isinstance(dts, datetime):
        return dts.strftime('%Y-%m-%d')
    elif hasattr(dts, 'date'):
        return dts.date().isoformat()
    elif hasattr(dts, 'strftime'):
        return dts.strftime('%Y-%m-%d')
    else:
        return str(dts)[:10]


def _parse_datetime_safe(valor):
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor[:19], '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError, IndexError):
            return None
    elif isinstance(valor, datetime):
        return valor
    elif hasattr(valor, 'to_pydatetime'):
        return valor.to_pydatetime()
    return None


def _parse_dt(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor[:19], '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            return None
    if hasattr(valor, 'strftime'):
        return valor
    return None


def _is_meia_noite(dt_obj):
    if not dt_obj:
        return True
    return dt_obj.hour == 0 and dt_obj.minute == 0 and dt_obj.second == 0


def _extrair_picagens_validas(registo):
    picagens = []
    for i in range(1, 9):
        ss_val = registo.get(f'ss_{str(i).zfill(2)}')
        ty_val = str(registo.get(f'ty_{str(i).zfill(2)}') or '').strip().lower()
        if not ss_val:
            continue
        dt_obj = _parse_dt(ss_val)
        if not dt_obj:
            continue
        if _is_meia_noite(dt_obj):
            continue
        picagens.append({'dt': dt_obj, 'tipo': ty_val, 'idx': i})

    # Ordenar sempre por hora real
    picagens.sort(key=lambda x: x['dt'])
    return picagens

def _calcular_horas_dia(registo):
    """
    Retorna dict com:
      minutos         - total de minutos trabalhados
      horas_fmt       - string "HH:MM"
      n_picagens      - número de picagens válidas
      tem_erro        - bool (picagens ímpares ou negativas)
      pausa_almoco    - bool (pausa de almoço detetada)
      minutos_pausa   - duração da pausa em minutos
      pares           - lista de pares {entrada, saida, minutos, horas_fmt}
      picagens_raw    - lista de picagens individuais para display
    """
    picagens = _extrair_picagens_validas(registo)
    n = len(picagens)

    pares = []
    total_minutos = 0
    tem_erro = False
    pausa_almoco = False
    minutos_pausa = 0

    if n == 0:
        return {
            'minutos': 0, 'horas_fmt': '00:00', 'n_picagens': 0,
            'tem_erro': False, 'pausa_almoco': False,
            'minutos_pausa': 0, 'pares': [],
            'picagens_raw': []
        }

    # ── Picagens ímpares → erro ──────────────────────────────
    if n % 2 != 0:
        tem_erro = True
        # Mesmo com erro, tentar calcular o que for possível
        # com os pares disponíveis
        for i in range(0, n - 1, 2):
            entrada = picagens[i]
            saida   = picagens[i + 1]
            diff_s  = (saida['dt'] - entrada['dt']).total_seconds()
            mins    = int(diff_s / 60)
            if 0 < mins <= 960:  # sanity: máx 16h
                total_minutos += mins
                pares.append({
                    'entrada':   entrada['dt'].strftime('%H:%M'),
                    'saida':     saida['dt'].strftime('%H:%M'),
                    'minutos':   mins,
                    'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}"
                })

        return {
            'minutos': total_minutos,
            'horas_fmt': f"{total_minutos // 60:02d}:{total_minutos % 60:02d}",
            'n_picagens': n,
            'tem_erro': True,
            'pausa_almoco': False,
            'minutos_pausa': 0,
            'pares': pares,
            'picagens_raw': [
                {
                    'hora': p['dt'].strftime('%H:%M:%S'),
                    'tipo': p['tipo'],
                    'idx':  p['idx']
                }
                for p in picagens
            ]
        }

    # ── 2 picagens: turno contínuo ───────────────────────────
    if n == 2:
        entrada = picagens[0]
        saida   = picagens[1]
        diff_s  = (saida['dt'] - entrada['dt']).total_seconds()
        mins    = int(diff_s / 60)

        if mins <= 0 or mins > 960:
            tem_erro = True
            mins = 0
        else:
            pares.append({
                'entrada':   entrada['dt'].strftime('%H:%M'),
                'saida':     saida['dt'].strftime('%H:%M'),
                'minutos':   mins,
                'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}"
            })

        return {
            'minutos': mins,
            'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}",
            'n_picagens': n,
            'tem_erro': tem_erro,
            'pausa_almoco': False,
            'minutos_pausa': 0,
            'pares': pares,
            'picagens_raw': [
                {
                    'hora': p['dt'].strftime('%H:%M:%S'),
                    'tipo': p['tipo'],
                    'idx':  p['idx']
                }
                for p in picagens
            ]
        }

    # ── 4 picagens: verificar pausa de almoço ────────────────
    if n == 4:
        # Par 1: P1→P2
        mins_par1 = int(
            (picagens[1]['dt'] - picagens[0]['dt']).total_seconds() / 60
        )
        # Pausa: P2→P3
        mins_pausa_real = int(
            (picagens[2]['dt'] - picagens[1]['dt']).total_seconds() / 60
        )
        # Par 2: P3→P4
        mins_par2 = int(
            (picagens[3]['dt'] - picagens[2]['dt']).total_seconds() / 60
        )

        # Validar pares
        if mins_par1 <= 0 or mins_par2 <= 0:
            tem_erro = True

        # Detetar pausa de almoço (45–75 min)
        if PAUSA_ALMOCO_MIN <= mins_pausa_real <= PAUSA_ALMOCO_MAX:
            pausa_almoco    = True
            minutos_pausa   = mins_pausa_real
            total_minutos   = mins_par1 + mins_par2
        else:
            # Pausa fora do intervalo → contar tudo como real
            pausa_almoco  = False
            minutos_pausa = mins_pausa_real
            total_minutos = mins_par1 + mins_pausa_real + mins_par2

        if mins_par1 > 0:
            pares.append({
                'entrada':   picagens[0]['dt'].strftime('%H:%M'),
                'saida':     picagens[1]['dt'].strftime('%H:%M'),
                'minutos':   mins_par1,
                'horas_fmt': f"{mins_par1 // 60:02d}:{mins_par1 % 60:02d}"
            })
        if mins_par2 > 0:
            pares.append({
                'entrada':   picagens[2]['dt'].strftime('%H:%M'),
                'saida':     picagens[3]['dt'].strftime('%H:%M'),
                'minutos':   mins_par2,
                'horas_fmt': f"{mins_par2 // 60:02d}:{mins_par2 % 60:02d}"
            })

        return {
            'minutos': total_minutos,
            'horas_fmt': f"{total_minutos // 60:02d}:{total_minutos % 60:02d}",
            'n_picagens': n,
            'tem_erro': tem_erro,
            'pausa_almoco': pausa_almoco,
            'minutos_pausa': minutos_pausa,
            'pares': pares,
            'picagens_raw': [
                {
                    'hora': p['dt'].strftime('%H:%M:%S'),
                    'tipo': p['tipo'],
                    'idx':  p['idx']
                }
                for p in picagens
            ]
        }

    # ── 6 ou 8 picagens: somar todos os pares ────────────────
    for i in range(0, n, 2):
        if i + 1 >= n:
            tem_erro = True
            break
        entrada = picagens[i]
        saida   = picagens[i + 1]
        mins    = int((saida['dt'] - entrada['dt']).total_seconds() / 60)

        if mins <= 0 or mins > 960:
            tem_erro = True
            continue

        # Verificar se existe pausa de almoço entre este par e o próximo
        if i + 2 < n:
            pausa_entre = int(
                (picagens[i + 2]['dt'] - saida['dt']).total_seconds() / 60
            )
            if PAUSA_ALMOCO_MIN <= pausa_entre <= PAUSA_ALMOCO_MAX:
                pausa_almoco  = True
                minutos_pausa = pausa_entre

        total_minutos += mins
        pares.append({
            'entrada':   entrada['dt'].strftime('%H:%M'),
            'saida':     saida['dt'].strftime('%H:%M'),
            'minutos':   mins,
            'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}"
        })

    return {
        'minutos': total_minutos,
        'horas_fmt': f"{total_minutos // 60:02d}:{total_minutos % 60:02d}",
        'n_picagens': n,
        'tem_erro': tem_erro,
        'pausa_almoco': pausa_almoco,
        'minutos_pausa': minutos_pausa,
        'pares': pares,
        'picagens_raw': [
            {
                'hora': p['dt'].strftime('%H:%M:%S'),
                'tipo': p['tipo'],
                'idx':  p['idx']
            }
            for p in picagens
        ]
    }

def _calcular_minutos_noturnos(pares_calculados, data_str):
    minutos_noturnos = 0
    try:
        data_base = datetime.strptime(data_str[:10], '%Y-%m-%d').date()
    except Exception:
        return 0

    for par in pares_calculados:
        try:
            entrada_dt = datetime.strptime(
                f"{data_str[:10]} {par['entrada']}", '%Y-%m-%d %H:%M'
            )
            saida_dt = datetime.strptime(
                f"{data_str[:10]} {par['saida']}", '%Y-%m-%d %H:%M'
            )
        except Exception:
            continue

        periodos = [
            (
                datetime.combine(data_base, dt_time(0, 0)),
                datetime.combine(data_base, dt_time(6, 0))
            ),
            (
                datetime.combine(data_base, dt_time(22, 0)),
                datetime.combine(data_base, dt_time(23, 59, 59))
            ),
        ]
        for pn_ini, pn_fim in periodos:
            overlap_ini = max(entrada_dt, pn_ini)
            overlap_fim = min(saida_dt, pn_fim)
            if overlap_fim > overlap_ini:
                minutos_noturnos += int(
                    (overlap_fim - overlap_ini).total_seconds() / 60
                )

    return minutos_noturnos

def _get_nomes_colaboradores(nums_list):
    nomes = {}
    if not nums_list:
        return nomes
    try:
        with connections[connSage100cName].cursor() as cs:
            placeholders = ','.join(['%s'] * len(nums_list))
            cs.execute(
                f"""
                SELECT NFUNC, NOME, DEPARTAMENTO
                FROM TRIMTEK_1GEP.dbo.FUNC1
                WHERE NFUNC IN ({placeholders})
                  AND DEMITIDO = 0
                """,
                nums_list
            )
            for row in cs.fetchall():
                nomes[row[0]] = {'nome': row[1], 'dep_sage': row[2]}
    except Exception as e:
        print(f"[WARN] _get_nomes_colaboradores: {e}")
    return nomes


def ajustar_picagens_turno_noturno(registos_db):
    if not registos_db:
        return registos_db
    
    print("\n" + "="*70)
    print("="*70)
    
    # Converter para dicts
    registos = []
    for reg in registos_db:
        if isinstance(reg, dict):
            registos.append(reg.copy())
        else:
            registos.append(dict(reg))
    
    # reordenar picagens em cada registo 
    print("reordenar picagens em cada registo")
    
    for registo in registos:
        picagens = _extrair_picagens(registo)
        
        if picagens:
            picagens_sorted = sorted(picagens, key=lambda x: x['datetime'])
            
            for i in range(1, 9):
                registo[f'ss_{i:02d}'] = None
            
            for idx, pic in enumerate(picagens_sorted[:8]):
                registo[f'ss_{(idx+1):02d}'] = pic['valor']
    
    # criar índice por (num, data) 
    print("criar índice de registos")
    
    indice_registos = {}
    for reg in registos:
        num = str(reg.get('num', '')).strip()
        data = _extract_date_str(reg.get('dts'))
        chave = (num, data)
        indice_registos[chave] = reg
    
    print(f"  Total de registos únicos: {len(indice_registos)}")
    
    #  combinar múltiplos registos do mesmo dia 
    print(" Combinando múltiplos registos do mesmo dia...")
    
    for chave, registo in list(indice_registos.items()):
        num, data = chave
        picagens = _extrair_picagens(registo)
        
        if not picagens:
            continue
        
        picagens_sorted = sorted(picagens, key=lambda x: x['datetime'])
        
        for i in range(1, 9):
            registo[f'ss_{i:02d}'] = None
        
        for idx, pic in enumerate(picagens_sorted[:8]):
            registo[f'ss_{(idx+1):02d}'] = pic['valor']
    
    print("Reassigning punches based on time...")
    
    new_indice = {}
    
    for (num, data), registo in indice_registos.items():
        picagens = _extrair_picagens(registo)
        
        for pic in picagens:
            punch_date = pic['datetime'].date()
            
            # Move 00:00-07:59 to previous day
            if 0 <= pic['hora'] < 8:
                adjusted_date = punch_date - timedelta(days=1)
            # Move 23:00-23:59 to next day
            elif 23 <= pic['hora'] <= 23:
                adjusted_date = punch_date + timedelta(days=1)
            # Keep 08:00-22:59 on original day
            else:
                adjusted_date = punch_date
            
            adjusted_date_str = adjusted_date.isoformat()
            chave = (num, adjusted_date_str)
            
            # Create or get the adjusted day's record
            if chave not in new_indice:
                new_indice[chave] = {
                    'num': num,
                    'dts': adjusted_date_str,
                    'dep': registo.get('dep', ''),
                    'tp_hor': registo.get('tp_hor', ''),
                    'ss_01': None, 'ss_02': None, 'ss_03': None, 'ss_04': None,
                    'ss_05': None, 'ss_06': None, 'ss_07': None, 'ss_08': None
                }
            
            # Add the punch to the adjusted day's record (next available slot)
            for i in range(1, 9):
                if new_indice[chave][f'ss_{i:02d}'] is None:
                    new_indice[chave][f'ss_{i:02d}'] = pic['valor']
                    break
    
    # Sort punches in each new record
    for registo in new_indice.values():
        picagens = _extrair_picagens(registo)
        if picagens:
            picagens_sorted = sorted(picagens, key=lambda x: x['datetime'])
            for i in range(1, 9):
                registo[f'ss_{i:02d}'] = None
            for idx, pic in enumerate(picagens_sorted[:8]):
                registo[f'ss_{(idx+1):02d}'] = pic['valor']

    registos_finais = list(new_indice.values())
    
    print("\n" + "="*70)
    print(f" CONCLUÍDO: {len(registos_finais)} registos finais")
    print("="*70 + "\n")
    
    return registos_finais


def load_template_workbook(dept_code):
    """Carrega o workbook template do departamento"""
    dept_folder = DEPT_MAPPING.get(dept_code)
    if not dept_folder:
        return None
    
    template_path = os.path.join(BASE_PATH, dept_folder, 'template.xlsx')
    
    if not os.path.exists(template_path):
        return None
    
    try:
        return openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f" Erro ao carregar template para {dept_code}: {str(e)}")
        return None


# ===== FUNÇÕES DE EXPORT =====

def _export_normal(registos_db, funcionarios_dict, start_dt, end_dt, d_start_str, d_end_str, fnum_filter):
    
    registos_por_funcionario = {}
    for registo in registos_db:
        num = registo.get('num', '').strip()
        data_str = parse_date_field(registo.get('dts'))
        
        if num not in registos_por_funcionario:
            registos_por_funcionario[num] = {}
        
        registos_por_funcionario[num][data_str] = registo
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    dias_pt = [
        "Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
        "Sexta-feira", "Sábado", "Domingo"
    ]
    headers = [
        'Nº', 'Departamento', 'Tipo Horário', 'Data', 'Dia da Semana',
        'P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'P7', 'P8', 'Duração'
    ]
    
    for num_func in sorted(registos_por_funcionario.keys()):
        nome_func = funcionarios_dict.get(num_func, num_func)
        registos_func = registos_por_funcionario[num_func]
        
        sheet_name = nome_func[:31]
        for char in ['/', '\\', ':', '*', '?', '[', ']']:
            sheet_name = sheet_name.replace(char, '-')
        
        if not sheet_name or sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = num_func
        
        ws = wb.create_sheet(title=sheet_name)
        
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        
        for i, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        curr_dt = start_dt
        
        while curr_dt <= end_dt:
            data_str = curr_dt.strftime('%Y-%m-%d')
            registo = registos_func.get(data_str, {})
            
            ws.cell(row=row_idx, column=1, value=num_func)
            ws.cell(row=row_idx, column=2, value=registo.get('dep', '').strip() if registo else '')
            ws.cell(row=row_idx, column=3, value=registo.get('tp_hor', '').strip() if registo else '')
            
            cell_data = ws.cell(row=row_idx, column=4, value=curr_dt)
            cell_data.number_format = 'DD/MM/YYYY'
            cell_data.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_idx, column=5, value=dias_pt[curr_dt.weekday()])
            
            picagens = []
            tem_picagens = False
            for i in range(1, 9):
                ss_key = f'ss_{i:02d}'
                ss_val = registo.get(ss_key) if registo else None
                
                if ss_val:
                    dt_pic = parse_datetime_field(ss_val)
                    if dt_pic and not is_midnight(dt_pic):
                        cell = ws.cell(row=row_idx, column=5 + i, value=dt_pic)
                        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                        cell.alignment = Alignment(horizontal='center')
                        picagens.append(dt_pic)
                        tem_picagens = True
            
            # Cálculo de duração
            r = row_idx
            if tem_picagens and len(picagens) >= 2:
                formula = (
                    f'=('
                    f'IF(AND(F{r}<>"",G{r}<>""),G{r}-F{r},0)+'
                    f'IF(AND(H{r}<>"",I{r}<>""),I{r}-H{r},0)+'
                    f'IF(AND(J{r}<>"",K{r}<>""),K{r}-J{r},0)+'
                    f'IF(AND(L{r}<>"",M{r}<>""),M{r}-L{r},0)'
                    f'-IF(('
                    f'IF(AND(F{r}<>"",G{r}<>""),1,0)+'
                    f'IF(AND(H{r}<>"",I{r}<>""),1,0)+'
                    f'IF(AND(J{r}<>"",K{r}<>""),1,0)+'
                    f'IF(AND(L{r}<>"",M{r}<>""),1,0)'
                    f')>0,1/24,0)'
                    f')'
                )
                dur_cell = ws.cell(row=r, column=14, value=formula)
                dur_cell.number_format = '[h]:mm'
                dur_cell.alignment = Alignment(horizontal='center')
            
            # Formatação
            is_weekend = curr_dt.weekday() >= 5
            weekend_fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )
            no_punch_fill = PatternFill(
                start_color="FFF9E6", end_color="FFF9E6", fill_type="solid"
            )
            
            for c in range(1, 15):
                cell = ws.cell(row=row_idx, column=c)
                if is_weekend:
                    cell.fill = weekend_fill
                elif not tem_picagens:
                    cell.fill = no_punch_fill
            
            curr_dt += timedelta(days=1)
            row_idx += 1
        
        # Ajuste de colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['N'].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    if fnum_filter:
        filename = f'Picagens_{fnum_filter}_{d_start_str}_a_{d_end_str}.xlsx'
    else:
        filename = f'Picagens_{d_start_str}_a_{d_end_str}.xlsx'
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    print(f" Exportação normal concluída: {filename}")
    return response


def _export_by_department_worker(registos_db, funcionarios_dict):
    DEPT_FOLDERS = {
        'DAF': 'Contabilidade e Financeiro', 'DSE': 'Manutenção e SI',
        'DPLAN': 'Planeamento', 'DPROD': 'Produção',
        'DQUAL': 'Qualidade', 'DRH': 'RH', 'DTEC': 'Técnico'
    }
    base_path = '/mnt/Picagem/Picagem 2026_2'
    temp_dir = '/tmp/excel_sync'
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        df_novo = pd.DataFrame(registos_db)
        df_novo['nome_funcionario'] = df_novo['num'].map(funcionarios_dict)
        pastas_reais = {d.name.lower().strip(): d.path for d in os.scandir(base_path) if d.is_dir()}

        for dept_code, group_df in df_novo.groupby('dep'):
            dept_code = str(dept_code).strip()
            folder_name = DEPT_FOLDERS.get(dept_code)
            dept_path = pastas_reais.get(folder_name.lower().strip()) if folder_name else None
            if not dept_path: 
                print(f"Pasta não encontrada para {dept_code}")
                continue

            try:
                filename = next(f for f in os.listdir(dept_path) if f.endswith('.xlsm') and not f.startswith('~$'))
                remote_path = os.path.join(dept_path, filename)
                local_path = os.path.join(temp_dir, f"sync_{dept_code}_{filename}")

                # 1. Cópia Binária
                with open(remote_path, 'rb') as f_in, open(local_path, 'wb') as f_out:
                    f_out.write(f_in.read())

                # 2. Carregar e Fazer APPEND
                book = openpyxl.load_workbook(local_path, keep_vba=True)
                if 'Raw Data' in book.sheetnames:
                    sheet = book['Raw Data']
                    data = list(sheet.values)
                    if len(data) > 0:
                        df_old = pd.DataFrame(data[1:], columns=data[0])
                        df_final = pd.concat([df_old, group_df]).drop_duplicates()
                    else:
                        df_final = group_df
                    book.remove(sheet)
                else:
                    df_final = group_df

                # 3. Guardar com Auto-Ajuste
                with pd.ExcelWriter(local_path, engine='openpyxl') as writer:
                    writer.book = book
                    df_final.to_excel(writer, sheet_name='Raw Data', index=False)
                    ws = writer.sheets['Raw Data']
                    for i, col in enumerate(df_final.columns):
                        width = max(df_final[col].astype(str).map(len).max(), len(col)) + 2
                        ws.column_dimensions[get_column_letter(i + 1)].width = width

                # 4. Devolver à rede (Binário)
                with open(local_path, 'rb') as f_src, open(remote_path, 'wb') as f_dst:
                    f_dst.write(f_src.read())
                
                os.remove(local_path)
                print(f" {dept_code} atualizado.")

            except Exception as e:
                print(f" Erro em {dept_code}: {e}")
    except Exception as e:
        print(f" Erro Crítico: {e}")



def _export_clean(registos_db, funcionarios_dict, d_start_str, d_end_str, fnum_filter):
    
    # ===== CONVERTER PARA PANDAS =====
    df = pd.DataFrame(registos_db)
    
    # ===== ADICIONAR NOME DO FUNCIONÁRIO =====
    df['NOME'] = df['num'].apply(lambda x: funcionarios_dict.get(str(x).strip(), ""))
    
    # ===== REORDENAR COLUNAS NA ORDEM SOLICITADA =====
    colunas_ordem = [
        'num',        # Número
        'tp_hor',     # Equipa
        'dep',        # Departamento
        'dts',        # Data
        'NOME',       # Nome
        'nt',         # Picagens (total)
        'ss_01',      # P01
        'ss_02',      # P02
        'ss_03',      # P03
        'ss_04',      # P04
        'ss_05',      # P05
        'ss_06',      # P06
        'ss_07',      # P07
        'ss_08'       # P08
    ]
    
    colunas_existentes = [c for c in colunas_ordem if c in df.columns]
    df = df[colunas_existentes]
    
    # ===== RENOMEAR COLUNAS PARA DISPLAY =====
    colunas_display = {
        'num': 'Número',
        'tp_hor': 'Equipa',
        'dep': 'Departamento',
        'dts': 'Data',
        'NOME': 'Nome',
        'nt': 'Picagens',
        'ss_01': 'P01',
        'ss_02': 'P02',
        'ss_03': 'P03',
        'ss_04': 'P04',
        'ss_05': 'P05',
        'ss_06': 'P06',
        'ss_07': 'P07',
        'ss_08': 'P08'
    }
    
    df_display = df.copy()
    df_display.columns = [colunas_display.get(c, c) for c in df_display.columns]
    
    # ===== LIMPEZA DE DADOS =====
    picagens = ['P01', 'P02', 'P03', 'P04', 'P05', 'P06', 'P07', 'P08']
    for col in picagens:
        if col in df_display.columns:
            df_display[col] = df_display[col].apply(
                lambda x: None if (x and isinstance(x, str) and '00:00:00' in x) else x
            )
    
    # ===== CRIAR WORKBOOK =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dados'
    
    # ===== ESTILOS =====
    cor_header = "4472C4"      # Azul
    cor_clara = "FFFFFF"        # Branco
    cor_escura = "D9E8F5"       # Azul claro
    
    font_header = Font(bold=True, size=11, color="FFFFFF")
    font_normal = Font(size=10)
    
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # ===== ESCREVER HEADER =====
    headers = list(df_display.columns)
    fill_header = PatternFill(start_color=cor_header, end_color=cor_header, fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # ===== ESCREVER DADOS =====
    for row_idx, (_, row) in enumerate(df_display.iterrows(), 2):
        # Linhas alternadas
        row_color = cor_clara if (row_idx - 2) % 2 == 0 else cor_escura
        fill_row = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        
        for col_idx, col_name in enumerate(headers, 1):
            value = row[col_name]
            
            # Tratar data para remover a parte da hora
            if col_name == 'Data' and value:
                value = str(value).split(' ')[0] if ' ' in str(value) else value
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_normal
            cell.fill = fill_row
            cell.border = thin_border
            
            # Alinhamento: texto à esquerda, números ao centro
            if col_name in ['Número', 'Nome', 'Departamento', 'Equipa']:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_center
    
    # ===== AUTO-AJUSTE DE COLUNAS =====
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Definir largura com margem
        adjusted_width = max_length + 2
        
        # Mínimos por coluna
        if column_letter == 'E':  # Nome
            adjusted_width = max(adjusted_width, 30)
        elif column_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:  # Picagens
            adjusted_width = max(adjusted_width, 15)
        else:  # Outras
            adjusted_width = max(adjusted_width, 12)
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'Listagem_{fnum_filter or "Geral"}_{d_start_str}_a_{d_end_str}.xlsx'
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    
    return response


# ===== ENDPOINT PRINCIPAL (ATUALIZADO) =====

@api_view(['POST'])
@renderer_classes([JSONRenderer])
def ExportRegistosExcel(request):
    try:
        filtros = request.data.get('filter', {})
        export_type = request.data.get('exportType')  # 'normal', 'department', 'clean'
        
        fdate_from_raw = filtros.get('fdateFrom')
        fdate_to_raw = filtros.get('fdateTo')
        fnum_filter = filtros.get('fnum', '').strip()
        
        if not fdate_from_raw or not fdate_to_raw:
            return HttpResponse("Filtros de data ausentes.", status=400)
        
        # ===== PARSE DAS DATAS =====
        d_start_str = parse_date_field(fdate_from_raw)
        d_end_str = parse_date_field(fdate_to_raw)
        d_next_day = (datetime.strptime(d_end_str, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
        
        start_dt = datetime.strptime(d_start_str, '%Y-%m-%d')
        end_dt = datetime.strptime(d_end_str, '%Y-%m-%d')
        
        # ===== QUERY SQL (MSSQL) =====
        dbmssql = connections[connMssqlName].cursor()  
        sql = """
            SELECT num, dts, nt, dep, tp_hor, ss_01, ss_02, ss_03, ss_04, ss_05, ss_06, ss_07, ss_08
            FROM rponto.dbo.time_registration
            WHERE dts >= %s AND dts < %s
        """
        params = [d_start_str, d_next_day]
        if fnum_filter:
            sql += " AND num = %s"
            params.append(format_fnum_filter(fnum_filter))
        
        with dbmssql as cursor:
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            registos_db = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        if not registos_db:
            return HttpResponse("Nenhum registo encontrado.", status=404)
        
        # ===== BUSCAR NOMES (SAGE) =====
        nums = list(set(r['num'] for r in registos_db if r.get('num')))
        funcionarios_dict = {}
        
        connSageName = connections[connSage100cName].cursor()
        with connSageName as cursor_sage:
            if nums:
                placeholders = ','.join(['%s'] * len(nums))
                cursor_sage.execute(
                    f"SELECT NFUNC, NOME FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE NFUNC IN ({placeholders})", 
                    nums
                )
                funcionarios_dict = {
                    row[0].strip(): row[1].strip() 
                    for row in cursor_sage.fetchall()
                }
        
        # ===== ROUTER DE EXPORTAÇÃO =====
        
        if export_type == 'clean':
            return _export_clean(
                registos_db,  # ← Registos SEM AJUSTE
                funcionarios_dict, 
                d_start_str, 
                d_end_str, 
                fnum_filter
            )
        
        # Para as outras exportações, FAZER O AJUSTE
        registos_db = ajustar_picagens_turno_noturno(registos_db)
        
        if export_type == 'normal':
            return _export_normal(
                registos_db, 
                funcionarios_dict, 
                start_dt, 
                end_dt, 
                d_start_str, 
                d_end_str, 
                fnum_filter
            )
        
        elif export_type == 'department':
            thread = threading.Thread(
                target=_export_by_department_worker, 
                args=(registos_db, funcionarios_dict)
            )
            thread.daemon = True
            thread.start()
            return JsonResponse(
                {"status": "processing", "message": "Exportação iniciada em background."}, 
                status=202
            )
        
        else:
            return HttpResponse(f"Tipo de exportação inválido: {export_type}", status=400)
        
    except Exception as e:
        print(f" Erro na exportação: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Erro: {str(e)}", status=500)

#region TESTE DE GESTÃO DE DEPARTAMENTO

def _get_colaborador_dep_tphor(num):
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT TOP 1 dep, tp_hor
                FROM rponto.dbo.time_registration
                WHERE num = %s
                  AND dep IS NOT NULL
                ORDER BY dts DESC
            """, [num])
            row = cursor.fetchone()
            if row:
                return row[0], row[1]
    except Exception:
        pass
    return None, None


# ────────────────────────────────────────────────────────────────
# HELPER: obter nome do colaborador via SAGE
# ────────────────────────────────────────────────────────────────
def _get_nome_colaborador(num):
    try:
        with connections[connSage100cName].cursor() as cursor:
            cursor.execute("""
                SELECT F1.NOME 
                FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                WHERE F1.NFUNC = %s AND F1.DEMITIDO = 0
            """, [num])
            row = cursor.fetchone()
            if row:
                return row[0]
    except Exception:
        pass
    return None


# ────────────────────────────────────────────────────────────────
# HELPER: verificar se o utilizador é chefe de algum departamento
# ────────────────────────────────────────────────────────────────
def _get_deps_do_chefe(num_chefe):
    """Devolve lista de dep_codigo que o utilizador chefia (ativos)."""
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT dep_codigo
                FROM rponto.dbo.rh_chefes_departamento
                WHERE num_chefe = %s
                  AND ativo = 1
                  AND (dt_fim IS NULL OR dt_fim >= GETDATE())
            """, [num_chefe])
            rows = cursor.fetchall()
            return [r[0] for r in rows]
    except Exception:
        return []


# ================================================================
# 1. JustificacoesList
#    Utilizado pelo colaborador (só vê as suas)
#    e pelo RH/Chefe (vê todas ou filtradas)
# ================================================================
def JustificacoesList(request, format=None):
    try:
        filter_data     = request.data.get('filter',     {})
        parameters_data = request.data.get('parameters', {})

        # ══════════════════════════════════════════════════════════
        # Papel do utilizador — vem SEMPRE do filter (enviado pelo JWT)
        # ══════════════════════════════════════════════════════════
        is_rh      = filter_data.get('isRH',      False)
        is_chefe   = filter_data.get('isChefe',   False)   # ← CORRIGIDO: filter, não parameters
        deps_chefe = filter_data.get('deps_chefe', [])     # ← ["DPLAN"] vindo do JWT

        # Número do utilizador autenticado (para colaborador normal)
        user_num = (
            getattr(request, 'user_num', None)
            or filter_data.get('_num')
            or filter_data.get('num_chefe')
        )

        # ══════════════════════════════════════════════════════════
        # SEGURANÇA — bloquear acesso sem papel válido
        # ══════════════════════════════════════════════════════════
        # Colaborador sem num → nada
        fnum = (filter_data.get('fnum') or '').strip()
        if not is_rh and not is_chefe and not fnum and not user_num:
            return Response({"rows": [], "total": 0, "status": "success",
                             "warn": "Sem identificação do utilizador"})

        # Chefe sem deps → nada (JWT não renovado)
        if is_chefe and not is_rh and not deps_chefe:
            return Response({"rows": [], "total": 0, "status": "success",
                             "warn": "Chefe sem departamentos — faça logout e login novamente"})

        # ══════════════════════════════════════════════════════════
        # Construir WHERE
        # ══════════════════════════════════════════════════════════
        where_parts = ["1=1"]
        params      = []

        # ── Colaborador normal → só as suas justificações ──────────
        if not is_rh and not is_chefe:
            if fnum:
                clean = fnum.replace('%', '').strip()
                where_parts.append("J.num = %s")
                params.append(clean)
            else:
                where_parts.append("J.num = %s")
                params.append(user_num)

        # ── Chefe → filtra pelos departamentos do JWT ──────────────
        # (dep_codigo na tabela justificacoes = dep do colaborador)
        elif is_chefe and not is_rh:
            placeholders = ', '.join(['%s'] * len(deps_chefe))
            where_parts.append(f"RTRIM(LTRIM(J.dep_codigo)) IN ({placeholders})")
            params.extend([d.strip() for d in deps_chefe])

            # Filtro adicional por num se o chefe pesquisar um colaborador
            if fnum:
                clean = fnum.replace('%', '').strip()
                where_parts.append("J.num LIKE %s")
                params.append(f"%{clean}%")

        # ── RH → vê tudo, pode filtrar por num ────────────────────
        elif is_rh:
            if fnum:
                clean = fnum.replace('%', '').strip()
                where_parts.append("J.num LIKE %s")
                params.append(f"%{clean}%")

        # ── Filtro por status ──────────────────────────────────────
        fstatus = filter_data.get('fstatus')
        if fstatus is not None and fstatus != '':
            where_parts.append("J.status = %s")
            params.append(int(fstatus))

        # ── Filtro por datas ───────────────────────────────────────
        fdata = filter_data.get('fdata')
        if isinstance(fdata, list) and len(fdata) == 2:
            start = str(fdata[0]).replace(">=", "").strip()[:10]
            end   = str(fdata[1]).replace("<=", "").strip()[:10]
            where_parts.append("CAST(J.dt_submissao AS DATE) >= %s")
            where_parts.append("CAST(J.dt_submissao AS DATE) <= %s")
            params.extend([start, end])

        # ── Filtro por departamento (extra, usado pelo RH) ─────────
        fdep = filter_data.get('fdep')
        if fdep and is_rh:
            where_parts.append("J.dep_codigo LIKE %s")
            params.append(fdep)

        where_clause = " AND ".join(where_parts)

        # ══════════════════════════════════════════════════════════
        # Paginação
        # ══════════════════════════════════════════════════════════
        dql       = dbmssql.dql(request.data, False)
        page      = dql.currentPage if hasattr(dql, 'currentPage') else 1
        page_size = dql.pageSize    if hasattr(dql, 'pageSize')    else 20
        offset    = (page - 1) * page_size

        with connections[connMssqlName].cursor() as cursor:
            # Total
            cursor.execute(f"""
                SELECT COUNT(*)
                FROM rponto.dbo.justificacoes J
                WHERE {where_clause}
            """, params)
            total = cursor.fetchone()[0]

            # Dados
            cursor.execute(f"""
                SELECT
                    J.id,
                    J.num,
                    J.dep_codigo,
                    J.tp_hor,
                    CONVERT(VARCHAR, J.dt_inicio,    23)  AS dt_inicio,
                    CONVERT(VARCHAR, J.dt_fim,       23)  AS dt_fim,
                    J.motivo_codigo,
                    M.descricao                            AS motivo_descricao,
                    J.descricao,
                    J.pdf_filename,
                    J.pdf_size,
                    CONVERT(VARCHAR, J.dt_submissao, 120) AS dt_submissao,
                    J.status_chefe,
                    CONVERT(VARCHAR, J.dt_chefe,     120) AS dt_chefe,
                    J.num_chefe,
                    J.obs_chefe,
                    J.status_rh,
                    CONVERT(VARCHAR, J.dt_rh,        120) AS dt_rh,
                    J.num_rh,
                    J.obs_rh,
                    J.status,
                    CASE J.status
                        WHEN 0 THEN 'Pendente (Chefe)'
                        WHEN 1 THEN 'Aguarda RH'
                        WHEN 2 THEN 'Aprovado'
                        WHEN 3 THEN 'Rejeitado pelo Chefe'
                        WHEN 4 THEN 'Rejeitado pelo RH'
                        ELSE        'Desconhecido'
                    END AS status_label
                FROM rponto.dbo.justificacoes J
                LEFT JOIN rponto.dbo.justificacoes_motivos M
                    ON J.motivo_codigo = M.codigo
                WHERE {where_clause}
                ORDER BY J.dt_submissao DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """, params + [offset, page_size])

            columns = [col[0] for col in cursor.description]
            rows    = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # ── Enriquecer com nomes via SAGE ──────────────────────────
        nums_unicos = list({r['num'] for r in rows if r.get('num')})
        nomes_dict  = {}
        if nums_unicos:
            try:
                with connections[connSage100cName].cursor() as cursor_sage:
                    placeholders = ', '.join(['%s'] * len(nums_unicos))
                    cursor_sage.execute(f"""
                        SELECT NFUNC, NOME
                        FROM TRIMTEK_1GEP.dbo.FUNC1
                        WHERE NFUNC IN ({placeholders})
                    """, nums_unicos)
                    for r in cursor_sage.fetchall():
                        nomes_dict[r[0]] = r[1]
            except Exception as e:
                print(f"[WARN] Nomes SAGE: {e}")

        for row in rows:
            row['nome_colaborador'] = nomes_dict.get(row['num'], '')

        return Response({
            "rows":     rows,
            "total":    total,
            "page":     page,
            "pageSize": page_size,
            "status":   "success"
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 2. JustificacaoCreate
#    Colaborador submete nova justificação
#    (sem PDF — o PDF é enviado via UploadJustificacaoPDF)
# ================================================================
def JustificacaoCreate(request, format=None):
    try:
        filter_data = request.data.get('filter', {})
        parameters_data = request.data.get('parameters', {})

        num = filter_data.get('num')
        dt_inicio = filter_data.get('dt_inicio')
        dt_fim = filter_data.get('dt_fim')
        motivo_codigo = filter_data.get('motivo_codigo')
        descricao = filter_data.get('descricao', '')

        if not all([num, dt_inicio, dt_fim, motivo_codigo]):
            return Response({
                "status": "error",
                "title": "Campos obrigatórios em falta: num, dt_inicio, dt_fim, motivo_codigo"
            })

        # Obter dep e tp_hor atual do colaborador
        dep_codigo, tp_hor = _get_colaborador_dep_tphor(num)

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                INSERT INTO rponto.dbo.justificacoes
                    (num, dep_codigo, tp_hor, dt_inicio, dt_fim,
                     motivo_codigo, descricao, dt_submissao, status)
                OUTPUT INSERTED.id
                VALUES (%s, %s, %s, %s, %s, %s, %s, GETDATE(), 0)
            """, [num, dep_codigo, tp_hor, dt_inicio, dt_fim,
                  motivo_codigo, descricao])
            row = cursor.fetchone()
            new_id = row[0] if row else None

        return Response({
            "status": "success",
            "title": "Justificação criada com sucesso!",
            "id": new_id
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 3. JustificacaoAprovar
#    Chefe ou RH aprova/rejeita uma justificação
# ================================================================
def JustificacaoAprovar(request, format=None):
    try:
        filter_data = request.data.get('filter', {})
        parameters_data = request.data.get('parameters', {})

        justificacao_id = filter_data.get('id')
        acao = filter_data.get('acao')          # 'aprovar' ou 'rejeitar'
        obs = filter_data.get('obs', '')
        num_aprovador = filter_data.get('num_aprovador')
        tipo_aprovador = filter_data.get('tipo')  # 'chefe' ou 'rh'

        if not all([justificacao_id, acao, num_aprovador, tipo_aprovador]):
            return Response({
                "status": "error",
                "title": "Campos obrigatórios em falta: id, acao, num_aprovador, tipo"
            })

        if acao not in ['aprovar', 'rejeitar']:
            return Response({"status": "error", "title": "Ação inválida. Use 'aprovar' ou 'rejeitar'."})

        if tipo_aprovador not in ['chefe', 'rh']:
            return Response({"status": "error", "title": "Tipo inválido. Use 'chefe' ou 'rh'."})

        with connections[connMssqlName].cursor() as cursor:
            # Verificar estado atual
            cursor.execute("""
                SELECT id, status, status_chefe, dep_codigo
                FROM rponto.dbo.justificacoes
                WHERE id = %s
            """, [justificacao_id])
            row = cursor.fetchone()

            if not row:
                return Response({"status": "error", "title": "Justificação não encontrada."})

            jid, status_atual, status_chefe_atual, dep_codigo = row

            # ── Chefe a aprovar/rejeitar ──
            if tipo_aprovador == 'chefe':
                if status_atual != 0:
                    return Response({
                        "status": "error",
                        "title": "Esta justificação já foi processada pelo chefe."
                    })

                # Verificar se o aprovador é realmente chefe deste dep
                deps_do_chefe = _get_deps_do_chefe(num_aprovador)
                if dep_codigo and dep_codigo not in deps_do_chefe:
                    return Response({
                        "status": "error",
                        "title": "Não tem permissão para aprovar justificações deste departamento."
                    })

                if acao == 'aprovar':
                    novo_status_chefe = 1
                    novo_status_global = 1  # Aguarda RH
                    titulo = "Aprovado pelo Chefe de Departamento. Aguarda aprovação do RH."
                else:
                    novo_status_chefe = 2
                    novo_status_global = 3  # Rejeitado pelo chefe
                    titulo = "Rejeitado pelo Chefe de Departamento."

                cursor.execute("""
                    UPDATE rponto.dbo.justificacoes
                    SET status_chefe = %s,
                        dt_chefe     = GETDATE(),
                        num_chefe    = %s,
                        obs_chefe    = %s,
                        status       = %s
                    WHERE id = %s
                """, [novo_status_chefe, num_aprovador, obs,
                      novo_status_global, justificacao_id])

            # ── RH a aprovar/rejeitar ──
            elif tipo_aprovador == 'rh':
                if status_atual != 1:
                    return Response({
                        "status": "error",
                        "title": "Esta justificação ainda não foi aprovada pelo chefe ou já foi processada."
                    })

                if acao == 'aprovar':
                    novo_status_rh = 1
                    novo_status_global = 2  # Aprovado final
                    titulo = "Aprovado pelo RH. Processo concluído."
                else:
                    novo_status_rh = 2
                    novo_status_global = 4  # Rejeitado pelo RH
                    titulo = "Rejeitado pelo RH."

                cursor.execute("""
                    UPDATE rponto.dbo.justificacoes
                    SET status_rh = %s,
                        dt_rh     = GETDATE(),
                        num_rh    = %s,
                        obs_rh    = %s,
                        status    = %s
                    WHERE id = %s
                """, [novo_status_rh, num_aprovador, obs,
                      novo_status_global, justificacao_id])

        return Response({"status": "success", "title": titulo})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 4. JustificacaoMotivos
#    Lista motivos disponíveis (para dropdown no frontend)
# ================================================================
def JustificacaoMotivos(request, format=None):
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT id, codigo, descricao
                FROM rponto.dbo.justificacoes_motivos
                WHERE ativo = 1
                ORDER BY id
            """)
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return Response({"rows": rows, "status": "success"})

    except Exception as e:
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 5. DepartamentosList
#    Lista departamentos (para gestão RH)
# ================================================================
def DepartamentosList(request, format=None):
    try:
        with connections[connMssqlName].cursor() as cursor:
            # Verificar se há departamentos em rh_departamentos
            cursor.execute("SELECT COUNT(*) FROM rponto.dbo.rh_departamentos")
            total = cursor.fetchone()[0]

            # Se estiver vazio, popular a partir da time_registration
            if total == 0:
                cursor.execute("""
                    INSERT INTO rponto.dbo.rh_departamentos (codigo, nome, ativo)
                    SELECT
                        RTRIM(LTRIM(dep)) AS codigo,
                        RTRIM(LTRIM(dep)) AS nome,
                        1
                    FROM rponto.dbo.time_registration
                    WHERE dep IS NOT NULL
                      AND RTRIM(LTRIM(dep)) != ''
                    GROUP BY RTRIM(LTRIM(dep))
                    ORDER BY RTRIM(LTRIM(dep))
                """)

            cursor.execute("""
                SELECT
                    id,
                    RTRIM(LTRIM(codigo)) AS codigo,
                    RTRIM(LTRIM(nome))   AS nome,
                    ativo
                FROM rponto.dbo.rh_departamentos
                ORDER BY ativo DESC, codigo
            """)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return Response({"rows": rows, "total": len(rows), "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 6. DepartamentoSave
#    Criar ou atualizar departamento (só RH)
# ================================================================
def DepartamentoSave(request, format=None):
    try:
        f      = request.data.get('filter', {})
        dep_id = f.get('id')
        codigo = f.get('codigo', '').strip().upper()
        nome   = f.get('nome', '').strip()
        ativo  = int(f.get('ativo', 1))

        if not codigo or not nome:
            return Response({"status": "error", "title": "Código e nome são obrigatórios"})

        with connections[connMssqlName].cursor() as cursor:
            if dep_id:
                cursor.execute("""
                    UPDATE rponto.dbo.rh_departamentos
                    SET nome = %s, ativo = %s
                    WHERE id = %s
                """, [nome, ativo, dep_id])
                msg = f"Departamento {codigo} atualizado"
            else:
                # Verificar se já existe
                cursor.execute(
                    "SELECT COUNT(*) FROM rponto.dbo.rh_departamentos WHERE codigo = %s",
                    [codigo]
                )
                if cursor.fetchone()[0] > 0:
                    return Response({
                        "status": "error",
                        "title": f"Departamento {codigo} já existe"
                    })
                cursor.execute("""
                    INSERT INTO rponto.dbo.rh_departamentos (codigo, nome, ativo)
                    VALUES (%s, %s, %s)
                """, [codigo, nome, ativo])
                msg = f"Departamento {codigo} criado com sucesso"

        return Response({"status": "success", "title": msg})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


def ColaboradoresSageLookup(request, format=None):
    """
    Lista colaboradores do SAGE para o Select do modal de chefes.
    Usa: SELECT NOME, NFUNC FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE DEMITIDO = 0
    filter: { search }  — filtra por nome ou número
    """
    try:
        f      = request.data.get('filter', {})
        search = f.get('search', '').strip()

        where  = ["DEMITIDO = 0"]
        params = []

        if search:
            where.append("(NOME LIKE %s OR NFUNC LIKE %s)")
            params.extend([f"%{search}%", f"%{search}%"])

        where_sql = "WHERE " + " AND ".join(where)

        with connections[connSage100cName].cursor() as cursor:
            cursor.execute(f"""
                SELECT TOP 50
                    RTRIM(LTRIM(NFUNC)) AS num,
                    RTRIM(LTRIM(NOME))  AS nome
                FROM TRIMTEK_1GEP.dbo.FUNC1
                {where_sql}
                ORDER BY NOME
            """, params)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return Response({"rows": rows, "total": len(rows), "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 7. ChefesList
#    Lista chefes por departamento
# ================================================================
def ChefesList(request, format=None):
    """
    Lista todos os chefes com nome do SAGE.
    filter: { dep_codigo }  — opcional
    """
    try:
        f         = request.data.get('filter', {})
        dep_codigo = f.get('dep_codigo', '').strip()

        where  = []
        params = []
        if dep_codigo:
            where.append("c.dep_codigo = %s")
            params.append(dep_codigo)

        where_sql = f"WHERE {' AND '.join(where)}" if where else ""

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    c.id,
                    RTRIM(LTRIM(c.dep_codigo)) AS dep_codigo,
                    RTRIM(LTRIM(c.num_chefe))  AS num_chefe,
                    CONVERT(VARCHAR(10), c.dt_inicio, 23) AS dt_inicio,
                    CONVERT(VARCHAR(10), c.dt_fim,    23) AS dt_fim,
                    c.ativo
                FROM rponto.dbo.rh_chefes_departamento c
                {where_sql}
                ORDER BY c.dep_codigo, c.ativo DESC, c.dt_inicio DESC
            """, params)
            cols  = [c[0] for c in cursor.description]
            rows  = [dict(zip(cols, r)) for r in cursor.fetchall()]

        # Enriquecer com nomes do SAGE
        nums  = list({r['num_chefe'] for r in rows if r.get('num_chefe')})
        nomes = _get_nomes_colaboradores(nums)
        for row in rows:
            row['nome_chefe'] = nomes.get(row['num_chefe'], {}).get('nome', '')

        return Response({"rows": rows, "total": len(rows), "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 8. ChefeSave
#    Associar/remover chefe a departamento (só RH)
# ================================================================
def ChefeSave(request, format=None):
    """
    Adiciona um chefe a um departamento.
    Desativa chefes anteriores do mesmo departamento se ativo=1.
    filter: { dep_codigo, num_chefe, dt_inicio, ativo }
    """
    try:
        f          = request.data.get('filter', {})
        dep_codigo = f.get('dep_codigo', '').strip()
        num_chefe  = f.get('num_chefe', '').strip()
        dt_inicio  = f.get('dt_inicio', '').strip()
        ativo      = int(f.get('ativo', 1))

        if not dep_codigo or not num_chefe or not dt_inicio:
            return Response({
                "status": "error",
                "title": "dep_codigo, num_chefe e dt_inicio são obrigatórios"
            })

        # Normalizar num_chefe
        if not num_chefe.startswith('F'):
            num_chefe = f"F{num_chefe.zfill(5)}"

        with connections[connMssqlName].cursor() as cursor:
            # Verificar se já é chefe deste departamento
            cursor.execute("""
                SELECT COUNT(*) FROM rponto.dbo.rh_chefes_departamento
                WHERE dep_codigo = %s AND num_chefe = %s AND ativo = 1
            """, [dep_codigo, num_chefe])
            if cursor.fetchone()[0] > 0:
                return Response({
                    "status": "error",
                    "title": "Este colaborador já é chefe ativo deste departamento"
                })

            # Inserir novo chefe
            cursor.execute("""
                INSERT INTO rponto.dbo.rh_chefes_departamento
                    (dep_codigo, num_chefe, dt_inicio, dt_fim, ativo)
                VALUES (%s, %s, %s, NULL, %s)
            """, [dep_codigo, num_chefe, dt_inicio, ativo])

        return Response({
            "status": "success",
            "title": f"Chefe adicionado ao departamento {dep_codigo}"
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})




# ================================================================
# 9. ChefeDelete
#    Remove (desativa) chefe de departamento
# ================================================================
def ChefeDelete(request, format=None):
    """
    Desativa (soft delete) um chefe.
    filter: { id }
    """
    try:
        chefe_id = request.data.get('filter', {}).get('id')
        if not chefe_id:
            return Response({"status": "error", "title": "id obrigatório"})

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                UPDATE rponto.dbo.rh_chefes_departamento
                SET ativo = 0, dt_fim = GETDATE()
                WHERE id = %s
            """, [chefe_id])

        return Response({"status": "success", "title": "Chefe removido com sucesso"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# 10. JustificacaoPendentesCount
#     Contador de pendentes (para badge no menu)
# ================================================================
def JustificacaoPendentesCount(request, format=None):
    try:
        filter_data = request.data.get('filter', {})
        parameters_data = request.data.get('parameters', {})

        num = filter_data.get('num')
        is_rh = parameters_data.get('isRH', False)
        is_chefe = parameters_data.get('isChefe', False)

        counts = {
            "colaborador": 0,
            "chefe": 0,
            "rh": 0
        }

        with connections[connMssqlName].cursor() as cursor:
            # Pendentes para o colaborador (os seus próprios)
            if num:
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM rponto.dbo.justificacoes
                    WHERE num = %s AND status IN (0, 1)
                """, [num])
                counts["colaborador"] = cursor.fetchone()[0]

            # Pendentes para o chefe
            if is_chefe:
                deps = _get_deps_do_chefe(num)
                if deps:
                    placeholders = ','.join(['%s'] * len(deps))
                    cursor.execute(f"""
                        SELECT COUNT(*)
                        FROM rponto.dbo.justificacoes
                        WHERE status = 0
                          AND dep_codigo IN ({placeholders})
                    """, deps)
                    counts["chefe"] = cursor.fetchone()[0]

            # Pendentes para o RH
            if is_rh:
                cursor.execute("""
                    SELECT COUNT(*)
                    FROM rponto.dbo.justificacoes
                    WHERE status = 1
                """)
                counts["rh"] = cursor.fetchone()[0]

        return Response({"counts": counts, "status": "success"})

    except Exception as e:
        return Response({"status": "error", "title": str(e)})


# ================================================================
# ENDPOINT DEDICADO: Upload de PDF
# Chamado via POST multipart/form-data (separado do def Sql)
# ================================================================
@api_view(['POST'])
@renderer_classes([JSONRenderer])
@permission_classes([IsAuthenticated])
def UploadJustificacaoPDF(request):
    try:
        justificacao_id = request.data.get('id')
        pdf_file = request.FILES.get('pdf')

        if not justificacao_id or not pdf_file:
            return Response({
                "status": "error",
                "title": "id e pdf são obrigatórios."
            })

        # Validar tipo de ficheiro
        allowed_types = ['application/pdf', 'image/jpeg', 'image/png']
        if pdf_file.content_type not in allowed_types:
            return Response({
                "status": "error",
                "title": "Tipo de ficheiro não permitido. Use PDF, JPG ou PNG."
            })

        # Validar tamanho (máx 10MB)
        max_size = 10 * 1024 * 1024
        if pdf_file.size > max_size:
            return Response({
                "status": "error",
                "title": "Ficheiro demasiado grande. Máximo permitido: 10MB."
            })

        # Criar pasta se não existir
        try:
            os.makedirs(JUSTIFICACOES_BASE_PATH, exist_ok=True)
        except Exception:
            pass

        # Gerar nome único para o ficheiro
        ext = os.path.splitext(pdf_file.name)[1].lower() or '.pdf'
        filename = f"just_{justificacao_id}_{uuid.uuid4().hex[:8]}{ext}"
        filepath = os.path.join(JUSTIFICACOES_BASE_PATH, filename)

        # Guardar ficheiro
        with open(filepath, 'wb') as f:
            for chunk in pdf_file.chunks():
                f.write(chunk)

        # Atualizar registo na BD
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                UPDATE rponto.dbo.justificacoes
                SET pdf_filename = %s,
                    pdf_path     = %s,
                    pdf_size     = %s
                WHERE id = %s
            """, [filename, filepath, pdf_file.size, justificacao_id])

        return Response({
            "status": "success",
            "title": "Ficheiro carregado com sucesso!",
            "pdf_filename": filename,
            "pdf_size": pdf_file.size
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# ENDPOINT DEDICADO: Download/Visualização de PDF
@api_view(['GET'])
@renderer_classes([JSONRenderer])
@permission_classes([IsAuthenticated])
def DownloadJustificacaoPDF(request, justificacao_id):
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT pdf_path, pdf_filename
                FROM rponto.dbo.justificacoes
                WHERE id = %s
            """, [justificacao_id])
            row = cursor.fetchone()

        if not row or not row[0]:
            raise Http404("PDF não encontrado.")

        pdf_path, pdf_filename = row

        if not os.path.exists(pdf_path):
            raise Http404("Ficheiro não encontrado no servidor.")

        # Determinar content type
        content_type, _ = mimetypes.guess_type(pdf_path)
        content_type = content_type or 'application/octet-stream'

        response = FileResponse(
            open(pdf_path, 'rb'),
            content_type=content_type
        )
        response['Content-Disposition'] = f'inline; filename="{pdf_filename}"'
        return response

    except Http404:
        raise
    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})

#region Teste processamento salarial

def _calcular_horas_dia(registo):
    """
    Retorna SEMPRE um dict com todas as chaves, mesmo em caso de erro.
    """
    # Resultado base com defaults seguros
    resultado_base = {
        'minutos': 0,
        'horas_fmt': '00:00',
        'n_picagens': 0,
        'tem_erro': False,
        'pausa_almoco': False,
        'minutos_pausa': 0,
        'pares': [],
        'picagens_raw': []
    }

    try:
        picagens = _extrair_picagens_validas(registo)
        n = len(picagens)

        resultado_base['n_picagens']   = n
        resultado_base['picagens_raw'] = [
            {
                'hora':  p['dt'].strftime('%H:%M:%S'),
                'tipo':  p['tipo'],
                'idx':   p['idx']
            }
            for p in picagens
        ]

        if n == 0:
            return resultado_base

        pares         = []
        total_minutos = 0
        tem_erro      = False
        pausa_almoco  = False
        minutos_pausa = 0

        # ── Picagens ímpares → erro ───────────────────────────
        if n % 2 != 0:
            tem_erro = True
            for i in range(0, n - 1, 2):
                entrada = picagens[i]
                saida   = picagens[i + 1]
                diff_s  = (saida['dt'] - entrada['dt']).total_seconds()
                mins    = int(diff_s / 60)
                if 0 < mins <= 960:
                    total_minutos += mins
                    pares.append({
                        'entrada':   entrada['dt'].strftime('%H:%M'),
                        'saida':     saida['dt'].strftime('%H:%M'),
                        'minutos':   mins,
                        'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}"
                    })
            resultado_base.update({
                'minutos':      total_minutos,
                'horas_fmt':    f"{total_minutos // 60:02d}:{total_minutos % 60:02d}",
                'tem_erro':     True,
                'pares':        pares,
            })
            return resultado_base

        # ── 2 picagens ────────────────────────────────────────
        if n == 2:
            diff_s = (picagens[1]['dt'] - picagens[0]['dt']).total_seconds()
            mins   = int(diff_s / 60)

            if mins <= 0 or mins > 960:
                resultado_base['tem_erro'] = True
                return resultado_base

            pares.append({
                'entrada':   picagens[0]['dt'].strftime('%H:%M'),
                'saida':     picagens[1]['dt'].strftime('%H:%M'),
                'minutos':   mins,
                'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}"
            })
            resultado_base.update({
                'minutos':   mins,
                'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}",
                'pares':     pares,
            })
            return resultado_base

        # ── 4 picagens ────────────────────────────────────────
        if n == 4:
            mins_par1       = int((picagens[1]['dt'] - picagens[0]['dt']).total_seconds() / 60)
            mins_pausa_real = int((picagens[2]['dt'] - picagens[1]['dt']).total_seconds() / 60)
            mins_par2       = int((picagens[3]['dt'] - picagens[2]['dt']).total_seconds() / 60)

            if mins_par1 <= 0 or mins_par2 <= 0:
                tem_erro = True

            if PAUSA_ALMOCO_MIN <= mins_pausa_real <= PAUSA_ALMOCO_MAX:
                pausa_almoco  = True
                minutos_pausa = mins_pausa_real
                total_minutos = mins_par1 + mins_par2
            else:
                # Pausa fora do intervalo — contar tudo como real
                total_minutos = mins_par1 + mins_pausa_real + mins_par2

            if mins_par1 > 0:
                pares.append({
                    'entrada':   picagens[0]['dt'].strftime('%H:%M'),
                    'saida':     picagens[1]['dt'].strftime('%H:%M'),
                    'minutos':   mins_par1,
                    'horas_fmt': f"{mins_par1 // 60:02d}:{mins_par1 % 60:02d}"
                })
            if mins_par2 > 0:
                pares.append({
                    'entrada':   picagens[2]['dt'].strftime('%H:%M'),
                    'saida':     picagens[3]['dt'].strftime('%H:%M'),
                    'minutos':   mins_par2,
                    'horas_fmt': f"{mins_par2 // 60:02d}:{mins_par2 % 60:02d}"
                })

            resultado_base.update({
                'minutos':      total_minutos,
                'horas_fmt':    f"{total_minutos // 60:02d}:{total_minutos % 60:02d}",
                'tem_erro':     tem_erro,
                'pausa_almoco': pausa_almoco,
                'minutos_pausa': minutos_pausa,
                'pares':        pares,
            })
            return resultado_base

        # ── 6 ou 8 picagens ───────────────────────────────────
        for i in range(0, n, 2):
            if i + 1 >= n:
                tem_erro = True
                break
            entrada = picagens[i]
            saida   = picagens[i + 1]
            mins    = int((saida['dt'] - entrada['dt']).total_seconds() / 60)

            if mins <= 0 or mins > 960:
                tem_erro = True
                continue

            # Verificar pausa almoço entre este par e o próximo
            if i + 2 < n:
                pausa_entre = int(
                    (picagens[i + 2]['dt'] - saida['dt']).total_seconds() / 60
                )
                if PAUSA_ALMOCO_MIN <= pausa_entre <= PAUSA_ALMOCO_MAX:
                    pausa_almoco  = True
                    minutos_pausa = pausa_entre

            total_minutos += mins
            pares.append({
                'entrada':   entrada['dt'].strftime('%H:%M'),
                'saida':     saida['dt'].strftime('%H:%M'),
                'minutos':   mins,
                'horas_fmt': f"{mins // 60:02d}:{mins % 60:02d}"
            })

        resultado_base.update({
            'minutos':       total_minutos,
            'horas_fmt':     f"{total_minutos // 60:02d}:{total_minutos % 60:02d}",
            'tem_erro':      tem_erro,
            'pausa_almoco':  pausa_almoco,
            'minutos_pausa': minutos_pausa,
            'pares':         pares,
        })
        return resultado_base

    except Exception as e:
        print(f"[ERROR] _calcular_horas_dia inesperado: {e}")
        traceback.print_exc()
        resultado_base['tem_erro'] = True
        return resultado_base


def _calcular_horas_noturnas(registo):
    """Conta minutos entre 00:00-06:00 e 22:00-24:00."""
    minutos_noturnos = 0
    for i in range(1, 9):
        ss_key = f'ss_{str(i).zfill(2)}'
        ty_key = f'ty_{str(i).zfill(2)}'
        ss_val = registo.get(ss_key)
        ty_val = str(registo.get(ty_key) or '').strip().lower()
        if not ss_val or ty_val != 'in':
            continue

        # Hora de entrada noturna
        ss_next = registo.get(f'ss_{str(i + 1).zfill(2)}')
        ty_next = str(registo.get(f'ty_{str(i + 1).zfill(2)}') or '').strip().lower()
        if not ss_next or ty_next != 'out':
            continue

        try:
            entrada = datetime.strptime(str(ss_val)[:19], '%Y-%m-%d %H:%M:%S')
            saida = datetime.strptime(str(ss_next)[:19], '%Y-%m-%d %H:%M:%S')
        except Exception:
            continue

        # Verificar sobreposição com período noturno
        dia = entrada.date()
        periodos_noturnos = [
            (datetime.combine(dia, __import__('datetime').time(0, 0)),
             datetime.combine(dia, __import__('datetime').time(6, 0))),
            (datetime.combine(dia, __import__('datetime').time(22, 0)),
             datetime.combine(dia, __import__('datetime').time(23, 59, 59))),
        ]
        for pn_inicio, pn_fim in periodos_noturnos:
            overlap_inicio = max(entrada, pn_inicio)
            overlap_fim = min(saida, pn_fim)
            if overlap_fim > overlap_inicio:
                minutos_noturnos += int(
                    (overlap_fim - overlap_inicio).total_seconds() / 60
                )
    return minutos_noturnos


# ================================================================
# DEPARTAMENTOS — lista distinct do dep em time_registration
# ================================================================
def DepartamentosDistinctList(request, format=None):
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT
                    RTRIM(LTRIM(dep)) AS codigo
                FROM rponto.dbo.time_registration
                WHERE dep IS NOT NULL
                  AND RTRIM(LTRIM(dep)) != ''
                ORDER BY codigo
            """)
            rows = [
                {'codigo': r[0], 'nome': r[0]}
                for r in cursor.fetchall()
            ]
        return Response({"rows": rows, "status": "success"})
    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# CHEFES — reutiliza rh_chefes_departamento mas alimentado
# com os deps distintos (sem tabela rh_departamentos)
# ================================================================
def ChefesDistinctList(request, format=None):
    """
    Lista chefes por departamento.
    Os departamentos vêm do distinct de time_registration.
    """
    try:
        filter_data = request.data.get('filter', {})
        fdep = filter_data.get('fdep')

        where = "WHERE 1=1"
        params = []
        if fdep:
            where += " AND C.dep_codigo = %s"
            params.append(fdep)

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    C.id,
                    C.dep_codigo,
                    C.num_chefe,
                    CONVERT(VARCHAR, C.dt_inicio, 23) AS dt_inicio,
                    CONVERT(VARCHAR, C.dt_fim, 23)    AS dt_fim,
                    C.ativo
                FROM rponto.dbo.rh_chefes_departamento C
                {where}
                  AND C.ativo = 1
                  AND (C.dt_fim IS NULL OR C.dt_fim >= GETDATE())
                ORDER BY C.dep_codigo, C.dt_inicio DESC
            """, params)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        # Enriquecer com nomes via SAGE
        nums = list(set(r['num_chefe'] for r in rows if r.get('num_chefe')))
        nomes_dict = {}
        if nums:
            try:
                with connections[connSage100cName].cursor() as cs:
                    placeholders = ','.join(['%s'] * len(nums))
                    cs.execute(f"""
                        SELECT NFUNC, NOME
                        FROM TRIMTEK_1GEP.dbo.FUNC1
                        WHERE NFUNC IN ({placeholders})
                    """, nums)
                    for r in cs.fetchall():
                        nomes_dict[r[0]] = r[1]
            except Exception:
                pass

        for row in rows:
            row['nome_chefe'] = nomes_dict.get(row['num_chefe'], '')

        return Response({"rows": rows, "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


def ProcessamentoSalarialDetalhe(request, format=None):
    try:
        f     = request.data.get('filter', {})
        num   = f.get('num')
        fdata = f.get('fdata')

        if not num:
            return Response({"status": "error", "title": "num é obrigatório"})
        if not num.startswith('F'):
            num = f"F{num.zfill(5)}"

        if not fdata or len(fdata) < 2:
            return Response({"status": "error", "title": "Intervalo de datas obrigatório"})

        d_ini_str = str(fdata[0]).replace(">=", "").strip()[:10]
        d_fim_str = str(fdata[1]).replace("<=", "").strip()[:10]

        try:
            d_ini = datetime.strptime(d_ini_str, '%Y-%m-%d').date()
            d_fim = datetime.strptime(d_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"status": "error", "title": "Formato de data inválido"})

        d_fim_exc_str = (d_fim + timedelta(days=1)).strftime('%Y-%m-%d')

        # ── Registos do colaborador ───────────────────────────
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    CONVERT(VARCHAR(10), TR.dts, 23)   AS dts,
                    RTRIM(LTRIM(ISNULL(TR.dep,'')))    AS dep,
                    RTRIM(LTRIM(ISNULL(TR.tp_hor,''))) AS tp_hor,
                    ISNULL(TR.nt, 0)                   AS nt,
                    TR.ss_01, ISNULL(TR.ty_01,'') AS ty_01,
                    TR.ss_02, ISNULL(TR.ty_02,'') AS ty_02,
                    TR.ss_03, ISNULL(TR.ty_03,'') AS ty_03,
                    TR.ss_04, ISNULL(TR.ty_04,'') AS ty_04,
                    TR.ss_05, ISNULL(TR.ty_05,'') AS ty_05,
                    TR.ss_06, ISNULL(TR.ty_06,'') AS ty_06,
                    TR.ss_07, ISNULL(TR.ty_07,'') AS ty_07,
                    TR.ss_08, ISNULL(TR.ty_08,'') AS ty_08
                FROM rponto.dbo.time_registration TR
                WHERE TR.num = %s
                  AND TR.dts >= %s
                  AND TR.dts <  %s
                ORDER BY TR.dts
                """,
                [num, d_ini_str, d_fim_exc_str]
            )
            cols     = [c[0] for c in cursor.description]
            registos = [dict(zip(cols, r)) for r in cursor.fetchall()]

        regs_por_data = {r['dts']: r for r in registos}

        # ── Nome via SAGE ─────────────────────────────────────
        info_sage = _get_nomes_colaboradores([num])
        nome   = info_sage.get(num, {}).get('nome', '')
        dep    = registos[0].get('dep', '')    if registos else ''
        tp_hor = registos[0].get('tp_hor', '') if registos else ''

        # ── Férias aprovadas no período ───────────────────────
        ferias_dias = set()
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT
                    CONVERT(VARCHAR(10), data_ini, 23) AS data_ini,
                    CONVERT(VARCHAR(10), data_fim, 23) AS data_fim
                FROM rponto.dbo.ferias_pedidos
                WHERE num    = %s
                  AND estado = 'aprovado_rh'
                  AND data_fim  >= %s
                  AND data_ini  <= %s
            """, [num, d_ini_str, d_fim_str])
            for row in cursor.fetchall():
                try:
                    fi = datetime.strptime(row[0][:10], '%Y-%m-%d').date()
                    ff = datetime.strptime(row[1][:10], '%Y-%m-%d').date()
                    curr_f = fi
                    while curr_f <= ff:
                        ferias_dias.add(curr_f.isoformat())
                        curr_f += timedelta(days=1)
                except Exception:
                    pass

        # ── Calendário completo do período ────────────────────
        nomes_dia = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']
        abrev_dia = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom']

        dias           = []
        total_mins     = 0
        total_mins_not = 0
        dias_com_reg   = 0
        dias_com_erro  = 0
        dias_pausa     = 0
        dias_ferias    = 0

        curr = d_ini
        while curr <= d_fim:
            data_str = curr.strftime('%Y-%m-%d')
            reg      = regs_por_data.get(data_str)
            is_fds   = curr.weekday() >= 5
            is_ferias_dia = data_str in ferias_dias

            if reg:
                calc     = _calcular_horas_dia(reg)
                mins_not = _calcular_minutos_noturnos(
                    calc.get('pares', []), data_str
                )
                total_mins     += calc['minutos']
                total_mins_not += mins_not

                if calc['n_picagens'] > 0:
                    dias_com_reg += 1
                if calc['tem_erro']:
                    dias_com_erro += 1
                if calc['pausa_almoco']:
                    dias_pausa += 1

                # Determinar tipo do dia
                if is_ferias_dia:
                    tipo_dia = 'ferias'
                    dias_ferias += 1
                elif is_fds:
                    tipo_dia = 'fds'
                else:
                    tipo_dia = 'normal'

                dias.append({
                    'data':             data_str,
                    'dia':              curr.day,
                    'dia_semana':       nomes_dia[curr.weekday()],
                    'dia_semana_abrev': abrev_dia[curr.weekday()],
                    'is_fds':           is_fds,
                    'is_ferias':        is_ferias_dia,
                    'tipo_dia':         tipo_dia,
                    'tem_registo':      calc['n_picagens'] > 0,
                    'n_picagens':       calc['n_picagens'],
                    'mins_trabalhados': calc['minutos'],
                    'horas_fmt':        calc['horas_fmt'],
                    'tem_erro':         calc['tem_erro'],
                    'pausa_almoco':     calc['pausa_almoco'],
                    'minutos_pausa':    calc['minutos_pausa'],
                    'mins_noturnos':    mins_not,
                    'pares':            calc['pares'],
                    'picagens':         calc['picagens_raw'],
                    'tp_hor':           reg.get('tp_hor', ''),
                })
            else:
                # Sem picagem — verificar se é férias
                if is_ferias_dia:
                    tipo_dia = 'ferias'
                    dias_ferias += 1
                elif is_fds:
                    tipo_dia = 'fds'
                else:
                    tipo_dia = 'ausencia'

                dias.append({
                    'data':             data_str,
                    'dia':              curr.day,
                    'dia_semana':       nomes_dia[curr.weekday()],
                    'dia_semana_abrev': abrev_dia[curr.weekday()],
                    'is_fds':           is_fds,
                    'is_ferias':        is_ferias_dia,
                    'tipo_dia':         tipo_dia,
                    'tem_registo':      False,
                    'n_picagens':       0,
                    'mins_trabalhados': 0,
                    'horas_fmt':        '00:00',
                    'tem_erro':         False,
                    'pausa_almoco':     False,
                    'minutos_pausa':    0,
                    'mins_noturnos':    0,
                    'pares':            [],
                    'picagens':         [],
                    'tp_hor':           '',
                })

            curr += timedelta(days=1)

        dias_uteis = sum(1 for d in dias if not d['is_fds'])

        return Response({
            "colaborador": {
                "num":    num,
                "nome":   nome,
                "dep":    dep,
                "tp_hor": tp_hor
            },
            "periodo": {
                "inicio":     d_ini_str,
                "fim":        d_fim_str,
                "dias_uteis": dias_uteis
            },
            "totais": {
                "mins_trabalhados":   total_mins,
                "horas_fmt":         f"{total_mins // 60:02d}:{total_mins % 60:02d}",
                "mins_noturnos":      total_mins_not,
                "horas_noturnas_fmt": f"{total_mins_not // 60:02d}:{total_mins_not % 60:02d}",
                "dias_com_registo":  dias_com_reg,
                "dias_com_erro":     dias_com_erro,
                "dias_pausa_almoco": dias_pausa,
                "dias_ferias":       dias_ferias,
            },
            "dias":   dias,
            "status": "success"
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# ProcessamentoSalarialResumo — versão corrigida
# Adiciona dias_ferias por colaborador
# ================================================================
def ProcessamentoSalarialResumo(request, format=None):
    try:
        f        = request.data.get('filter', {})
        fdep     = f.get('fdep')
        fnum     = f.get('fnum')
        fdata    = f.get('fdata')

        if not fdata or len(fdata) < 2:
            return Response({"status": "error", "title": "Intervalo de datas obrigatório"})

        d_ini_str = str(fdata[0]).replace(">=", "").strip()[:10]
        d_fim_str = str(fdata[1]).replace("<=", "").strip()[:10]

        try:
            d_ini = datetime.strptime(d_ini_str, '%Y-%m-%d').date()
            d_fim = datetime.strptime(d_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"status": "error", "title": "Formato de data inválido"})

        if d_ini > d_fim:
            return Response({"status": "error", "title": "Data início não pode ser superior à data fim"})

        d_fim_exc_str = (d_fim + timedelta(days=1)).strftime('%Y-%m-%d')

        dias_uteis = sum(
            1 for i in range((d_fim - d_ini).days + 1)
            if (d_ini + timedelta(days=i)).weekday() < 5
        )

        where  = ["TR.dts >= %s", "TR.dts < %s"]
        params = [d_ini_str, d_fim_exc_str]

        if fdep:
            where.append("RTRIM(LTRIM(TR.dep)) = %s")
            params.append(fdep.strip())
        if fnum:
            num_fmt = fnum.strip()
            if not num_fmt.startswith('F'):
                num_fmt = f"F{num_fmt.zfill(5)}"
            where.append("TR.num = %s")
            params.append(num_fmt)

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    TR.num,
                    RTRIM(LTRIM(ISNULL(TR.dep,'')))    AS dep,
                    RTRIM(LTRIM(ISNULL(TR.tp_hor,''))) AS tp_hor,
                    CONVERT(VARCHAR(10), TR.dts, 23)   AS dts,
                    ISNULL(TR.nt, 0)                   AS nt,
                    TR.ss_01, ISNULL(TR.ty_01,'') AS ty_01,
                    TR.ss_02, ISNULL(TR.ty_02,'') AS ty_02,
                    TR.ss_03, ISNULL(TR.ty_03,'') AS ty_03,
                    TR.ss_04, ISNULL(TR.ty_04,'') AS ty_04,
                    TR.ss_05, ISNULL(TR.ty_05,'') AS ty_05,
                    TR.ss_06, ISNULL(TR.ty_06,'') AS ty_06,
                    TR.ss_07, ISNULL(TR.ty_07,'') AS ty_07,
                    TR.ss_08, ISNULL(TR.ty_08,'') AS ty_08
                FROM rponto.dbo.time_registration TR
                WHERE {' AND '.join(where)}
                ORDER BY TR.num, TR.dts
            """, params)
            cols     = [c[0] for c in cursor.description]
            registos = [dict(zip(cols, r)) for r in cursor.fetchall()]

        if not registos:
            return Response({
                "rows": [], "total": 0,
                "periodo": {"inicio": d_ini_str, "fim": d_fim_str, "dias_uteis": dias_uteis},
                "status": "success"
            })

        nums_unicos = list({r['num'] for r in registos if r.get('num')})
        nomes_dict  = _get_nomes_colaboradores(nums_unicos)

        # Pré-carregar férias de todos os colaboradores de uma vez
        # (1 query por colaborador — aceitável; alternativa seria 1 query com IN)
        ferias_por_num = {}
        for num_c in nums_unicos:
            ferias_por_num[num_c] = _get_ferias_aprovadas_set(
                num_c, d_ini_str, d_fim_str
            )

        from collections import defaultdict
        por_colab = defaultdict(list)
        for reg in registos:
            por_colab[reg['num']].append(reg)

        resultado = []

        for num_c, regs in por_colab.items():
            info_sage   = nomes_dict.get(num_c, {})
            ferias_set  = ferias_por_num.get(num_c, set())

            total_mins       = 0
            dias_com_registo = 0
            dias_com_erro    = 0
            dias_fds         = 0
            dias_pausa       = 0
            dias_ferias      = 0
            alertas          = []

            # Conjunto de datas com registo
            datas_com_reg = set()
            for reg in regs:
                dts_str = reg.get('dts', '')
                try:
                    dts_date = datetime.strptime(dts_str[:10], '%Y-%m-%d').date()
                except Exception:
                    continue

                try:
                    calc = _calcular_horas_dia(reg)
                except Exception:
                    calc = {'minutos': 0, 'horas_fmt': '00:00', 'n_picagens': 0,
                            'tem_erro': True, 'pausa_almoco': False,
                            'minutos_pausa': 0, 'pares': [], 'picagens_raw': []}

                is_fds    = dts_date.weekday() >= 5
                is_ferias = dts_str in ferias_set

                if is_fds:
                    dias_fds += 1
                if calc['n_picagens'] > 0:
                    dias_com_registo += 1
                    datas_com_reg.add(dts_str)
                if calc['tem_erro']:
                    dias_com_erro += 1
                    alertas.append({
                        'tipo': 'erro_picagem',
                        'data': dts_str[:10],
                        'msg':  f"Picagens inconsistentes ({calc['n_picagens']}) em {dts_str[:10]}"
                    })
                if calc['pausa_almoco']:
                    dias_pausa += 1
                if is_ferias and not is_fds:
                    dias_ferias += 1

                total_mins += calc['minutos']

            # Dias de férias sem picagem (contam como férias, não como falta)
            for data_f in ferias_set:
                if data_f not in datas_com_reg:
                    dias_ferias += 1

            resultado.append({
                'num':    num_c,
                'nome':   info_sage.get('nome', ''),
                'dep':    regs[0].get('dep', ''),
                'tp_hor': regs[0].get('tp_hor', ''),
                'total_mins':        total_mins,
                'total_horas_fmt':   f"{total_mins // 60:02d}:{total_mins % 60:02d}",
                'dias_uteis':        dias_uteis,
                'dias_com_registo':  dias_com_registo,
                'dias_com_erro':     dias_com_erro,
                'dias_fds':          dias_fds,
                'dias_pausa_almoco': dias_pausa,
                'dias_ferias':       dias_ferias,
                'n_alertas':         len(alertas),
                'alertas':           alertas,
            })

        resultado.sort(key=lambda x: (x.get('dep', ''), x.get('nome', '')))

        return Response({
            "rows":   resultado,
            "total":  len(resultado),
            "periodo": {"inicio": d_ini_str, "fim": d_fim_str, "dias_uteis": dias_uteis},
            "status": "success"
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


# ================================================================
# EXPORT EXCEL — Processamento Salarial
# Gera ficheiro Excel com resumo + detalhe por colaborador
# ================================================================
@api_view(['POST'])
@renderer_classes([JSONRenderer])
@permission_classes([IsAuthenticated])
def ExportProcessamentoExcel(request):
    """
    Gera Excel com:
      - Folha "Resumo" — 1 linha por colaborador
      - 1 folha por colaborador com detalhe dia a dia
    """
    try:
        filter_data = request.data.get('filter', {})
        ano  = int(filter_data.get('ano',  datetime.now().year))
        mes  = int(filter_data.get('mes',  datetime.now().month))
        fdep = filter_data.get('fdep')
        fnum = filter_data.get('fnum')

        # Reutiliza a lógica do resumo
        request_fake = type('obj', (object,), {
            'data': {
                'filter': filter_data,
                'parameters': {'method': 'ProcessamentoSalarialResumo'},
                'pagination': {'enabled': False}
            }
        })()
        resumo_resp = ProcessamentoSalarialResumo(request_fake)
        resumo_data = resumo_resp.data

        if resumo_data.get('status') != 'success':
            return HttpResponse("Erro ao gerar dados.", status=500)

        rows = resumo_data.get('rows', [])
        periodo = resumo_data.get('periodo', {})
        meses_pt = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        nome_mes = meses_pt[mes]

        wb = openpyxl.Workbook()

        # ── Estilos ──────────────────────────────────────────────
        azul_escuro   = "1E3A5F"
        azul_medio    = "2D6A9F"
        azul_claro    = "D9E8F5"
        verde_claro   = "E8F5E9"
        vermelho_claro= "FFEBEE"
        amarelo_claro = "FFF9C4"
        cinza_claro   = "F5F5F5"
        branco        = "FFFFFF"

        def header_style(cell, bg=azul_escuro, fg=branco, bold=True, size=11):
            cell.font = Font(bold=bold, color=fg, size=size)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

        def data_style(cell, bg=branco, bold=False, align='center', color='000000'):
            cell.font = Font(bold=bold, color=color, size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=align, vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='EEEEEE'),
                right=Side(style='thin', color='EEEEEE'),
                top=Side(style='thin', color='EEEEEE'),
                bottom=Side(style='thin', color='EEEEEE')
            )

        # ── Folha RESUMO ─────────────────────────────────────────
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"

        # Título
        ws_resumo.merge_cells('A1:N1')
        titulo = ws_resumo['A1']
        titulo.value = (
            f"PROCESSAMENTO SALARIAL — {nome_mes.upper()} {ano}"
        )
        titulo.font = Font(bold=True, size=14, color=branco)
        titulo.fill = PatternFill("solid", fgColor=azul_escuro)
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws_resumo.row_dimensions[1].height = 35

        # Subtítulo
        ws_resumo.merge_cells('A2:N2')
        sub = ws_resumo['A2']
        sub.value = (
            f"Período: {periodo.get('inicio')} a {periodo.get('fim')} "
            f"| Dias Úteis: {periodo.get('dias_uteis')}"
        )
        sub.font = Font(italic=True, size=10, color='555555')
        sub.alignment = Alignment(horizontal='center')
        ws_resumo.row_dimensions[2].height = 18

        # Headers
        headers_resumo = [
            'Nº', 'Nome', 'Departamento', 'Tipo Horário',
            'Dias Úteis', 'Dias c/ Registo', 'Faltas',
            'Erros Picagem', 'H. Trabalhadas', 'H. Previstas',
            'Diferença', 'H. Noturnas', 'FDS Trabalhados', 'Alertas'
        ]
        for col, h in enumerate(headers_resumo, 1):
            cell = ws_resumo.cell(row=3, column=col, value=h)
            header_style(cell)
        ws_resumo.row_dimensions[3].height = 28

        # Dados
        for row_idx, row in enumerate(rows, 4):
            row_bg = branco if (row_idx - 4) % 2 == 0 else cinza_claro
            tem_alertas = row.get('n_alertas', 0) > 0
            tem_erros   = row.get('dias_com_erro', 0) > 0
            tem_faltas  = row.get('dias_sem_registo', 0) > 0

            valores = [
                row.get('num', ''),
                row.get('nome', ''),
                row.get('dep', ''),
                row.get('tp_hor', ''),
                row.get('dias_uteis', 0),
                row.get('dias_com_registo', 0),
                row.get('dias_sem_registo', 0),
                row.get('dias_com_erro', 0),
                row.get('total_horas_fmt', ''),
                row.get('total_horas_prev_fmt', '00:00'),
                row.get('diferenca_fmt', ''),
                row.get('horas_noturnas_fmt', '00:00'),
                row.get('dias_fds_trabalhados', 0),
                row.get('n_alertas', 0),
            ]

            for col_idx, val in enumerate(valores, 1):
                cell = ws_resumo.cell(row=row_idx, column=col_idx, value=val)
                bg = row_bg

                # Destacar faltas e erros
                if col_idx == 7 and tem_faltas:
                    bg = vermelho_claro
                elif col_idx == 8 and tem_erros:
                    bg = amarelo_claro
                elif col_idx == 11:
                    dif = row.get('diferenca_mins', 0)
                    bg = verde_claro if dif >= 0 else vermelho_claro

                align = 'left' if col_idx <= 2 else 'center'
                data_style(cell, bg=bg, align=align,
                           bold=(col_idx == 1),
                           color='CC0000' if (col_idx == 14 and tem_alertas)
                                         else '000000')

        # Auto-largura
        col_widths = [12, 35, 15, 15, 10, 14, 8, 13, 14, 13, 12, 13, 14, 8]
        for i, w in enumerate(col_widths, 1):
            ws_resumo.column_dimensions[get_column_letter(i)].width = w

        ws_resumo.freeze_panes = 'A4'
        ws_resumo.auto_filter.ref = (
            f"A3:{get_column_letter(len(headers_resumo))}3"
        )

        # ── Folha por colaborador ──────────────────��──────────────
        dias_semana_bg = {
            'Sáb': "EDE7F6",
            'Dom': "EDE7F6"
        }

        for colab in rows:
            nome_sheet = (colab.get('nome', colab['num']) or colab['num'])[:28]
            for c in ['/', '\\', ':', '*', '?', '[', ']']:
                nome_sheet = nome_sheet.replace(c, '-')

            ws = wb.create_sheet(title=nome_sheet)

            # Cabeçalho do colaborador
            ws.merge_cells('A1:J1')
            c1 = ws['A1']
            c1.value = (
                f"{colab.get('nome', '')} ({colab['num']}) "
                f"— {nome_mes} {ano}"
            )
            c1.font = Font(bold=True, size=12, color=branco)
            c1.fill = PatternFill("solid", fgColor=azul_medio)
            c1.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 28

            # Info
            ws.merge_cells('A2:J2')
            c2 = ws['A2']
            c2.value = (
                f"Dept: {colab.get('dep', '—')} | "
                f"Horário: {colab.get('tp_hor', '—')} | "
                f"Total: {colab.get('total_horas_fmt', '00:00')}h | "
                f"Faltas: {colab.get('dias_sem_registo', 0)} dias"
            )
            c2.font = Font(size=9, color='444444')
            c2.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 16

            # Headers detalhe
            headers_det = [
                'Data', 'Dia', 'P1 Ent.', 'P1 Saí.',
                'P2 Ent.', 'P2 Saí.', 'P3 Ent.', 'P3 Saí.',
                'H. Trab.', 'Estado'
            ]
            for col, h in enumerate(headers_det, 1):
                cell = ws.cell(row=3, column=col, value=h)
                header_style(cell, bg=azul_medio)
            ws.row_dimensions[3].height = 24

            # Dias
            for di, dia in enumerate(colab.get('dias', []), 4):
                is_fds = dia.get('is_fds', False)
                tem_err = dia.get('tem_erro', False)
                tem_reg = dia.get('tem_registo', False)

                if is_fds:
                    row_bg = "EDE7F6"
                elif not tem_reg:
                    row_bg = "FFEBEE"
                elif tem_err:
                    row_bg = "FFF9C4"
                else:
                    row_bg = branco if di % 2 == 0 else cinza_claro

                pares = dia.get('pares', [])

                def get_par(idx, campo):
                    if idx < len(pares):
                        return pares[idx].get(campo, '')
                    return ''

                valores_dia = [
                    dia.get('data', ''),
                    dia.get('dia_semana_abrev', ''),
                    get_par(0, 'entrada'),
                    get_par(0, 'saida'),
                    get_par(1, 'entrada'),
                    get_par(1, 'saida'),
                    get_par(2, 'entrada'),
                    get_par(2, 'saida'),
                    dia.get('horas_fmt', '00:00'),
                    '⚠️ Erro' if tem_err else (
                        'FDS' if is_fds else (
                            '—' if not tem_reg else '✓'
                        )
                    )
                ]

                for col_idx, val in enumerate(valores_dia, 1):
                    cell = ws.cell(row=di, column=col_idx, value=val)
                    data_style(cell, bg=row_bg,
                               bold=(col_idx == 9),
                               align='left' if col_idx == 2 else 'center')

            # Larguras
            for i, w in enumerate([12, 6, 8, 8, 8, 8, 8, 8, 10, 10], 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            ws.freeze_panes = 'A4'

        # ── Guardar ───────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ProcessamentoSalarial_{nome_mes}_{ano}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            )
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Erro: {str(e)}", status=500)
    
#region FÉRIAS

def _dias_uteis(d_ini, d_fim):
    total = 0
    curr  = d_ini
    while curr <= d_fim:
        if curr.weekday() < 5:
            total += 1
        curr += timedelta(days=1)
    return total


def FeriasSubmeter(request, format=None):
    try:
        f        = request.data.get('filter', {})
        num      = (f.get('num') or '').strip()
        dep      = (f.get('dep') or '').strip()   # pode vir vazio — obtemos da BD
        data_ini = (f.get('data_ini') or '').strip()
        data_fim = (f.get('data_fim') or '').strip()
        obs      = (f.get('obs') or '').strip()

        # ✅ CORRIGIDO: dep NÃO é obrigatório no payload
        # — será lido da BD se vier vazio
        if not num or not data_ini or not data_fim:
            return Response({'status': 'error', 'title': 'Campos obrigatórios em falta (num, data_ini, data_fim)'})

        d_ini = datetime.strptime(data_ini, '%Y-%m-%d').date()
        d_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()

        if d_ini > d_fim:
            return Response({'status': 'error', 'title': 'Data início não pode ser superior à data fim'})

        n_dias = _dias_uteis(d_ini, d_fim)
        if n_dias == 0:
            return Response({'status': 'error', 'title': 'Nenhum dia útil no período selecionado'})

        # ── Obter dep da BD se não vier no payload ─────────────────
        if not dep:
            with connections[connMssqlName].cursor() as cursor:
                cursor.execute("""
                    SELECT TOP 1 RTRIM(LTRIM(dep)) AS dep
                    FROM rponto.dbo.time_registration
                    WHERE num = %s
                      AND dep IS NOT NULL
                      AND dep != ''
                    ORDER BY dts DESC
                """, [num])
                row = cursor.fetchone()
                dep = row[0] if row else ''

        # ── Verificar sobreposição ─────────────────────────────────
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) FROM rponto.dbo.ferias_pedidos
                WHERE num = %s
                  AND estado NOT IN ('rejeitado_chefe', 'rejeitado_rh', 'cancelado')
                  AND (
                      (data_ini <= %s AND data_fim >= %s) OR
                      (data_ini <= %s AND data_fim >= %s) OR
                      (data_ini >= %s AND data_fim <= %s)
                  )
            """, [num, data_fim, data_ini, data_ini, data_ini, d_ini, d_fim])
            overlap = cursor.fetchone()[0]

        if overlap > 0:
            return Response({
                'status': 'error',
                'title':  'Já existe um pedido de férias para esse período'
            })

        # ── Inserir ────────────────────────────────────────────────
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                INSERT INTO rponto.dbo.ferias_pedidos
                    (num, dep, data_ini, data_fim, n_dias, obs_colab, estado, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, 'pendente', GETDATE(), GETDATE())
            """, [num, dep, data_ini, data_fim, n_dias, obs or None])

        return Response({
            'status': 'success',
            'title':  f'Pedido de férias submetido com sucesso! ({n_dias} dias úteis)',
            'n_dias': n_dias
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def FeriasListar(request, format=None):
    try:
        f      = request.data.get('filter', {})
        params = request.data.get('parameters', {})
        role   = params.get('role', 'colaborador')

        # ── Parâmetros de papel ────────────────────────────────────
        is_rh      = f.get('isRH',       False)
        is_chefe   = f.get('isChefe',    False)
        deps_chefe = f.get('deps_chefe', [])

        # ── Filtros opcionais ──────────────────────────────────────
        fnum    = (f.get('num') or f.get('fnum') or '').strip()
        festado = (f.get('estado') or f.get('festado') or '').strip()
        fdata   = f.get('fdata')

        # ══════════════════════════════════════════════════════════
        # SEGURANÇA — bloquear acesso sem papel válido
        # ══════════════════════════════════════════════════════════
        # Colaborador sem num → nada
        if not is_rh and not is_chefe and not fnum:
            return Response({"rows": [], "total": 0, "status": "success",
                             "warn": "Colaborador sem número"})

        # Chefe sem deps → nada (JWT inválido ou não renovado)
        if is_chefe and not is_rh and not deps_chefe:
            return Response({"rows": [], "total": 0, "status": "success",
                             "warn": "Chefe sem departamentos configurados"})

        # ══════════════════════════════════════════════════════════
        # Construir WHERE
        # ══════════════════════════════════════════════════════════
        where = []
        qp    = []

        # ── Colaborador → só os seus pedidos ──────────────────────
        if not is_rh and not is_chefe:
            clean = fnum.replace('%', '').strip()
            if not clean.startswith('F'):
                clean = f"F{clean.zfill(5)}"
            where.append("FP.num = %s")
            qp.append(clean)

        # ── Chefe → filtra pelos departamentos dele ────────────────
        # (a tabela ferias_pedidos tem coluna dep com o dep do colaborador
        #  que foi preenchido no momento do pedido via FeriasSubmeter)
        elif is_chefe and not is_rh:
            placeholders = ', '.join(['%s'] * len(deps_chefe))
            where.append(f"RTRIM(LTRIM(FP.dep)) IN ({placeholders})")
            qp.extend([d.strip() for d in deps_chefe])

            # Filtro adicional por num se o chefe pesquisar um colaborador
            if fnum:
                clean = fnum.replace('%', '').strip()
                where.append("FP.num LIKE %s")
                qp.append(f"%{clean}%")

        # ── RH → vê tudo, pode filtrar por num se quiser ──────────
        elif is_rh:
            if fnum:
                clean = fnum.replace('%', '').strip()
                where.append("FP.num LIKE %s")
                qp.append(f"%{clean}%")

        # ── Filtro por estado ──────────────────────────────────────
        if festado:
            where.append("FP.estado = %s")
            qp.append(festado)

        # ── Filtro por data ────────────────────────────────────────
        if fdata and len(fdata) >= 2:
            d_ini = str(fdata[0]).replace('>=', '').strip()[:10]
            d_fim = str(fdata[1]).replace('<=', '').strip()[:10]
            where.append("FP.data_ini >= %s")
            qp.append(d_ini)
            where.append("FP.data_fim <= %s")
            qp.append(d_fim)

        where_sql = f"WHERE {' AND '.join(where)}" if where else ""

        # ══════════════════════════════════════════════════════════
        # Query principal
        # ══════════════════════════════════════════════════════════
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    FP.id,
                    FP.num,
                    FP.dep,
                    CONVERT(VARCHAR(10), FP.data_ini,   23)  AS data_ini,
                    CONVERT(VARCHAR(10), FP.data_fim,   23)  AS data_fim,
                    FP.n_dias,
                    FP.obs_colab,
                    FP.estado,
                    FP.chefe_num,
                    CONVERT(VARCHAR(19), FP.chefe_data, 120) AS chefe_data,
                    FP.chefe_obs,
                    FP.rh_num,
                    CONVERT(VARCHAR(19), FP.rh_data,    120) AS rh_data,
                    FP.rh_obs,
                    CONVERT(VARCHAR(19), FP.created_at, 120) AS created_at
                FROM rponto.dbo.ferias_pedidos FP
                {where_sql}
                ORDER BY FP.created_at DESC
            """, qp)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        # ── Enriquecer com nomes ───────────────────────────────────
        nums_unicos = list({r['num'] for r in rows if r.get('num')})
        nomes_dict  = _get_nomes_colaboradores(nums_unicos)

        for row in rows:
            info = nomes_dict.get(row['num'], {})
            row['nome']         = info.get('nome', '')
            row['estado_label'] = FERIAS_ESTADOS.get(row['estado'], {}).get('label', row['estado'])
            row['estado_cor']   = FERIAS_ESTADOS.get(row['estado'], {}).get('cor', 'default')

        return Response({
            'rows':   rows,
            'total':  len(rows),
            'status': 'success'
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def FeriasAprovarChefe(request, format=None):
    try:
        f         = request.data.get('filter', {})
        pedido_id = f.get('id')
        acao      = f.get('acao', '').strip()
        chefe_num = f.get('chefe_num', '').strip()
        obs       = f.get('obs', '').strip()

        if not pedido_id or acao not in ('aprovar', 'rejeitar'):
            return Response({'status': 'error', 'title': 'Parâmetros inválidos'})

        novo_estado = 'aprovado_chefe' if acao == 'aprovar' else 'rejeitado_chefe'

        with connections[connMssqlName].cursor() as cursor:
            # Verificar que o pedido está pendente
            cursor.execute(
                "SELECT estado FROM rponto.dbo.ferias_pedidos WHERE id = %s",
                [pedido_id]
            )
            row = cursor.fetchone()
            if not row:
                return Response({'status': 'error', 'title': 'Pedido não encontrado'})
            if row[0] != 'pendente':
                return Response({
                    'status': 'error',
                    'title':  f'Pedido já foi processado (estado: {row[0]})'
                })

            cursor.execute("""
                UPDATE rponto.dbo.ferias_pedidos
                SET
                    estado      = %s,
                    chefe_num   = %s,
                    chefe_data  = GETDATE(),
                    chefe_obs   = %s,
                    updated_at  = GETDATE()
                WHERE id = %s
            """, [novo_estado, chefe_num or None, obs or None, pedido_id])

        acao_label = 'aprovado' if acao == 'aprovar' else 'rejeitado'
        return Response({
            'status': 'success',
            'title':  f'Pedido {acao_label} com sucesso pelo chefe de departamento'
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def FeriasAprovarRH(request, format=None):
    try:
        f         = request.data.get('filter', {})
        pedido_id = f.get('id')
        acao      = f.get('acao', '').strip()
        rh_num    = f.get('rh_num', '').strip()
        obs       = f.get('obs', '').strip()

        if not pedido_id or acao not in ('aprovar', 'rejeitar'):
            return Response({'status': 'error', 'title': 'Parâmetros inválidos'})

        novo_estado = 'aprovado_rh' if acao == 'aprovar' else 'rejeitado_rh'

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(
                "SELECT estado FROM rponto.dbo.ferias_pedidos WHERE id = %s",
                [pedido_id]
            )
            row = cursor.fetchone()
            if not row:
                return Response({'status': 'error', 'title': 'Pedido não encontrado'})
            if row[0] != 'aprovado_chefe':
                return Response({
                    'status': 'error',
                    'title':  f'Pedido tem de estar aprovado pelo chefe primeiro (estado atual: {row[0]})'
                })

            cursor.execute("""
                UPDATE rponto.dbo.ferias_pedidos
                SET
                    estado     = %s,
                    rh_num     = %s,
                    rh_data    = GETDATE(),
                    rh_obs     = %s,
                    updated_at = GETDATE()
                WHERE id = %s
            """, [novo_estado, rh_num or None, obs or None, pedido_id])

        acao_label = 'aprovado' if acao == 'aprovar' else 'rejeitado'
        return Response({
            'status': 'success',
            'title':  f'Pedido {acao_label} com sucesso pelos Recursos Humanos'
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def FeriasCancelar(request, format=None):
    try:
        f         = request.data.get('filter', {})
        pedido_id = f.get('id')
        num       = f.get('num', '').strip()

        if not pedido_id or not num:
            return Response({'status': 'error', 'title': 'Parâmetros em falta'})

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(
                "SELECT estado, num FROM rponto.dbo.ferias_pedidos WHERE id = %s",
                [pedido_id]
            )
            row = cursor.fetchone()
            if not row:
                return Response({'status': 'error', 'title': 'Pedido não encontrado'})
            if row[1] != num:
                return Response({'status': 'error', 'title': 'Sem permissão para cancelar este pedido'})
            if row[0] not in ('pendente',):
                return Response({
                    'status': 'error',
                    'title':  'Só é possível cancelar pedidos no estado pendente'
                })

            cursor.execute("""
                UPDATE rponto.dbo.ferias_pedidos
                SET estado = 'cancelado', updated_at = GETDATE()
                WHERE id = %s
            """, [pedido_id])

        return Response({'status': 'success', 'title': 'Pedido cancelado com sucesso'})

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def FeriasSummary(request, format=None):
    try:
        p    = request.data.get('parameters', {})
        role = p.get('role', 'colaborador')
        dep  = p.get('dep', '').strip()
        num  = p.get('num', '').strip()

        with connections[connMssqlName].cursor() as cursor:
            if role == 'rh':
                cursor.execute("""
                    SELECT estado, COUNT(*) AS total
                    FROM rponto.dbo.ferias_pedidos
                    GROUP BY estado
                """)
            elif role == 'chefe':
                cursor.execute("""
                    SELECT estado, COUNT(*) AS total
                    FROM rponto.dbo.ferias_pedidos
                    WHERE RTRIM(LTRIM(dep)) = %s
                    GROUP BY estado
                """, [dep])
            else:
                cursor.execute("""
                    SELECT estado, COUNT(*) AS total
                    FROM rponto.dbo.ferias_pedidos
                    WHERE num = %s
                    GROUP BY estado
                """, [num])

            rows = cursor.fetchall()

        contagens = {r[0]: r[1] for r in rows}

        pendentes_chefe = contagens.get('pendente', 0)
        pendentes_rh    = contagens.get('aprovado_chefe', 0)

        return Response({
            'contagens':      contagens,
            'pendentes_chefe': pendentes_chefe,
            'pendentes_rh':    pendentes_rh,
            'status':         'success'
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def HorariosConfigList(request, format=None):
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT
                    t.sigla                                          AS tp_hor,
                    t.nome,
                    CONVERT(VARCHAR(5), t.hora_inicio, 108)         AS hora_inicio,
                    CONVERT(VARCHAR(5), t.hora_fim,    108)         AS hora_fim,
                    t.cor_hex,
                    t.tipo,
                    t.descricao                                      AS turno_descricao,
                    CONVERT(VARCHAR(5), hc.entrada_esperada, 108)   AS entrada_esperada,
                    CONVERT(VARCHAR(5), hc.saida_esperada,   108)   AS saida_esperada,
                    ISNULL(hc.tolerancia_minutos, 15)               AS tolerancia_minutos,
                    CONVERT(VARCHAR(5), hc.almoco_inicio, 108)      AS almoco_inicio,
                    CONVERT(VARCHAR(5), hc.almoco_fim,    108)      AS almoco_fim,
                    hc.horas_dia,
                    hc.dias_semana,
                    hc.descricao                                     AS config_descricao,
                    (
                        SELECT COUNT(DISTINCT tr.num)
                        FROM rponto.dbo.time_registration tr
                        WHERE RTRIM(LTRIM(tr.tp_hor)) = t.sigla
                    )                                                AS n_colaboradores
                FROM rponto.dbo.turnos t
                LEFT JOIN rponto.dbo.horarios_config hc ON hc.tp_hor = t.sigla
                ORDER BY
                    CASE t.tipo
                        WHEN 'trabalho' THEN 0
                        WHEN 'reforco'  THEN 1
                        ELSE 2
                    END,
                    t.sigla
            """)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return Response({"rows": rows, "total": len(rows), "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


def HorariosConfigSave(request, format=None):
    try:
        f = request.data.get('filter', {})

        tp_hor             = f.get('tp_hor', '').strip().upper()
        entrada_esperada   = f.get('entrada_esperada')   or None
        saida_esperada     = f.get('saida_esperada')     or None
        tolerancia         = f.get('tolerancia_minutos', 15)
        almoco_inicio      = f.get('almoco_inicio')      or None
        almoco_fim         = f.get('almoco_fim')         or None
        horas_dia          = f.get('horas_dia')
        dias_semana        = f.get('dias_semana', '1,2,3,4,5')
        descricao          = f.get('descricao', '')

        if not tp_hor:
            return Response({"status": "error", "title": "tp_hor obrigatório"})

        # Calcular horas_dia automaticamente se não fornecido
        if horas_dia is None and entrada_esperada and saida_esperada:
            try:
                h_ini   = datetime.strptime(entrada_esperada[:5], '%H:%M')
                h_fim   = datetime.strptime(saida_esperada[:5],   '%H:%M')
                minutos = int((h_fim - h_ini).total_seconds() / 60)
                if minutos < 0:
                    minutos += 1440  # overnight (ex: TAR 16:00→00:00)
                if almoco_inicio and almoco_fim:
                    a_ini   = datetime.strptime(almoco_inicio[:5], '%H:%M')
                    a_fim   = datetime.strptime(almoco_fim[:5],    '%H:%M')
                    minutos -= int((a_fim - a_ini).total_seconds() / 60)
                horas_dia = round(minutos / 60, 2)
            except Exception:
                horas_dia = 8.0

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM rponto.dbo.horarios_config WHERE tp_hor = %s",
                [tp_hor]
            )
            existe = cursor.fetchone()[0] > 0

            if existe:
                cursor.execute("""
                    UPDATE rponto.dbo.horarios_config SET
                        entrada_esperada   = %s,
                        saida_esperada     = %s,
                        tolerancia_minutos = %s,
                        almoco_inicio      = %s,
                        almoco_fim         = %s,
                        horas_dia          = %s,
                        dias_semana        = %s,
                        descricao          = %s
                    WHERE tp_hor = %s
                """, [
                    entrada_esperada, saida_esperada, tolerancia,
                    almoco_inicio, almoco_fim, horas_dia,
                    dias_semana, descricao, tp_hor
                ])
                msg = f"Horário {tp_hor} atualizado com sucesso"
            else:
                cursor.execute("""
                    INSERT INTO rponto.dbo.horarios_config
                        (tp_hor, entrada_esperada, saida_esperada,
                         tolerancia_minutos, almoco_inicio, almoco_fim,
                         horas_dia, dias_semana, descricao)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    tp_hor, entrada_esperada, saida_esperada,
                    tolerancia, almoco_inicio, almoco_fim,
                    horas_dia, dias_semana, descricao
                ])
                msg = f"Horário {tp_hor} criado com sucesso"

        return Response({"status": "success", "title": msg})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


def EsquemasTeoricos(request, format=None):
    try:
        esquema_tipo = request.data.get('filter', {}).get('esquema_tipo', 'GER')

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT
                    et.esquema_tipo,
                    et.esquema_numero,
                    et.dia_semana_iso,
                    CASE et.dia_semana_iso
                        WHEN 1 THEN 'Segunda'
                        WHEN 2 THEN 'Terça'
                        WHEN 3 THEN 'Quarta'
                        WHEN 4 THEN 'Quinta'
                        WHEN 5 THEN 'Sexta'
                        WHEN 6 THEN 'Sábado'
                        WHEN 7 THEN 'Domingo'
                    END                                         AS dia_nome,
                    et.turno_sigla,
                    t.nome                                      AS turno_nome,
                    t.tipo                                      AS turno_tipo,
                    t.cor_hex,
                    CONVERT(VARCHAR(5), t.hora_inicio,   108)  AS hora_inicio,
                    CONVERT(VARCHAR(5), t.hora_fim,      108)  AS hora_fim,
                    CONVERT(VARCHAR(5), hc.almoco_inicio,108)  AS almoco_inicio,
                    CONVERT(VARCHAR(5), hc.almoco_fim,   108)  AS almoco_fim,
                    hc.tolerancia_minutos,
                    hc.horas_dia
                FROM rponto.dbo.esquemas_teoricos   et
                LEFT JOIN rponto.dbo.turnos          t  ON t.sigla   = et.turno_sigla
                LEFT JOIN rponto.dbo.horarios_config hc ON hc.tp_hor = et.turno_sigla
                WHERE et.esquema_tipo = %s
                ORDER BY et.esquema_numero, et.dia_semana_iso
            """, [esquema_tipo])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return Response({"rows": rows, "total": len(rows), "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})


def ColaboradoresPorHorario(request, format=None):
    try:
        tp_hor = request.data.get('filter', {}).get('tp_hor', '').strip()

        where  = []
        params = []
        if tp_hor:
            where.append("RTRIM(LTRIM(tr.tp_hor)) = %s")
            params.append(tp_hor)

        where_sql = f"WHERE {' AND '.join(where)}" if where else ""

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(f"""
                SELECT DISTINCT
                    tr.num,
                    RTRIM(LTRIM(ISNULL(tr.dep,'')))    AS dep,
                    RTRIM(LTRIM(ISNULL(tr.tp_hor,''))) AS tp_hor,
                    t.nome                             AS turno_nome,
                    t.cor_hex,
                    t.tipo                             AS turno_tipo
                FROM rponto.dbo.time_registration tr
                LEFT JOIN rponto.dbo.turnos t ON t.sigla = RTRIM(LTRIM(tr.tp_hor))
                {where_sql}
                ORDER BY tr.dep, tr.num
            """, params)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        # Enriquecer com nomes SAGE
        nums  = [r['num'] for r in rows if r.get('num')]
        nomes = _get_nomes_colaboradores(nums)
        for row in rows:
            row['nome'] = nomes.get(row['num'], {}).get('nome', '')

        # Agrupar por tp_hor
        from collections import defaultdict
        grupos = defaultdict(list)
        for row in rows:
            grupos[row['tp_hor']].append(row)

        resultado = [
            {
                'tp_hor':        k,
                'turno_nome':    v[0].get('turno_nome', ''),
                'cor_hex':       v[0].get('cor_hex', ''),
                'turno_tipo':    v[0].get('turno_tipo', ''),
                'colaboradores': v,
                'total':         len(v)
            }
            for k, v in sorted(grupos.items())
        ]

        return Response({
            "rows":   resultado,
            "total":  len(resultado),
            "status": "success"
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})

def _get_ferias_aprovadas_set(num, data_inicio_str, data_fim_str):
    """
    Devolve um set de strings 'YYYY-MM-DD' com todos os dias úteis
    de férias aprovadas pelo RH para o colaborador no intervalo dado.
    """
    ferias_set = set()
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT data_ini, data_fim
                FROM rponto.dbo.ferias_pedidos
                WHERE num    = %s
                  AND estado = 'aprovado_rh'
                  AND data_fim >= %s
                  AND data_ini <= %s
            """, [num, data_inicio_str, data_fim_str])
            for row in cursor.fetchall():
                d_ini_f = row[0] if isinstance(row[0], date) \
                          else datetime.strptime(str(row[0])[:10], '%Y-%m-%d').date()
                d_fim_f = row[1] if isinstance(row[1], date) \
                          else datetime.strptime(str(row[1])[:10], '%Y-%m-%d').date()
                curr_f = d_ini_f
                while curr_f <= d_fim_f:
                    if curr_f.weekday() < 5:          # só dias úteis
                        ferias_set.add(curr_f.isoformat())
                    curr_f += timedelta(days=1)
    except Exception as e:
        print(f"[WARN] _get_ferias_aprovadas_set: {e}")
    return ferias_set


DEP_PARA_ESQUEMA = {
    'DPROD': 'Laboracao_Continua',
    'DARM':  'Armazem',
    'DAF':   'Armazem',
    'DARQ':  'Armazem',
    # acrescentar outros conforme os resultados da query debug_08
}
ESQUEMA_DEFAULT = 'Laboracao_Continua'   # fallback se dep não mapeado
EQUIPAS_ROTATIVAS = ('A', 'B', 'C', 'D', 'E')


def _dep_to_esquema(dep):
    """Converte um dep no seu esquema_tipo no ciclo_laboracao."""
    if not dep:
        return ESQUEMA_DEFAULT
    return DEP_PARA_ESQUEMA.get(dep.strip().upper(), ESQUEMA_DEFAULT)


def GetTurnosColaborador(request, format=None):
    try:
        parameters  = request.data.get('parameters', {})
        filter_data = request.data.get('filter', {})

        # ── Datas ──────────────────────────────────────────────────
        data_inicio_str = parameters.get('data_inicio')
        data_fim_str    = parameters.get('data_fim')

        if not data_inicio_str:
            dt_inicio       = datetime.now().replace(day=1)
            data_inicio_str = dt_inicio.strftime('%Y-%m-%d')
        else:
            dt_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')

        if not data_fim_str:
            proximo_mes  = (dt_inicio.replace(day=28) + timedelta(days=4)).replace(day=1)
            data_fim_str = (proximo_mes - timedelta(days=1)).strftime('%Y-%m-%d')

        # ── Papel do utilizador ────────────────────────────────────
        num_auth   = (filter_data.get('num') or '').strip()
        is_rh      = bool(filter_data.get('isRH',    False))
        is_admin   = bool(filter_data.get('isAdmin', False))
        is_chefe   = bool(filter_data.get('isChefe', False))
        deps_chefe = filter_data.get('deps_chefe', []) or []

        # ── Obter tp_hor e dep do colaborador ──────────────────────
        tp_hor_auth = None
        dep_auth    = None

        if num_auth:
            with connections[connMssqlName].cursor() as cursor:
                cursor.execute("""
                    SELECT TOP 1
                        RTRIM(LTRIM(tp_hor)) AS tp_hor,
                        RTRIM(LTRIM(dep))    AS dep
                    FROM rponto.dbo.time_registration
                    WHERE num    = %s
                      AND tp_hor IS NOT NULL AND tp_hor != ''
                      AND dep    IS NOT NULL AND dep    != ''
                    ORDER BY dts DESC
                """, [num_auth])
                row = cursor.fetchone()
                if row:
                    tp_hor_auth = row[0]
                    dep_auth    = row[1]

        print(f"[GetTurnos] num={num_auth} tp_hor={tp_hor_auth} dep={dep_auth} "
              f"isRH={is_rh} isChefe={is_chefe} deps_chefe={deps_chefe}")

        # ══════════════════════════════════════════════════════════
        # FÉRIAS APROVADAS — vista pessoal
        # ══════════════════════════════════════════════════════════
        ferias_colab = set()
        if num_auth and not is_rh and not is_admin and not is_chefe:
            with connections[connMssqlName].cursor() as cursor:
                cursor.execute("""
                    SELECT
                        CONVERT(VARCHAR(10), data_ini, 23),
                        CONVERT(VARCHAR(10), data_fim, 23)
                    FROM rponto.dbo.ferias_pedidos
                    WHERE num    = %s
                      AND estado = 'aprovado_rh'
                      AND data_fim  >= %s
                      AND data_ini  <= %s
                """, [num_auth, data_inicio_str, data_fim_str])
                for row in cursor.fetchall():
                    try:
                        fi = datetime.strptime(row[0][:10], '%Y-%m-%d').date()
                        ff = datetime.strptime(row[1][:10], '%Y-%m-%d').date()
                        c  = fi
                        while c <= ff:
                            ferias_colab.add(c.isoformat())
                            c += timedelta(days=1)
                    except Exception:
                        pass

        # ══════════════════════════════════════════════════════════
        # TROCAS APROVADAS — vista pessoal
        # ══════════════════════════════════════════════════════════
        trocas_colab = {}  # data_str -> novo_turno_sigla
        if num_auth and not is_rh and not is_admin and not is_chefe:
            with connections[connMssqlName].cursor() as cursor:
                # Como requerente: na data_req fica com turno_dest,
                #                  na data_dest fica com turno_req
                cursor.execute("""
                    SELECT
                        CONVERT(VARCHAR(10), data_req,  23),
                        CONVERT(VARCHAR(10), data_dest, 23),
                        turno_req, turno_dest
                    FROM rponto.dbo.trocas_turno
                    WHERE num_req = %s AND estado = 'aprovado_chefe'
                      AND (data_req  BETWEEN %s AND %s
                        OR data_dest BETWEEN %s AND %s)
                """, [num_auth,
                      data_inicio_str, data_fim_str,
                      data_inicio_str, data_fim_str])
                for r in cursor.fetchall():
                    dr, dd, treq, tdest = r
                    if dr: trocas_colab[dr] = tdest
                    if dd: trocas_colab[dd] = treq

                # Como destino
                cursor.execute("""
                    SELECT
                        CONVERT(VARCHAR(10), data_req,  23),
                        CONVERT(VARCHAR(10), data_dest, 23),
                        turno_req, turno_dest
                    FROM rponto.dbo.trocas_turno
                    WHERE num_dest = %s AND estado = 'aprovado_chefe'
                      AND (data_req  BETWEEN %s AND %s
                        OR data_dest BETWEEN %s AND %s)
                """, [num_auth,
                      data_inicio_str, data_fim_str,
                      data_inicio_str, data_fim_str])
                for r in cursor.fetchall():
                    dr, dd, treq, tdest = r
                    if dd: trocas_colab[dd] = treq
                    if dr: trocas_colab[dr] = tdest

        # ══════════════════════════════════════════════════════════
        # CARREGAR INFO DE TODOS OS TURNOS (cache)
        # ══════════════════════════════════════════════════════════
        turno_cache = {}
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT sigla, nome,
                       CONVERT(VARCHAR(5), hora_inicio, 108),
                       CONVERT(VARCHAR(5), hora_fim,    108),
                       cor_hex, tipo
                FROM rponto.dbo.turnos
            """)
            for row in cursor.fetchall():
                turno_cache[row[0]] = {
                    'nome':        row[1],
                    'hora_inicio': row[2],
                    'hora_fim':    row[3],
                    'cor_hex':     row[4] or '#94a3b8',
                    'tipo':        row[5],
                }

        # ══════════════════════════════════════════════════════════
        # DETERMINAR EQUIPAS + ESQUEMAS A CARREGAR
        # ══════════════════════════════════════════════════════════
        #
        # Regras:
        #   RH/Admin   → todas as equipas, TODOS os esquemas (vista geral)
        #   Chefe       → equipas do(s) dep(s), esquemas do(s) dep(s)
        #   Colaborador → só a sua equipa, só o seu esquema
        #
        # Estrutura de consulta:
        #   lista de (equipa, esquema) pares a carregar

        if is_rh or is_admin:
            # Pegar todos os pares (equipa, esquema) distintos existentes
            with connections[connMssqlName].cursor() as cursor:
                cursor.execute("""
                    SELECT DISTINCT equipa_letra, esquema_tipo
                    FROM rponto.dbo.ciclo_laboracao
                    ORDER BY equipa_letra, esquema_tipo
                """)
                pares_eq_esq = [(r[0], r[1]) for r in cursor.fetchall()]

        elif is_chefe:
            # Para cada dep do chefe, descobrir as equipas e esquemas
            pares_eq_esq = []
            if deps_chefe:
                with connections[connMssqlName].cursor() as cursor:
                    placeholders = ','.join(['%s'] * len(deps_chefe))
                    cursor.execute(f"""
                        SELECT DISTINCT
                            RTRIM(LTRIM(tp_hor)) AS equipa
                        FROM rponto.dbo.time_registration
                        WHERE RTRIM(LTRIM(dep)) IN ({placeholders})
                          AND tp_hor IN {str(EQUIPAS_ROTATIVAS).replace(',)', ')')}
                    """, deps_chefe)
                    equipas_chefe = [r[0] for r in cursor.fetchall()]

                # Para cada equipa do chefe, determinar os esquemas dos seus deps
                esquemas_chefe = list({
                    _dep_to_esquema(d) for d in deps_chefe
                })

                with connections[connMssqlName].cursor() as cursor:
                    if equipas_chefe and esquemas_chefe:
                        eq_ph  = ','.join(["'%s'" % e for e in equipas_chefe])
                        esq_ph = ','.join(["'%s'" % e for e in esquemas_chefe])
                        cursor.execute(f"""
                            SELECT DISTINCT equipa_letra, esquema_tipo
                            FROM rponto.dbo.ciclo_laboracao
                            WHERE equipa_letra IN ({eq_ph})
                              AND esquema_tipo  IN ({esq_ph})
                            ORDER BY equipa_letra, esquema_tipo
                        """)
                        pares_eq_esq = [(r[0], r[1]) for r in cursor.fetchall()]

        else:
            # Colaborador normal
            if tp_hor_auth in EQUIPAS_ROTATIVAS:
                esquema_colab = _dep_to_esquema(dep_auth)
                # Verificar se existe exactamente este par na BD
                with connections[connMssqlName].cursor() as cursor:
                    cursor.execute("""
                        SELECT TOP 1 equipa_letra, esquema_tipo
                        FROM rponto.dbo.ciclo_laboracao
                        WHERE equipa_letra = %s AND esquema_tipo = %s
                    """, [tp_hor_auth, esquema_colab])
                    row = cursor.fetchone()
                    if row:
                        pares_eq_esq = [(tp_hor_auth, esquema_colab)]
                    else:
                        # Fallback: primeiro esquema disponível para esta equipa
                        cursor.execute("""
                            SELECT TOP 1 equipa_letra, esquema_tipo
                            FROM rponto.dbo.ciclo_laboracao
                            WHERE equipa_letra = %s
                            ORDER BY esquema_tipo
                        """, [tp_hor_auth])
                        row2 = cursor.fetchone()
                        pares_eq_esq = [(tp_hor_auth, row2[1])] if row2 else []
            else:
                pares_eq_esq = []

        # ══════════════════════════════════════════════════════════
        # CARREGAR O CICLO — uma query por par (equipa, esquema)
        # ou uma query única filtrando IN com AND por par
        # ══════════════════════════════════════════════════════════
        # escalas_rotativos: { data_str: [ {equipa, esquema, turno_sigla, ...} ] }
        escalas_rotativos = {}

        if pares_eq_esq:
            # Construir um filtro WHERE eficiente:
            # WHERE (equipa_letra = 'A' AND esquema_tipo = 'Laboracao_Continua')
            #    OR (equipa_letra = 'B' AND esquema_tipo = 'Laboracao_Continua')
            #    OR (equipa_letra = 'B' AND esquema_tipo = 'Armazem')
            #    ...
            par_conditions = " OR ".join([
                f"(c.equipa_letra = '{eq}' AND c.esquema_tipo = '{esq}')"
                for eq, esq in pares_eq_esq
            ])

            query_ciclo = f"""
                SET DATEFIRST 1;
                SELECT
                    FORMAT(DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01'),
                           'yyyy-MM-dd')                       AS data,
                    c.equipa_letra                             AS equipa,
                    c.esquema_tipo                             AS esquema,
                    c.turno_sigla,
                    COALESCE(t.nome, c.turno_sigla)            AS turno_nome,
                    CONVERT(VARCHAR(5), t.hora_inicio, 108)    AS hora_inicio,
                    CONVERT(VARCHAR(5), t.hora_fim,    108)    AS hora_fim,
                    ISNULL(t.cor_hex, '#94a3b8')               AS cor_hex,
                    CASE WHEN c.ordem_rotacao IN (1,358,359,365) THEN 1 ELSE 0
                    END                                        AS is_feriado,
                    h.name                                     AS nome_feriado
                FROM rponto.dbo.ciclo_laboracao c
                LEFT JOIN rponto.dbo.turnos t
                    ON t.sigla = c.turno_sigla
                LEFT JOIN rponto.dbo.holidays h
                    ON h.holiday_date =
                       DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01')
                WHERE DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01')
                        BETWEEN '{data_inicio_str}' AND '{data_fim_str}'
                  AND ({par_conditions})
                ORDER BY data, c.equipa_letra, c.esquema_tipo
            """

            with connections[connMssqlName].cursor() as cursor:
                cursor.execute(query_ciclo)
                cols_ciclo = [c[0] for c in cursor.description]
                for row in cursor.fetchall():
                    r = dict(zip(cols_ciclo, row))
                    d = r['data']
                    if d not in escalas_rotativos:
                        escalas_rotativos[d] = []
                    escalas_rotativos[d].append({
                        'equipa':       r['equipa'],
                        'esquema':      r['esquema'],
                        'turno_sigla':  r['turno_sigla'],
                        'turno_nome':   r['turno_nome'],
                        'hora_inicio':  r['hora_inicio'],
                        'hora_fim':     r['hora_fim'],
                        'cor_hex':      r['cor_hex'],
                        'is_feriado':   bool(r['is_feriado']),
                        'nome_feriado': r['nome_feriado'],
                        'is_troca':     False,
                        'is_ferias':    False,
                    })

        # ══════════════════════════════════════════════════════════
        # COLABORADORES GER (horário geral fixo)
        # ══════════════════════════════════════════════════════════
        colaboradores_ger = []

        carregar_ger = (
            is_rh or is_admin or is_chefe
            or (not is_rh and not is_admin and not is_chefe
                and tp_hor_auth not in EQUIPAS_ROTATIVAS)
        )

        if carregar_ger:
            where_ger  = ["RTRIM(LTRIM(tr.tp_hor)) NOT IN ('A','B','C','D','E')"]
            params_ger = []

            if not is_rh and not is_admin and not is_chefe and num_auth:
                where_ger.append("tr.num = %s")
                params_ger.append(num_auth)
            elif deps_chefe:
                placeholders = ','.join(['%s'] * len(deps_chefe))
                where_ger.append(f"RTRIM(LTRIM(tr.dep)) IN ({placeholders})")
                params_ger.extend(deps_chefe)
            # RH/Admin: sem restrição por dep

            with connections[connMssqlName].cursor() as cursor:
                cursor.execute(f"""
                    SELECT DISTINCT
                        tr.num,
                        RTRIM(LTRIM(tr.dep))    AS dep,
                        RTRIM(LTRIM(tr.tp_hor)) AS tp_hor
                    FROM rponto.dbo.time_registration tr
                    WHERE {' AND '.join(where_ger)}
                """, params_ger)
                cols_ger = [c[0] for c in cursor.description]
                colaboradores_ger = [dict(zip(cols_ger, r))
                                     for r in cursor.fetchall()]

        # Obter nomes dos colaboradores GER via SAGE
        nums_ger  = [c['num'] for c in colaboradores_ger]
        nomes_ger = {}
        if nums_ger:
            try:
                nomes_ger = _get_nomes_colaboradores(nums_ger)
            except Exception:
                pass
        for c in colaboradores_ger:
            c['nome'] = nomes_ger.get(c['num'], {}).get('nome', '')

        # ══════════════════════════════════════════════════════════
        # MONTAR CALENDÁRIO DIA A DIA
        # ══════════════════════════════════════════════════════════
        nomes_dia = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']
        dt_current = dt_inicio
        dt_end     = datetime.strptime(data_fim_str, '%Y-%m-%d')
        escalas    = []

        while dt_current <= dt_end:
            data_str    = dt_current.strftime('%Y-%m-%d')
            iso_weekday = dt_current.isoweekday()   # 1=Seg ... 7=Dom
            is_fds      = iso_weekday >= 6

            dia_obj = {
                'data':        data_str,
                'dia_semana':  nomes_dia[dt_current.weekday()],
                'iso_weekday': iso_weekday,
                'is_fds':      is_fds,
                # equipas: lista de turnos rotativos para esta data
                # (pode ter múltiplos esquemas/equipas na vista geral)
                'equipas':     list(escalas_rotativos.get(data_str, [])),
                # ger: lista de colaboradores com horário geral
                'ger':         [],
                'tem_ferias':  False,
                'tem_troca':   False,
            }

            # GER — turno calculado pelo dia da semana
            for colab in colaboradores_ger:
                turno_sigla = 'GER' if iso_weekday <= 5 else 'DSC'
                ti          = turno_cache.get(turno_sigla, {})
                dia_obj['ger'].append({
                    'num':         colab['num'],
                    'nome':        colab.get('nome', ''),
                    'dep':         colab['dep'],
                    'turno_sigla': turno_sigla,
                    'turno_nome':  ti.get('nome', 'Geral' if iso_weekday <= 5 else 'Descanso'),
                    'hora_inicio': ti.get('hora_inicio') if iso_weekday <= 5 else None,
                    'hora_fim':    ti.get('hora_fim')    if iso_weekday <= 5 else None,
                    'cor_hex':     ti.get('cor_hex', '#F59E0B') if iso_weekday <= 5 else '#DC2626',
                    'is_troca':    False,
                    'is_ferias':   False,
                })

            # ── OVERRIDE FÉRIAS (só vista pessoal) ────────────────
            if data_str in ferias_colab:
                dia_obj['tem_ferias'] = True
                for eq in dia_obj['equipas']:
                    eq['turno_sigla_original'] = eq['turno_sigla']
                    eq['turno_nome_original']  = eq['turno_nome']
                    eq['turno_sigla'] = 'FER'
                    eq['turno_nome']  = 'Férias'
                    eq['cor_hex']     = '#9333EA'
                    eq['is_ferias']   = True
                for g in dia_obj['ger']:
                    g['turno_sigla_original'] = g['turno_sigla']
                    g['turno_sigla'] = 'FER'
                    g['turno_nome']  = 'Férias'
                    g['cor_hex']     = '#9333EA'
                    g['is_ferias']   = True

            # ── OVERRIDE TROCAS (só vista pessoal) ────────────────
            if data_str in trocas_colab:
                dia_obj['tem_troca']     = True
                novo_sigla               = trocas_colab[data_str]
                info_novo                = turno_cache.get(novo_sigla, {})
                for eq in dia_obj['equipas']:
                    if not eq.get('is_ferias'):
                        eq['turno_sigla_original'] = eq['turno_sigla']
                        eq['turno_sigla'] = novo_sigla
                        eq['turno_nome']  = info_novo.get('nome', novo_sigla)
                        eq['hora_inicio'] = info_novo.get('hora_inicio')
                        eq['hora_fim']    = info_novo.get('hora_fim')
                        eq['cor_hex']     = info_novo.get('cor_hex', '#94a3b8')
                        eq['is_troca']    = True
                for g in dia_obj['ger']:
                    if not g.get('is_ferias'):
                        g['turno_sigla_original'] = g['turno_sigla']
                        g['turno_sigla'] = novo_sigla
                        g['turno_nome']  = info_novo.get('nome', novo_sigla)
                        g['hora_inicio'] = info_novo.get('hora_inicio')
                        g['hora_fim']    = info_novo.get('hora_fim')
                        g['cor_hex']     = info_novo.get('cor_hex', '#94a3b8')
                        g['is_troca']    = True

            escalas.append(dia_obj)
            dt_current += timedelta(days=1)

        # ── Metadados do utilizador ────────────────────────────────
        user_info = {
            'num':          num_auth,
            'tp_hor':       tp_hor_auth,
            'dep':          dep_auth,
            'esquema':      _dep_to_esquema(dep_auth) if dep_auth else ESQUEMA_DEFAULT,
            'pares_carregados': pares_eq_esq,
            'is_rh':        is_rh,
            'is_admin':     is_admin,
            'is_chefe':     is_chefe,
            'deps_chefe':   deps_chefe,
            'role':         ('rh' if (is_rh or is_admin)
                             else 'chefe' if is_chefe
                             else 'colaborador'),
        }

        print(f"[GetTurnos] OK escalas={len(escalas)} "
              f"pares={pares_eq_esq} ferias={len(ferias_colab)} "
              f"trocas={len(trocas_colab)} ger={len(colaboradores_ger)}")

        return Response({
            'success':     True,
            'data_inicio': data_inicio_str,
            'data_fim':    data_fim_str,
            'total_dias':  len(escalas),
            'escalas':     escalas,
            'user_info':   user_info,
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'success': False, 'error': str(e)}, status=500)




def DashboardResumo(request, format=None):
    filter_data  = request.data.get('filter', {})
    num          = filter_data.get('num')
    is_rh        = filter_data.get('isRH', False)
    is_chefe     = filter_data.get('isChefe', False)
    deps_chefe   = filter_data.get('deps_chefe', [])

    conn_rponto = connections[connMssqlName].cursor()
    conn_sage   = connections[connSage100cName].cursor()

    try:
        data = {}
        today = date.today()

        # ── Colaborador normal ─────────────────────────────────────────
        if not is_rh and not is_chefe and num:
            # Picagens hoje
            conn_rponto.execute(
                "SELECT nt FROM rponto.dbo.time_registration WHERE num = ? AND dts = ?",
                [num, today.strftime('%Y-%m-%d')]
            )
            row = conn_rponto.fetchone()
            data['picagens_hoje'] = row[0] if row else 0

            # Férias
            conn_rponto.execute(
                "SELECT COUNT(*) FROM rponto.dbo.ferias WHERE num = ? AND estado = 'pendente'",
                [num]
            )
            data['ferias_pendentes'] = (conn_rponto.fetchone() or [0])[0]

            conn_rponto.execute(
                "SELECT COUNT(*) FROM rponto.dbo.ferias WHERE num = ? AND estado = 'aprovado_rh'",
                [num]
            )
            data['ferias_aprovadas'] = (conn_rponto.fetchone() or [0])[0]

            # Justificações
            conn_rponto.execute(
                "SELECT COUNT(*) FROM rponto.dbo.justificacoes WHERE num = ? AND status IN (0, 1)",
                [num]
            )
            data['justificacoes_pendentes'] = (conn_rponto.fetchone() or [0])[0]

        # ── Chefe ─────────────────────────────────────────────────────
        elif is_chefe and not is_rh and deps_chefe:
            deps_placeholder = ', '.join(['?' for _ in deps_chefe])

            # Justificações pendentes do dep
            conn_rponto.execute(
                f"SELECT COUNT(*) FROM rponto.dbo.justificacoes WHERE status = 0 AND RTRIM(LTRIM(dep_codigo)) IN ({deps_placeholder})",
                deps_chefe
            )
            data['justificacoes_pendentes'] = (conn_rponto.fetchone() or [0])[0]

            # Últimas 3 justificações pendentes
            conn_rponto.execute(
                f"""SELECT TOP 3 j.id, j.num, j.dt_inicio, j.dt_fim, j.motivo_descricao, j.dep_codigo
                    FROM rponto.dbo.justificacoes j
                    WHERE j.status = 0 AND RTRIM(LTRIM(j.dep_codigo)) IN ({deps_placeholder})
                    ORDER BY j.dt_submissao DESC""",
                deps_chefe
            )
            rows_just = conn_rponto.fetchall()
            
            # Enriquecer com nomes
            just_recentes = []
            for r in (rows_just or []):
                conn_sage.execute(
                    "SELECT NOME FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE NFUNC = ? AND DEMITIDO = 0", [r[1]]
                )
                nome_row = conn_sage.fetchone()
                just_recentes.append({
                    'num': r[1],
                    'nome': nome_row[0] if nome_row else r[1],
                    'dt_inicio': str(r[2]) if r[2] else None,
                    'motivo': r[4] or '—',
                    'dep': r[5] or '—',
                })
            data['justificacoes_recentes'] = just_recentes

            # Férias pendentes
            conn_rponto.execute(
                f"SELECT COUNT(*) FROM rponto.dbo.ferias WHERE estado = 'pendente' AND RTRIM(LTRIM(dep)) IN ({deps_placeholder})",
                deps_chefe
            )
            data['ferias_pendentes'] = (conn_rponto.fetchone() or [0])[0]

            # Últimas 3 férias pendentes
            conn_rponto.execute(
                f"""SELECT TOP 3 f.id, f.num, f.data_ini, f.data_fim, f.n_dias
                    FROM rponto.dbo.ferias f
                    WHERE f.estado = 'pendente' AND RTRIM(LTRIM(f.dep)) IN ({deps_placeholder})
                    ORDER BY f.created_at DESC""",
                deps_chefe
            )
            rows_fer = conn_rponto.fetchall()
            ferias_recentes = []
            for r in (rows_fer or []):
                conn_sage.execute(
                    "SELECT NOME FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE NFUNC = ? AND DEMITIDO = 0", [r[1]]
                )
                nome_row = conn_sage.fetchone()
                ferias_recentes.append({
                    'num': r[1],
                    'nome': nome_row[0] if nome_row else r[1],
                    'data_ini': str(r[2]) if r[2] else None,
                    'data_fim': str(r[3]) if r[3] else None,
                    'n_dias': r[4] or 0,
                })
            data['ferias_recentes'] = ferias_recentes

            # Nº colaboradores do dep
            conn_sage.execute(
                f"SELECT COUNT(*) FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE DEPARTAMENTO IN ({deps_placeholder}) AND DEMITIDO = 0",
                deps_chefe
            )
            data['n_colaboradores'] = (conn_sage.fetchone() or [0])[0]

            # Trocas pendentes
            try:
                conn_rponto.execute(
                    f"SELECT COUNT(*) FROM rponto.dbo.trocas_turno WHERE estado IN ('pendente','aceite_dest') AND (RTRIM(LTRIM(dep_req)) IN ({deps_placeholder}) OR RTRIM(LTRIM(dep_dest)) IN ({deps_placeholder}))",
                    deps_chefe + deps_chefe
                )
                data['trocas_pendentes'] = (conn_rponto.fetchone() or [0])[0]
            except Exception:
                data['trocas_pendentes'] = 0

        # ── RH ─────────────────────────────────────────────────────────
        elif is_rh:
            # Justificações a aguardar RH
            conn_rponto.execute("SELECT COUNT(*) FROM rponto.dbo.justificacoes WHERE status = 1")
            data['justificacoes_rh'] = (conn_rponto.fetchone() or [0])[0]

            # Justificações chefes pendentes
            conn_rponto.execute("SELECT COUNT(*) FROM rponto.dbo.justificacoes WHERE status = 0")
            data['justificacoes_chefes'] = (conn_rponto.fetchone() or [0])[0]

            # Férias aguardam RH
            conn_rponto.execute("SELECT COUNT(*) FROM rponto.dbo.ferias WHERE estado = 'aprovado_chefe'")
            data['ferias_rh'] = (conn_rponto.fetchone() or [0])[0]

            # Total pendentes
            data['total_pendentes'] = data['justificacoes_rh'] + data['ferias_rh']

            # Total colaboradores
            conn_sage.execute("SELECT COUNT(*) FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE DEMITIDO = 0")
            data['total_colaboradores'] = (conn_sage.fetchone() or [0])[0]

            # Picagens hoje
            conn_rponto.execute(
                "SELECT COUNT(*) FROM rponto.dbo.time_registration WHERE dts = ?",
                [today.strftime('%Y-%m-%d')]
            )
            data['picagens_hoje'] = (conn_rponto.fetchone() or [0])[0]

            # Últimas justificações a aguardar RH
            conn_rponto.execute("""
                SELECT TOP 5 j.id, j.num, j.dt_inicio, j.motivo_descricao, j.dep_codigo
                FROM rponto.dbo.justificacoes j
                WHERE j.status = 1
                ORDER BY j.dt_submissao DESC
            """)
            rows_rh = conn_rponto.fetchall()
            just_rh = []
            for r in (rows_rh or []):
                conn_sage.execute(
                    "SELECT NOME FROM TRIMTEK_1GEP.dbo.FUNC1 WHERE NFUNC = ? AND DEMITIDO = 0", [r[1]]
                )
                nome_row = conn_sage.fetchone()
                just_rh.append({
                    'num': r[1],
                    'nome': nome_row[0] if nome_row else r[1],
                    'dt_inicio': str(r[2]) if r[2] else None,
                    'motivo': r[3] or '—',
                    'dep': r[4] or '—',
                })
            data['justificacoes_recentes_rh'] = just_rh

        return Response({'status': 'success', **data})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})
    finally:
        conn_rponto.close()
        conn_sage.close()


#region trocas da prod
# ================================================================
# TROCAS DE TURNO
# ================================================================

def _get_equipa_colaborador(num):
    """Devolve tp_hor e dep do colaborador (últimos dados)."""
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT TOP 1
                    RTRIM(LTRIM(tp_hor)) AS tp_hor,
                    RTRIM(LTRIM(dep))    AS dep
                FROM rponto.dbo.time_registration
                WHERE num = %s
                  AND tp_hor IS NOT NULL AND tp_hor != ''
                  AND dep    IS NOT NULL AND dep    != ''
                ORDER BY dts DESC
            """, [num])
            row = cursor.fetchone()
            if row:
                return row[0], row[1]   # tp_hor, dep
    except Exception:
        pass
    return None, None


def _get_turno_equipa_em_data(equipa, data_str):
    """
    Devolve o turno_sigla que a equipa tem na data indicada,
    usando ciclo_laboracao (a mesma lógica do GetTurnosColaborador).
    """
    try:
        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT TOP 1 c.turno_sigla
                FROM rponto.dbo.ciclo_laboracao c
                WHERE c.equipa_letra = %s
                  AND DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01') = %s
            """, [equipa, data_str])
            row = cursor.fetchone()
            if row:
                return row[0]
    except Exception:
        pass
    return None


def TrocasSolicitar(request, format=None):
    try:
        filter_data = request.data.get('filter', {})

        num_req   = filter_data.get('num_req')
        num_dest  = filter_data.get('num_dest')
        data_req  = filter_data.get('data_req')
        data_dest = filter_data.get('data_dest')
        obs       = filter_data.get('obs_req', '')

        if not all([num_req, num_dest, data_req, data_dest]):
            return Response({
                "status": "error",
                "title": "Campos obrigatórios: num_req, num_dest, data_req, data_dest"
            })

        if num_req == num_dest:
            return Response({"status": "error", "title": "Os dois colaboradores têm de ser diferentes."})

        def _get_turno_para_data(cursor, num, data_str):
            try:
                cursor.execute("""
                    SELECT TOP 1 c.turno_sigla
                    FROM rponto.dbo.ciclo_laboracao c
                    JOIN rponto.dbo.time_registration tr
                        ON tr.tp_hor = c.equipa_letra
                    WHERE tr.num = %s
                      AND DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01') = %s
                      AND tr.dep IS NOT NULL
                    ORDER BY tr.dts DESC
                """, [num, data_str])
                row = cursor.fetchone()
                if row and row[0]:
                    return row[0].strip()
            except Exception:
                pass
            return 'DSC'

        with connections[connMssqlName].cursor() as cursor:
            # Dep + equipa do requerente
            cursor.execute("""
                SELECT TOP 1 dep, tp_hor
                FROM rponto.dbo.time_registration
                WHERE num = %s AND dep IS NOT NULL
                ORDER BY dts DESC
            """, [num_req])
            row = cursor.fetchone()
            dep_req    = row[0].strip() if row else None
            equipa_req = row[1].strip() if row and row[1] else ''

            # Dep + equipa do destino
            cursor.execute("""
                SELECT TOP 1 dep, tp_hor
                FROM rponto.dbo.time_registration
                WHERE num = %s AND dep IS NOT NULL
                ORDER BY dts DESC
            """, [num_dest])
            row = cursor.fetchone()
            dep_dest    = row[0].strip() if row else None
            equipa_dest = row[1].strip() if row and row[1] else ''

            # Validar mesmo departamento
            if dep_req and dep_dest and dep_req != dep_dest:
                return Response({
                    "status": "error",
                    "title": f"Os colaboradores são de departamentos diferentes ({dep_req} / {dep_dest})."
                })

            # Obter turnos
            turno_req  = _get_turno_para_data(cursor, num_req,  data_req)
            turno_dest = _get_turno_para_data(cursor, num_dest, data_dest)

            cursor.execute("""
                INSERT INTO rponto.dbo.trocas_turno
                    (num_req, num_dest, dep_req, dep_dest,
                     equipa_req, equipa_dest,
                     turno_req, turno_dest,
                     data_req, data_dest,
                     obs_req, estado, created_at)
                OUTPUT INSERTED.id
                VALUES (%s, %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, 'pendente', GETDATE())
            """, [
                num_req, num_dest, dep_req, dep_dest,
                equipa_req, equipa_dest,
                turno_req, turno_dest,
                data_req, data_dest,
                obs
            ])
            row = cursor.fetchone()
            new_id = row[0] if row else None

        return Response({
            "status": "success",
            "title": f"Pedido de troca enviado com sucesso! (ID #{new_id})",
            "id": new_id
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})



def TrocasSolicitarChefe(request, format=None):
    try:
        filter_data = request.data.get('filter', {})

        chefe_num = filter_data.get('chefe_num')
        num_req   = filter_data.get('num_req')
        num_dest  = filter_data.get('num_dest')
        data_req  = filter_data.get('data_req')
        data_dest = filter_data.get('data_dest')
        obs       = filter_data.get('obs_req', '')

        if not all([chefe_num, num_req, num_dest, data_req, data_dest]):
            return Response({
                "status": "error",
                "title": "Campos obrigatórios em falta: chefe_num, num_req, num_dest, data_req, data_dest"
            })

        if num_req == num_dest:
            return Response({"status": "error", "title": "Os dois colaboradores têm de ser diferentes."})

        deps_do_chefe = _get_deps_do_chefe(chefe_num)

        def _get_turno_para_data(cursor, num, data_str):
            """
            Tenta obter o turno (sigla) do colaborador para uma data específica.
            Consulta ciclo_laboracao via tp_hor (equipa) e a data.
            Devolve a sigla do turno ou 'DSC' como fallback.
            """
            try:
                cursor.execute("""
                    SELECT TOP 1 c.turno_sigla
                    FROM rponto.dbo.ciclo_laboracao c
                    JOIN rponto.dbo.time_registration tr
                        ON tr.tp_hor = c.equipa_letra
                    WHERE tr.num = %s
                      AND DATEADD(DAY, c.ordem_rotacao - 1, '2026-01-01') = %s
                      AND tr.dep IS NOT NULL
                    ORDER BY tr.dts DESC
                """, [num, data_str])
                row = cursor.fetchone()
                if row and row[0]:
                    return row[0].strip()
            except Exception:
                pass
            return 'DSC'  # fallback seguro

        with connections[connMssqlName].cursor() as cursor:

            # ── Dep + equipa (tp_hor) de cada colaborador ──────────
            cursor.execute("""
                SELECT TOP 1 dep, tp_hor
                FROM rponto.dbo.time_registration
                WHERE num = %s AND dep IS NOT NULL
                ORDER BY dts DESC
            """, [num_req])
            row = cursor.fetchone()
            dep_req    = row[0].strip() if row else None
            equipa_req = row[1].strip() if row and row[1] else None

            cursor.execute("""
                SELECT TOP 1 dep, tp_hor
                FROM rponto.dbo.time_registration
                WHERE num = %s AND dep IS NOT NULL
                ORDER BY dts DESC
            """, [num_dest])
            row = cursor.fetchone()
            dep_dest    = row[0].strip() if row else None
            equipa_dest = row[1].strip() if row and row[1] else None

            # ── Validar departamentos ──────────────────────────────
            if dep_req and dep_req not in deps_do_chefe:
                return Response({
                    "status": "error",
                    "title": f"Colaborador {num_req} não pertence ao seu departamento ({dep_req})."
                })
            if dep_dest and dep_dest not in deps_do_chefe:
                return Response({
                    "status": "error",
                    "title": f"Colaborador {num_dest} não pertence ao seu departamento ({dep_dest})."
                })

            equipa_req  = equipa_req  or ''
            equipa_dest = equipa_dest or ''

            # ── Obter turno de cada colaborador na data indicada ───
            turno_req  = _get_turno_para_data(cursor, num_req,  data_req)
            turno_dest = _get_turno_para_data(cursor, num_dest, data_dest)

            # ── Inserir troca já aprovada pelo chefe ───────────────
            cursor.execute("""
                INSERT INTO rponto.dbo.trocas_turno
                    (num_req, num_dest, dep_req, dep_dest,
                     equipa_req, equipa_dest,
                     turno_req, turno_dest,
                     data_req, data_dest,
                     obs_req, estado,
                     dest_data, chefe_num, chefe_data, chefe_obs,
                     created_at)
                OUTPUT INSERTED.id
                VALUES (%s, %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, 'aprovado_chefe',
                        GETDATE(), %s, GETDATE(), %s,
                        GETDATE())
            """, [
                num_req, num_dest, dep_req, dep_dest,
                equipa_req, equipa_dest,
                turno_req, turno_dest,
                data_req, data_dest,
                obs,
                chefe_num, obs
            ])
            row = cursor.fetchone()
            new_id = row[0] if row else None

        return Response({
            "status": "success",
            "title": f"Troca registada e aprovada com sucesso! (ID #{new_id})",
            "id": new_id
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})



def TrocasListar(request, format=None):
    """
    Lista trocas filtradas por papel do utilizador.
    filter: { num, isRH, isChefe, deps_chefe, festado }
    """
    try:
        f          = request.data.get('filter', {})
        num        = (f.get('num') or '').strip()
        is_rh      = bool(f.get('isRH',    False))
        is_chefe   = bool(f.get('isChefe', False))
        deps_chefe = f.get('deps_chefe', []) or []
        festado    = (f.get('festado') or '').strip()

        where  = ["1=1"]
        params = []

        if is_rh:
            # RH vê tudo — só conhecimento
            pass
        elif is_chefe and deps_chefe:
            # Chefe vê as do seu dep
            placeholders = ','.join(['%s'] * len(deps_chefe))
            where.append(f"(RTRIM(LTRIM(dep_req)) IN ({placeholders}) "
                         f"OR RTRIM(LTRIM(dep_dest)) IN ({placeholders}))")
            params.extend([d.strip() for d in deps_chefe])
            params.extend([d.strip() for d in deps_chefe])
        elif num:
            # Colaborador vê as suas
            where.append("(num_req = %s OR num_dest = %s)")
            params.extend([num, num])
        else:
            return Response({"rows": [], "total": 0, "status": "success"})

        if festado:
            where.append("estado = %s")
            params.append(festado)

        where_sql = " AND ".join(where)

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    id,
                    num_req,  dep_req,  equipa_req,
                    CONVERT(VARCHAR(10), data_req,  23) AS data_req,
                    turno_req,
                    num_dest, dep_dest, equipa_dest,
                    CONVERT(VARCHAR(10), data_dest, 23) AS data_dest,
                    turno_dest,
                    estado,
                    CONVERT(VARCHAR(19), dest_data,  120) AS dest_data,
                    dest_obs,
                    chefe_num,
                    CONVERT(VARCHAR(19), chefe_data, 120) AS chefe_data,
                    chefe_obs,
                    obs_req,
                    CONVERT(VARCHAR(19), created_at, 120) AS created_at
                FROM rponto.dbo.trocas_turno
                WHERE {where_sql}
                ORDER BY created_at DESC
            """, params)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        # Enriquecer com nomes
        nums_set = set()
        for r in rows:
            if r.get('num_req'):  nums_set.add(r['num_req'])
            if r.get('num_dest'): nums_set.add(r['num_dest'])
        nomes = _get_nomes_colaboradores(list(nums_set)) if nums_set else {}

        for r in rows:
            r['nome_req']  = nomes.get(r['num_req'],  {}).get('nome', '')
            r['nome_dest'] = nomes.get(r['num_dest'], {}).get('nome', '')

        return Response({'rows': rows, 'total': len(rows), 'status': 'success'})

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def TrocasResponderDest(request, format=None):
    """
    Colaborador destino aceita ou recusa a troca.
    filter: { id, num_dest, acao: 'aceitar'|'recusar', obs }
    """
    try:
        f        = request.data.get('filter', {})
        troca_id = f.get('id')
        num_dest = (f.get('num_dest') or '').strip()
        acao     = (f.get('acao') or '').strip()
        obs      = (f.get('obs') or '').strip()

        if not troca_id or acao not in ('aceitar', 'recusar'):
            return Response({'status': 'error', 'title': 'Parâmetros inválidos'})

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT id, num_dest, estado
                FROM rponto.dbo.trocas_turno
                WHERE id = %s
            """, [troca_id])
            row = cursor.fetchone()

            if not row:
                return Response({'status': 'error', 'title': 'Troca não encontrada'})

            _, dest_bd, estado_atual = row

            if dest_bd.strip() != num_dest:
                return Response({'status': 'error', 'title': 'Sem permissão'})

            if estado_atual != 'pendente':
                return Response({
                    'status': 'error',
                    'title':  f'Troca já foi processada (estado: {estado_atual})'
                })

            novo_estado = 'aceite_dest' if acao == 'aceitar' else 'rejeitado_dest'

            cursor.execute("""
                UPDATE rponto.dbo.trocas_turno
                SET estado     = %s,
                    dest_data  = GETDATE(),
                    dest_obs   = %s,
                    updated_at = GETDATE()
                WHERE id = %s
            """, [novo_estado, obs or None, troca_id])

        label = 'aceite' if acao == 'aceitar' else 'recusado'
        return Response({
            'status': 'success',
            'title':  f'Pedido de troca {label}. '
                      + ('Aguarda aprovação do chefe.' if acao == 'aceitar' else '')
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def TrocasAprovarChefe(request, format=None):
    """
    Chefe aprova ou rejeita uma troca já aceite pelo destino.
    filter: { id, chefe_num, acao: 'aprovar'|'rejeitar', obs }
    """
    try:
        f         = request.data.get('filter', {})
        troca_id  = f.get('id')
        chefe_num = (f.get('chefe_num') or '').strip()
        acao      = (f.get('acao') or '').strip()
        obs       = (f.get('obs') or '').strip()

        if not troca_id or acao not in ('aprovar', 'rejeitar'):
            return Response({'status': 'error', 'title': 'Parâmetros inválidos'})

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT id, dep_req, estado
                FROM rponto.dbo.trocas_turno
                WHERE id = %s
            """, [troca_id])
            row = cursor.fetchone()

            if not row:
                return Response({'status': 'error', 'title': 'Troca não encontrada'})

            _, dep_req, estado_atual = row

            if estado_atual != 'aceite_dest':
                return Response({
                    'status': 'error',
                    'title':  'Troca ainda não foi aceite pelo colaborador destino '
                              f'(estado: {estado_atual})'
                })

            # Verificar que o chefe é do departamento
            deps_do_chefe = _get_deps_do_chefe(chefe_num)
            if dep_req and dep_req.strip() not in deps_do_chefe:
                return Response({
                    'status': 'error',
                    'title':  'Sem permissão para aprovar trocas deste departamento'
                })

            novo_estado = 'aprovado_chefe' if acao == 'aprovar' else 'rejeitado_chefe'

            cursor.execute("""
                UPDATE rponto.dbo.trocas_turno
                SET estado     = %s,
                    chefe_num  = %s,
                    chefe_data = GETDATE(),
                    chefe_obs  = %s,
                    updated_at = GETDATE()
                WHERE id = %s
            """, [novo_estado, chefe_num, obs or None, troca_id])

        label = 'aprovada' if acao == 'aprovar' else 'rejeitada'
        return Response({
            'status': 'success',
            'title':  f'Troca {label} com sucesso.'
                      + (' Os horários foram actualizados.' if acao == 'aprovar' else '')
        })

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})


def TrocasCancelar(request, format=None):
    """
    Requerente cancela o pedido (só se ainda estiver pendente).
    filter: { id, num_req }
    """
    try:
        f        = request.data.get('filter', {})
        troca_id = f.get('id')
        num_req  = (f.get('num_req') or '').strip()

        if not troca_id or not num_req:
            return Response({'status': 'error', 'title': 'id e num_req obrigatórios'})

        with connections[connMssqlName].cursor() as cursor:
            cursor.execute("""
                SELECT num_req, estado FROM rponto.dbo.trocas_turno WHERE id = %s
            """, [troca_id])
            row = cursor.fetchone()

            if not row:
                return Response({'status': 'error', 'title': 'Troca não encontrada'})
            if row[0].strip() != num_req:
                return Response({'status': 'error', 'title': 'Sem permissão'})
            if row[1] not in ('pendente',):
                return Response({
                    'status': 'error',
                    'title':  'Só é possível cancelar trocas pendentes'
                })

            cursor.execute("""
                UPDATE rponto.dbo.trocas_turno
                SET estado = 'cancelado', updated_at = GETDATE()
                WHERE id = %s
            """, [troca_id])

        return Response({'status': 'success', 'title': 'Troca cancelada'})

    except Exception as e:
        traceback.print_exc()
        return Response({'status': 'error', 'title': str(e)})



def ColaboradoresDepartamento(request, format=None):
    try:
        filter_data = request.data.get('filter', {})
        is_rh       = filter_data.get('isRH',       False)
        is_chefe    = filter_data.get('isChefe',    False)
        deps_chefe  = filter_data.get('deps_chefe', [])
        dep_filter = None

        if is_rh:
            dep_filter = None  
        elif is_chefe and deps_chefe:
            dep_filter = [d.strip() for d in deps_chefe if d.strip()]
        else:
            num = filter_data.get('num') or filter_data.get('_num')
            if num:
                try:
                    with connections[connMssqlName].cursor() as cur:
                        cur.execute("""
                            SELECT TOP 1 RTRIM(LTRIM(dep))
                            FROM rponto.dbo.time_registration
                            WHERE num = %s
                              AND dep IS NOT NULL
                              AND RTRIM(LTRIM(dep)) != ''
                            ORDER BY dts DESC
                        """, [num])
                        row = cur.fetchone()
                        if row and row[0]:
                            dep_filter = [row[0]]
                except Exception as e:
                    print(f"[WARN] ColaboradoresDepartamento dep lookup: {e}")

        where  = ["F1.DEMITIDO = 0"]
        params = []

        if dep_filter:
            placeholders = ', '.join(['%s'] * len(dep_filter))
            where.append(f"RTRIM(LTRIM(F1.DEPARTAMENTO)) IN ({placeholders})")
            params.extend(dep_filter)

        where_sql = "WHERE " + " AND ".join(where)

        with connections[connSage100cName].cursor() as cursor:
            cursor.execute(f"""
                SELECT
                    RTRIM(LTRIM(F1.NFUNC))        AS num,
                    RTRIM(LTRIM(F1.NOME))         AS nome,
                    RTRIM(LTRIM(F1.DEPARTAMENTO)) AS dep
                FROM TRIMTEK_1GEP.dbo.FUNC1 F1
                {where_sql}
                ORDER BY F1.NOME
            """, params)
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        return Response({"rows": rows, "total": len(rows), "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"status": "error", "title": str(e)})

#region Notificações

def NotificacoesCount(request, format=None):
    """
    Contagem de notificações não lidas.

    Casos cobertos:
      A) Colaborador é destino de troca pendente de resposta        → estado='pendente'
      B) Chefe DPROD: trocas aceites aguardam aprovação             → estado='aceite_dest'
      C) Colaborador (req OU dest) de troca aprovada pelo chefe     → estado='aprovado_chefe'
         (o chefe registou a troca directamente — ambos são notificados)
    """
    try:
        filter_data = request.data.get('filter', {})
        num         = (filter_data.get('num') or '').strip()
        is_chefe    = bool(filter_data.get('isChefe', False))
        deps_chefe  = [str(d).strip() for d in (filter_data.get('deps_chefe') or [])]

        if not num:
            return Response({"count": 0, "status": "success"})

        count = 0

        with connections[connMssqlName].cursor() as cursor:

            # A) Trocas pendentes onde o colaborador é destino
            cursor.execute("""
                SELECT COUNT(*)
                FROM rponto.dbo.trocas_turno
                WHERE num_dest = %s
                  AND estado   = 'pendente'
            """, [num])
            count += cursor.fetchone()[0]

            # B) Chefe DPROD: trocas aceites aguardam aprovação do chefe
            if is_chefe and 'DPROD' in deps_chefe:
                cursor.execute("""
                    SELECT COUNT(*)
                    FROM rponto.dbo.trocas_turno
                    WHERE estado = 'aceite_dest'
                      AND (RTRIM(LTRIM(dep_req))  = 'DPROD'
                        OR RTRIM(LTRIM(dep_dest)) = 'DPROD')
                """)
                count += cursor.fetchone()[0]

            # C) Trocas aprovadas pelo chefe onde o colaborador é requerente OU destino
            #    e ainda não foram vistas (created_at nas últimas 7 dias como proxy de "nova")
            cursor.execute("""
                SELECT COUNT(*)
                FROM rponto.dbo.trocas_turno
                WHERE estado      = 'aprovado_chefe'
                  AND (num_req  = %s OR num_dest = %s)
                  AND chefe_data >= DATEADD(DAY, -7, GETDATE())
            """, [num, num])
            count += cursor.fetchone()[0]

        return Response({"count": count, "status": "success"})

    except Exception as e:
        traceback.print_exc()
        return Response({"count": 0, "status": "error", "title": str(e)})


def NotificacoesList(request, format=None):
    """
    Lista detalhada de notificações.

    Casos cobertos:
      A) Colaborador é destino de troca pendente de resposta
      B) Chefe DPROD: trocas aceites aguardam aprovação
      C) Colaborador (req OU dest) de troca aprovada pelo chefe (últimos 7 dias)
    """
    try:
        filter_data = request.data.get('filter', {})
        num         = (filter_data.get('num') or '').strip()
        is_chefe    = bool(filter_data.get('isChefe', False))
        deps_chefe  = [str(d).strip() for d in (filter_data.get('deps_chefe') or [])]

        if not num:
            return Response({"rows": [], "total": 0, "status": "success"})

        notificacoes = []

        with connections[connMssqlName].cursor() as cursor:

            # ── A) Trocas pendentes — colaborador é destino ────────
            cursor.execute("""
                SELECT
                    id,
                    num_req,
                    CONVERT(VARCHAR(10), data_req,  23) AS data_req,
                    CONVERT(VARCHAR(10), data_dest, 23) AS data_dest,
                    turno_req,
                    turno_dest,
                    obs_req,
                    CONVERT(VARCHAR(19), created_at, 120) AS created_at
                FROM rponto.dbo.trocas_turno
                WHERE num_dest = %s
                  AND estado   = 'pendente'
                ORDER BY created_at DESC
            """, [num])
            cols = [c[0] for c in cursor.description]
            trocas_pendentes = [dict(zip(cols, r)) for r in cursor.fetchall()]

            nums_enriquecer = list({r['num_req'] for r in trocas_pendentes if r.get('num_req')})
            nomes = _get_nomes_colaboradores(nums_enriquecer) if nums_enriquecer else {}

            for t in trocas_pendentes:
                nome_req = nomes.get(t['num_req'], {}).get('nome', t['num_req'])
                notificacoes.append({
                    "id":       f"troca_{t['id']}",
                    "troca_id": t['id'],
                    "tipo":     "troca_pendente",
                    "titulo":   "Pedido de Troca de Turno",
                    "mensagem": (
                        f"{nome_req} quer trocar consigo: "
                        f"o seu turno de {t['data_dest']} ({t['turno_dest']}) "
                        f"pelo turno de {t['data_req']} ({t['turno_req']})"
                    ),
                    "data":     t['created_at'],
                    "lida":     False,
                    "acao":     "responder",
                })

            # ── B) Chefe DPROD: trocas aceites aguardam aprovação ──
            if is_chefe and 'DPROD' in deps_chefe:
                cursor.execute("""
                    SELECT
                        id,
                        num_req,
                        num_dest,
                        CONVERT(VARCHAR(10), data_req,  23) AS data_req,
                        CONVERT(VARCHAR(10), data_dest, 23) AS data_dest,
                        turno_req,
                        turno_dest,
                        CONVERT(VARCHAR(19), created_at, 120) AS created_at
                    FROM rponto.dbo.trocas_turno
                    WHERE estado = 'aceite_dest'
                      AND (RTRIM(LTRIM(dep_req))  = 'DPROD'
                        OR RTRIM(LTRIM(dep_dest)) = 'DPROD')
                    ORDER BY created_at DESC
                """)
                cols_b = [c[0] for c in cursor.description]
                trocas_aceites = [dict(zip(cols_b, r)) for r in cursor.fetchall()]

                nums_all = list(
                    {r['num_req'] for r in trocas_aceites} |
                    {r['num_dest'] for r in trocas_aceites}
                )
                nomes_all = _get_nomes_colaboradores(nums_all) if nums_all else {}

                for t in trocas_aceites:
                    nome_req  = nomes_all.get(t['num_req'],  {}).get('nome', t['num_req'])
                    nome_dest = nomes_all.get(t['num_dest'], {}).get('nome', t['num_dest'])
                    notificacoes.append({
                        "id":       f"chefe_troca_{t['id']}",
                        "troca_id": t['id'],
                        "tipo":     "troca_aprovar",
                        "titulo":   "Troca Aguarda Aprovação",
                        "mensagem": (
                            f"Troca entre {nome_req} ({t['data_req']}) "
                            f"e {nome_dest} ({t['data_dest']}) foi aceite — aguarda a sua aprovação"
                        ),
                        "data":     t['created_at'],
                        "lida":     False,
                        "acao":     "aprovar",
                    })

            # ── C) Trocas aprovadas pelo chefe (últimos 7 dias) ────
            #    O colaborador (req ou dest) é notificado do resultado
            cursor.execute("""
                SELECT
                    id,
                    num_req,
                    num_dest,
                    CONVERT(VARCHAR(10), data_req,  23) AS data_req,
                    CONVERT(VARCHAR(10), data_dest, 23) AS data_dest,
                    turno_req,
                    turno_dest,
                    chefe_num,
                    CONVERT(VARCHAR(19), chefe_data, 120) AS chefe_data
                FROM rponto.dbo.trocas_turno
                WHERE estado      = 'aprovado_chefe'
                  AND (num_req  = %s OR num_dest = %s)
                  AND chefe_data >= DATEADD(DAY, -7, GETDATE())
                ORDER BY chefe_data DESC
            """, [num, num])
            cols_c = [c[0] for c in cursor.description]
            trocas_aprovadas = [dict(zip(cols_c, r)) for r in cursor.fetchall()]

            nums_c = list(
                {r['num_req'] for r in trocas_aprovadas} |
                {r['num_dest'] for r in trocas_aprovadas}
            )
            nomes_c = _get_nomes_colaboradores(nums_c) if nums_c else {}

            for t in trocas_aprovadas:
                nome_req  = nomes_c.get(t['num_req'],  {}).get('nome', t['num_req'])
                nome_dest = nomes_c.get(t['num_dest'], {}).get('nome', t['num_dest'])

                # Mensagem personalizada conforme o papel do utilizador na troca
                if t['num_req'] == num:
                    mensagem = (
                        f"A sua troca com {nome_dest} foi aprovada pelo chefe: "
                        f"{t['data_req']} ({t['turno_req']}) ↔ {t['data_dest']} ({t['turno_dest']})"
                    )
                else:
                    mensagem = (
                        f"A troca com {nome_req} foi aprovada pelo chefe: "
                        f"{t['data_req']} ({t['turno_req']}) ↔ {t['data_dest']} ({t['turno_dest']})"
                    )

                notificacoes.append({
                    "id":       f"aprovada_{t['id']}_{num}",
                    "troca_id": t['id'],
                    "tipo":     "troca_aprovada",
                    "titulo":   "Troca de Turno Aprovada ✓",
                    "mensagem": mensagem,
                    "data":     t['chefe_data'],
                    "lida":     False,
                    "acao":     "info",
                })

        # Ordenar por data desc
        notificacoes.sort(key=lambda x: x['data'] or '', reverse=True)

        return Response({
            "rows":   notificacoes,
            "total":  len(notificacoes),
            "status": "success"
        })

    except Exception as e:
        traceback.print_exc()
        return Response({"rows": [], "total": 0, "status": "error", "title": str(e)})


def MarcarNotificacaoLida(request, format=None):
    """
    Stub — as notificações derivam do estado das trocas.
    O frontend usa este endpoint para limpar o badge localmente
    após o utilizador abrir o dropdown.
    """
    return Response({"status": "success"})