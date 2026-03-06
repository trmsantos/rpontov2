import base64
from operator import eq
from pyexpat import features
import re
from typing import List
from wsgiref.util import FileWrapper
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework.views import APIView
from django.http import Http404, request, StreamingHttpResponse
from rest_framework.response import Response
from django.http.response import HttpResponse
from django.http import FileResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from rest_framework import status
import mimetypes
from datetime import datetime, timedelta
from io import BytesIO
import zipfile
# import cups
import os, tempfile

from pyodbc import Cursor, Error, connect, lowercase
from datetime import datetime
from django.http.response import JsonResponse
from rest_framework.decorators import api_view, authentication_classes, permission_classes, renderer_classes
from django.db import connections, transaction
from support.database import encloseColumn, Filters, DBSql, TypeDml, fetchall, Check
from support.myUtils import  ifNull

from rest_framework.renderers import JSONRenderer, MultiPartRenderer, BaseRenderer
from rest_framework.utils import encoders, json
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
import collections
import hmac
import hashlib
import math
from django.core.files.storage import FileSystemStorage
from sistema.settings.appSettings import AppSettings
import time
import requests
import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

def max_length(worksheet):
    column_widths = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                if cell.column in column_widths:
                    column_widths[cell.column_letter] = max(column_widths[cell.column_letter], len(str(cell.value)))
                else:
                    column_widths[cell.column_letter] = len(str(cell.value))
    for i, column_width in column_widths.items():
        worksheet.column_dimensions[i].width = column_width + 10


def exportRunxlslist(req,dbi,conn):
    try:
        response = dbi.executeSimpleList(req["sql"], conn, req["data"])
        df = pd.DataFrame(response["rows"])
        cols = list(filter(lambda x: x in df.columns.values.tolist(), req["cols"].keys()))
        rcols = []
        for v in cols:
            if ("format" in req["cols"][v]):
                if req["cols"][v]["format"]=='0':
                    df[v] = df[v].fillna(0).astype(int)
                elif req["cols"][v]["format"]=='0.0':
                    df[v] = df[v].fillna(0).astype(float).round(1)
                elif req["cols"][v]["format"]=='0.00':
                    df[v] = df[v].fillna(0).astype(float).round(2)
                elif req["cols"][v]["format"]=='0.000':
                    df[v] = df[v].fillna(0).astype(float).round(3)
                elif req["cols"][v]["format"]=='0.0000':
                    df[v] = df[v].fillna(0).astype(float).round(4)
            if "title" in req["cols"][v]:
                df.rename(columns = {v:req["cols"][v]["title"]}, inplace = True)
                rcols.append(req["cols"][v]["title"])
            else:
                rcols.append(v)

        #print(req["cols"].keys())
        #df = df.loc[:, df.columns.isin(req["cols"].keys())]
        # Create a new Excel file and add the dataframe to it
        #book = Workbook()
        excel_file  = BytesIO()
        #writer = pd.ExcelWriter(sio, engine='openpyxl') 
        #writer.book = book
        df.to_excel(excel_file, engine='openpyxl', index=False,columns=rcols)
        
        excel_file.seek(0)
        wb = load_workbook(filename=excel_file)
        ws = wb.active
        header = ws[1]
        
        # for cell in header:
        #     if "format" in req["cols"][cell.value]:
        #         print(cell.column_letter)
        #         print(req["cols"][cell.value]["format"])
        #         cell.number_format = req["cols"][cell.value]["format"]
        #     cell.value = req["cols"][cell.value]["title"]
        #     cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type = "solid")


        excel_file.close()
        excel_file  = BytesIO()
        wb.save(excel_file)
        #sio.seek(0)
        #wb = load_workbook(sio)
        #ws = wb.active


        # Get the worksheet and format the header
        #worksheet = writer.book['Sheet1']
        #header = worksheet[1] 
        #for cell in header:
        #    cell.fill = openpyxl.styles.PatternFill(start_color="0066CC", end_color="0066CC", fill_type = "solid")

        # Save the changes
        #x  = BytesIO()
        #wb.save(excel_file)
        excel_file.seek(0)
        response =  HttpResponse(excel_file.read(), content_type="application/ms-excel")
        response['Content-Disposition'] = 'attachment; filename=list.xlsx'
        return response
        

    except Exception as error:
        print(str(error))
        return Response({"status": "error", "title": str(error)})

def exportRunxlstemplatelist(req,dbi,conn):
    try:
        print(req["sql"])
        response = dbi.executeSimpleList(req["sql"], conn, req["data"])
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "a",zipfile.ZIP_DEFLATED, False) as zip_file:
            for v in req.get("filter").get("months"):
                df = pd.DataFrame([r for r in response["rows"] if r['m'] == int(v.get("key"))])
                df = df[df.columns.intersection(list(req["cols"].keys()))]
                book = load_workbook(f'./doc_templates/{req.get("template")}')
                sheet = book['dataplan']
                sheet.append(list(req["cols"].keys()))
                # Append the new data to the existing sheet
                for index,row in df.iterrows():
                    r = []
                    for c in list(req["cols"].keys()):
                        r.append(row[c])
                    sheet.append(r)
                # Save the modified file to a buffer
                output = BytesIO()
                book.save(output)
                output.seek(0)
                zip_file.writestr(f"""{req.get("filter").get("fnum")}_{v.get("key")}_{req.get("filter").get("y")}.xlsx""" , output.read())
        zip_file.close()


        # df = pd.DataFrame(response["rows"])
        # df = df[df.columns.intersection(list(req["cols"].keys()))]
        # book = load_workbook(f'./doc_templates/{req.get("template")}')
        # sheet = book['dataplan']
        # sheet.append(list(req["cols"].keys()))
        # # Append the new data to the existing sheet
        # for index,row in df.iterrows():
        #     r = []
        #     for c in list(req["cols"].keys()):
        #         r.append(row[c])
        #     sheet.append(r)
        # # Save the modified file to a buffer
        # output = BytesIO()
        # book.save(output)
        # output.seek(0)

        response =  HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response['Content-Disposition'] = 'attachment; filename=registo_diario.zip'
        return response       

    except Exception as error:
        print(str(error))
        return Response({"status": "error", "title": str(error)})

def export(sql, db_parameters, parameters,conn_name,dbi=None,conn=None):
    if ("export" in parameters and parameters["export"] is not None):
        dbparams={}
        for key, value in db_parameters.items():
            if f"%({key})s" not in sql: 
                continue
            dbparams[key] = value
        if parameters["export"] == "clean-excel":
            req = {
                "sql":sql,
                "data":dbparams,
                "cols":parameters["cols"]
            }
            return exportRunxlslist(req,dbi,conn)
        if parameters["export"] == "template-p-xls":
            req = {
                "sql":sql,
                "data":dbparams,
                "cols":parameters["cols"],
                "template":parameters["template"],
                "filter":parameters.get("filter")
            }
            return exportRunxlstemplatelist(req,dbi,conn)
        
        sql = sql.replace(f"%({key})s",f":{key}")        
        hash = base64.b64encode(hmac.new(bytes("SA;PA#Jct\"#f.+%UxT[vf5B)XW`mssr$" , 'utf-8'), msg = bytes(sql , 'utf-8'), digestmod = hashlib.sha256).hexdigest().upper().encode()).decode()
        req = {
            
            "conn-name":conn_name,
            "sql":sql,
            "hash":hash,
            "data":dbparams,
            **parameters
        }
        wService = "runxlslist" if parameters["export"] == "clean-excel" else "runlist"
        fstream = requests.post(f'http://192.168.0.16:8080/ReportsGW/{wService}', json=req)
        
        if (fstream.status_code==200):
            resp =  HttpResponse(fstream.content, content_type=fstream.headers["Content-Type"])
            if (parameters["export"] == "pdf"):
                resp['Content-Disposition'] = "inline; filename=list.pdf"
            elif (parameters["export"] == "excel"):
                resp['Content-Disposition'] = "inline; filename=list.xlsx"
            elif (parameters["export"] == "word"):
                resp['Content-Disposition'] = "inline; filename=list.docx"
            if (parameters["export"] == "csv"):
                resp['Content-Disposition'] = "inline; filename=list.csv"
            return resp